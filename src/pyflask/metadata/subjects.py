from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_SUBJECTS,SCHEMA_NAME_SUBJECTS
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
import numpy as np
from utils import validate_schema, get_sds_headers
from openpyxl.styles import Font
from .helpers import transposeMatrix, getMetadataCustomFields, sortedSubjectsTableData, upload_metadata_file, get_template_path


def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path(SDS_FILE_SUBJECTS)


    subjects = soda["dataset_metadata"]["subjects"]

    validate_schema(subjects, SCHEMA_NAME_SUBJECTS)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_SUBJECTS) if upload_boolean else local_destination
    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]



    # 1. see if the length of datastructure[0] == length of datastructure. If yes, go ahead. If no, add new columns from headers[n-1] onward.
    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )


    # 1.1 (optional) add custom fields to the headers of the workbook
    
    # for column, header in zip(
    #     excel_columns(start_index=29), subjects[0].keys()
    # ):
    #     cell = column + str(1)
    #     ws1[cell] = header
    #     ws1[cell].fill = orangeFill
    #     ws1[cell].font = Font(bold=True, size=12, name="Calibri")

    row = 2
    ascii_headers = excel_columns(start_index=0)
    custom_headers_to_column = {}
    sds_headers = get_sds_headers(SCHEMA_NAME_SUBJECTS)

    # 2. populate excel file with the data 
    for subject in subjects:
        ws1[ascii_headers[0] + str(row)] = subject.get("subject_id", "")
        ws1[ascii_headers[0] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[1] + str(row)] = subject.get("pool_id", "")
        ws1[ascii_headers[1] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[2] + str(row)] = subject.get("subject_experimental_group", "")
        ws1[ascii_headers[2] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[3] + str(row)] = subject.get("age", "")
        ws1[ascii_headers[3] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[4] + str(row)] = subject.get("sex", "")
        ws1[ascii_headers[4] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[5] + str(row)] = subject.get("species", "")
        ws1[ascii_headers[5] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[6] + str(row)] = subject.get("strain", "")
        ws1[ascii_headers[6] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[7] + str(row)] = subject.get("rrid_for_strain", "")
        ws1[ascii_headers[7] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[8] + str(row)] = subject.get("age_category", "")
        ws1[ascii_headers[8] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[9] + str(row)] = subject.get("also_in_dataset", "")
        ws1[ascii_headers[9] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[10] + str(row)] = subject.get("member_of", "")
        ws1[ascii_headers[10] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[11] + str(row)] = subject.get("metadata_only", "")
        ws1[ascii_headers[11] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[12] + str(row)] = subject.get("number_of_directly_derived_samples", "")
        ws1[ascii_headers[12] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[13] + str(row)] = subject.get("laboratory_internal_id", "")
        ws1[ascii_headers[13] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[14] + str(row)] = subject.get("date_of_birth", "")
        ws1[ascii_headers[14] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[15] + str(row)] = subject.get("age_range_min", "")
        ws1[ascii_headers[15] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[16] + str(row)] = subject.get("age_range_max", "")
        ws1[ascii_headers[16] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[17] + str(row)] = normalize_body_mass(subject.get("body_mass", ""))
        ws1[ascii_headers[17] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[18] + str(row)] = subject.get("genotype", "")
        ws1[ascii_headers[18] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[19] + str(row)] = subject.get("phenotype", "")
        ws1[ascii_headers[19] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[20] + str(row)] = subject.get("handedness", "")
        ws1[ascii_headers[20] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[21] + str(row)] = subject.get("reference_atlas", "")
        ws1[ascii_headers[21] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[22] + str(row)] = subject.get("experimental_log_file_path", "")
        ws1[ascii_headers[22] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[23] + str(row)] = subject.get("experiment_date", "")
        ws1[ascii_headers[23] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[24] + str(row)] = subject.get("disease", "")
        ws1[ascii_headers[24] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[25] + str(row)] = subject.get("intervention", "")
        ws1[ascii_headers[25] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[26] + str(row)] = subject.get("disease_model", "")
        ws1[ascii_headers[26] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[27] + str(row)] = subject.get("protocol_title", "")
        ws1[ascii_headers[27] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[28] + str(row)] = subject.get("protocol_url_or_doi", "")
        ws1[ascii_headers[28] + str(row)].font = Font(bold=False, size=11, name="Arial")

        # handle custom fields
        for field_name, field in subject.items():
            if field_name in sds_headers:
                continue

            # check if the field is already in the custom_headers_to_column dictionary
            if field_name not in custom_headers_to_column:
                print(field_name)


                custom_headers_to_column[field_name] = len(custom_headers_to_column.keys()) + 1

                # create the column header in the excel file
                offset_from_final_sds_header = custom_headers_to_column[field_name]
                ws1[ascii_headers[28 + offset_from_final_sds_header] + "1"] = field_name
                ws1[ascii_headers[28 + offset_from_final_sds_header] + "1"].fill = orangeFill
                ws1[ascii_headers[28 + offset_from_final_sds_header] + "1"].font = Font(bold=True, size=12, name="Calibri")



            # add the field value to the corresponding cell in the excel file
            offset_from_final_sds_header = custom_headers_to_column[field_name]

            ws1[ascii_headers[28 + offset_from_final_sds_header] + str(row)] = field
            ws1[ascii_headers[28 + offset_from_final_sds_header] + str(row)].font = Font(bold=False, size=11, name="Arial")


        row += 1

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload_boolean:
        upload_metadata_file(SDS_FILE_SUBJECTS, soda, destination, True)

    return {size: size}



def normalize_body_mass(body_mass_value):
    """
    Ensures that a body_mass value is normalized to a string for Excel export.
    Handles cases where 'body_mass' is a string, number, or an object with 'value' and 'unit'.
    """
    # If already a string or empty, leave as is
    if isinstance(body_mass_value, str):
        return body_mass_value
    # If a number, convert to string
    if isinstance(body_mass_value, (int, float)):
        return str(body_mass_value)
    # If an object with value and unit
    if isinstance(body_mass_value, dict):
            value = body_mass_value.get("value")
            unit = body_mass_value.get("unit")
            if value is not None and unit:
                return f"{value} {unit}"
            elif value is not None:
                return str(value)
            else:
                return ""
    else:
        return ""