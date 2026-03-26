from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_MANIFEST, SCHEMA_NAME_MANIFEST
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import Font, PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema, get_schema_path
from .helpers import upload_metadata_file
from .helpers import get_template_path
import os


from json import load as json_load



def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path("manifest.xlsx")
    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_MANIFEST) if upload_boolean else local_destination
    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]
    manifest = soda["dataset_metadata"]["manifest_file"]
    validate_schema(manifest, SCHEMA_NAME_MANIFEST)
    ascii_headers = excel_columns(start_index=0)
    custom_headers_to_column = {}

    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )

    # Load schema to get standard headers
    schema_path =get_schema_path(SCHEMA_NAME_MANIFEST)
    with open(schema_path, "r") as f:
        schema = json_load(f)
    # The schema is an array, so get the first item's properties
    item_schema = schema["items"][0]
    standard_headers = list(item_schema["properties"].keys())

    # Write standard headers to the first row
    for idx, header in enumerate(standard_headers):
        ws1[ascii_headers[idx] + "1"] = header.replace("_", " ")
        ws1[ascii_headers[idx] + "1"].font = Font(bold=True, size=12, name="Calibri")

    row = 2
    for entry in manifest:
        # Write standard fields
        for idx, header in enumerate(standard_headers):
            value = entry.get(header, "")
            if isinstance(value, list):
                value = " ".join(value)
            ws1[ascii_headers[idx] + str(row)] = value
            ws1[ascii_headers[idx] + str(row)].font = Font(bold=False, size=11, name="Arial")

        # Handle custom fields
        for field_name, field_value in entry.items():
            if field_name in standard_headers:
                continue
            if field_name not in custom_headers_to_column:
                custom_headers_to_column[field_name] = len(custom_headers_to_column) + len(standard_headers)
                col_idx = custom_headers_to_column[field_name]
                ws1[ascii_headers[col_idx] + "1"] = field_name
                ws1[ascii_headers[col_idx] + "1"].fill = orangeFill
                ws1[ascii_headers[col_idx] + "1"].font = Font(bold=True, size=12, name="Calibri")
            col_idx = custom_headers_to_column[field_name]
            ws1[ascii_headers[col_idx] + str(row)] = field_value
            ws1[ascii_headers[col_idx] + str(row)].font = Font(bold=False, size=11, name="Arial")
        row += 1

    # Rename additional metadata header to Additional Metadata header
    # ws1[ascii_headers[len(standard_headers)] + "1"] = "Additional Metadata"

    wb.save(destination)
    size = getsize(destination)
    if upload_boolean:
        upload_metadata_file(SDS_FILE_MANIFEST, soda, destination, True)

    return {"size": size}




def load_existing_manifest_file(manifest_file_path):
   # check that a file exists at the path 
    if not os.path.exists(manifest_file_path):
        raise FileNotFoundError(f"Manifest file not found at {manifest_file_path}")
    
   # load the xlsx file and store its first row as a headers array and the rest of the rows in a data key 
    wb = load_workbook(manifest_file_path)
    ws1 = wb["Sheet1"]
    headers = []
    data = []

    for row in ws1.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    for row in ws1.iter_rows(min_row=2, values_only=True):
        # grab the item in row 5
        new_row = []
        for col_data in row:
            if isinstance(col_data, list):
                # space separate the list into a string
                col_data = " ".join(col_data)
            new_row.append(col_data)
        data.append(new_row)

    return {"headers": headers, "data": data}





