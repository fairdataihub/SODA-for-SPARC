from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_CODE_DESCRIPTION, SCHEMA_NAME_CODE_DESCRIPTION
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill, Font
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema
from .helpers import upload_metadata_file, get_template_path



# TODO: Handle optional entries when coupled with provided entries
# TODO: Handle extending columns and filling with color when more entries are provided than the template default handles
def create_excel(soda, upload, local_destination):
    source = get_template_path(SDS_FILE_CODE_DESCRIPTION)
    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_CODE_DESCRIPTION) if upload else local_destination
    shutil.copyfile(source, destination)

    validate_schema(soda["dataset_metadata"]["code_description"], SCHEMA_NAME_CODE_DESCRIPTION)

    wb = load_workbook("./" +destination)
    print(wb.sheetnames)
    ws1 = wb[wb.sheetnames[0]]

    populate_input_output_information(ws1, soda)

    populate_basic_information(ws1, soda)

    populate_ten_simple_rules(ws1, soda)

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload:
        upload_metadata_file(
            SDS_FILE_CODE_DESCRIPTION, soda, destination, True
        )

    return {"size": size}





# TODO: Handle optional entries
def populate_input_output_information(ws1, soda):
    # populate from row 27 and column 4 up to column n, depending upon the amount of items in the array for each input output information entry
    input_output_information = soda["dataset_metadata"]["input_output_information"]

    row = 27

    excel_ascii = excel_columns(start_index=3)[0]
    ws1[excel_ascii + str(row)] = input_output_information["number_of_inputs"]
    ws1[excel_ascii + str(row)].font = Font(bold=False, size=11, name="Arial")

    for input, column in zip(input_output_information["inputs"], excel_columns(start_index=3)):
        row = 28
        ws1[column + str(row)] = input["input_parameter_name"]
        ws1[column + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 1)] = input["input parameter type"]
        ws1[column + str(row + 1)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 2)] = input["input_parameter_description"]
        ws1[column + str(row + 2)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 3)] = input["input_units"]
        ws1[column + str(row + 3)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 4)] = input["input_default_value"]
        ws1[column + str(row + 4)].font = Font(bold=False, size=11, name="Arial")

    # populate number of outputs into row 34
    row = 34
    ws1[excel_ascii + str(row)] = input_output_information["number_of_outputs"]
    ws1[excel_ascii + str(row)].font = Font(bold=False, size=11, name="Arial")

    # populate the outputs from row 35 - 39
    for output, column in zip(input_output_information["outputs"], excel_columns(start_index=3)):
        row = 35
        ws1[column + str(row)] = output["output_parameter_name"]
        ws1[column + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 1)] = output["output_parameter_type"]
        ws1[column + str(row + 1)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 2)] = output["output_parameter_description"]
        ws1[column + str(row + 2)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 3)] = output["output_units"]
        ws1[column + str(row + 3)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 4)] = output["output_default_value"]
        ws1[column + str(row + 4)].font = Font(bold=False, size=11, name="Arial")


def populate_basic_information(ws1, soda):
    basic_information = soda["dataset_metadata"]["basic_information"]

    # fill out basic information from row 2 - 5 starting from col 3
    row = 2
    for info, column in zip(basic_information, excel_columns(start_index=3)):
        ws1[column + str(row)] = info["RRID_term"]
        ws1[column + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 1)] = info["RRID_identifier"]
        ws1[column + str(row + 1)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 2)] = info["ontology_term"]
        ws1[column + str(row + 2)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[column + str(row + 3)] = info["ontology_identifier"]
        ws1[column + str(row + 3)].font = Font(bold=False, size=11, name="Arial")


def populate_ten_simple_rules(ws1, soda):
    ten_simple_rules = soda["dataset_metadata"]["ten_simple_rules"]
    row = 8
    ascii_cols = excel_columns(start_index=3)
    for _, rule in ten_simple_rules.items():
        ws1[ascii_cols[0] + str(row)] = rule.get("Link", "")
        ws1[ascii_cols[0] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_cols[1] + str(row)] = rule.get("Rating", "")
        ws1[ascii_cols[1] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_cols[2] + str(row)] = rule.get("Target", "")
        ws1[ascii_cols[2] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_cols[3] + str(row)] = rule.get("Target Justification", "")
        ws1[ascii_cols[3] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_cols[4] + str(row)] = rule.get("Text", "")
        ws1[ascii_cols[4] + str(row)].font = Font(bold=False, size=11, name="Arial")
        row += 1






    

