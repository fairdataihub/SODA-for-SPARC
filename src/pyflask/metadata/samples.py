from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_SAMPLES, SCHEMA_NAME_SAMPLES
from .excel_utils import excel_columns
from openpyxl.styles import PatternFill, Font
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema, get_sds_headers
from .helpers import upload_metadata_file, get_template_path


def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path(SDS_FILE_SAMPLES)

    samples = soda["dataset_metadata"]["samples"]

    validate_schema(samples, SCHEMA_NAME_SAMPLES)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_SAMPLES) if upload_boolean else local_destination
    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )

    row = 2
    ascii_headers = excel_columns(start_index=0)
    custom_headers_to_column = {}
    sds_headers = get_sds_headers(SCHEMA_NAME_SAMPLES)

    # Populate the Excel file with the data
    for sample in samples:
        ws1[ascii_headers[0] + str(row)] = sample.get("sample_id", "")
        ws1[ascii_headers[0] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[1] + str(row)] = sample.get("subject_id", "")
        ws1[ascii_headers[1] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[2] + str(row)] = sample.get("was_derived_from", "")
        ws1[ascii_headers[2] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[3] + str(row)] = sample.get("pool_id", "")
        ws1[ascii_headers[3] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[4] + str(row)] = sample.get("sample_experimental_group", "")
        ws1[ascii_headers[4] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[5] + str(row)] = sample.get("sample_type", "")
        ws1[ascii_headers[5] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[6] + str(row)] = handle_anatomical_location_field(sample)
        ws1[ascii_headers[6] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[7] + str(row)] = sample.get("also_in_dataset", "")
        ws1[ascii_headers[7] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[8] + str(row)] = sample.get("member_of", "")
        ws1[ascii_headers[8] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[9] + str(row)] = sample.get("metadata_only", "")
        ws1[ascii_headers[9] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[10] + str(row)] = sample.get("number_of_directly_derived_samples", "")
        ws1[ascii_headers[10] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[11] + str(row)] = sample.get("laboratory_internal_id", "")
        ws1[ascii_headers[11] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[12] + str(row)] = sample.get("date_of_derivation", "")
        ws1[ascii_headers[12] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[13] + str(row)] = sample.get("experimental_log_file_path", "")
        ws1[ascii_headers[13] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[14] + str(row)] = sample.get("reference_atlas", "")
        ws1[ascii_headers[14] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[15] + str(row)] = sample.get("pathology", "")
        ws1[ascii_headers[15] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[16] + str(row)] = sample.get("laterality", "")
        ws1[ascii_headers[16] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[17] + str(row)] = sample.get("cell_type", "")
        ws1[ascii_headers[17] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[18] + str(row)] = sample.get("plane_of_section", "")
        ws1[ascii_headers[18] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[19] + str(row)] = sample.get("protocol_title", "")
        ws1[ascii_headers[19] + str(row)].font = Font(bold=False, size=11, name="Arial")

        ws1[ascii_headers[20] + str(row)] = sample.get("protocol_url_or_doi", "")
        ws1[ascii_headers[20] + str(row)].font = Font(bold=False, size=11, name="Arial")

        # Handle custom fields
        for field_name, field in sample.items():
            if field_name in sds_headers:
                continue

            # Check if the field is already in the custom_headers_to_column dictionary
            if field_name not in custom_headers_to_column:
                custom_headers_to_column[field_name] = len(custom_headers_to_column.keys()) + 1

                # Create the column header in the Excel file
                offset_from_final_sds_header = custom_headers_to_column[field_name]
                ws1[ascii_headers[20 + offset_from_final_sds_header] + "1"] = field_name
                ws1[ascii_headers[20 + offset_from_final_sds_header] + "1"].fill = orangeFill
                ws1[ascii_headers[20 + offset_from_final_sds_header] + "1"].font = Font(bold=True, size=12, name="Calibri")

            # Add the field value to the corresponding cell in the Excel file
            offset_from_final_sds_header = custom_headers_to_column[field_name]
            ws1[ascii_headers[20 + offset_from_final_sds_header] + str(row)] = field
            ws1[ascii_headers[20 + offset_from_final_sds_header] + str(row)].font = Font(bold=False, size=11, name="Arial")

        row += 1

    wb.save(destination)

    size = getsize(destination)

    # If generating directly on Pennsieve, call upload function
    if upload_boolean:
        upload_metadata_file(SDS_FILE_SAMPLES, soda, destination, True)

    return size



def handle_anatomical_location_field(sample):
    anatomical_location = sample.get("sample_anatomical_location", "")
    if isinstance(anatomical_location, list):
        return " ".join(anatomical_location)
    return anatomical_location