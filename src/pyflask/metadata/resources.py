from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_RESOURCES, SCHEMA_NAME_RESOURCES
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill, Font
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema
from .helpers import upload_metadata_file, get_template_path

def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path(SDS_FILE_RESOURCES)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_RESOURCES) if upload_boolean else local_destination

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    resources = soda["dataset_metadata"]["resources"]

    validate_schema(resources, SCHEMA_NAME_RESOURCES)


    # get the ascii column headers
    row = 2
    ascii_headers = excel_columns(start_index=0)
    for resource in resources: 
        ws1[ascii_headers[0] + str(row)] = resource.get("rrid", "")
        ws1[ascii_headers[0] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[1] + str(row)] = resource.get("type", "")
        ws1[ascii_headers[1] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[2] + str(row)] = resource.get("name", "")
        ws1[ascii_headers[2] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[3] + str(row)] = resource.get("url", "")
        ws1[ascii_headers[3] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[4] + str(row)] = resource.get("vendor", "")
        ws1[ascii_headers[4] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[5] + str(row)] = resource.get("version", "")
        ws1[ascii_headers[5] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[6] + str(row)] = resource.get("id_in_protocol", "")
        ws1[ascii_headers[6] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[7] + str(row)] = resource.get("additional_metadata", "")
        ws1[ascii_headers[7] + str(row)].font = Font(bold=False, size=11, name="Arial")

        row += 1

    wb.save(destination)


    size = getsize(destination)


    ## if generating directly on Pennsieve, call upload function
    if upload_boolean:
        upload_metadata_file(SDS_FILE_RESOURCES, soda,  destination, True)

    return {"size": size}



