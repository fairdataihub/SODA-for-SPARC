from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_DATASET_DESCRIPTION, SCHEMA_NAME_DATASET_DESCRIPTION
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from .excel_utils import rename_headers, excel_columns
from copy import copy
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils import validate_schema
from .helpers import upload_metadata_file, get_template_path



def create_excel(
    soda,
    upload_boolean,
    local_destination,
):
    source = get_template_path(SDS_FILE_DATASET_DESCRIPTION)
    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_DATASET_DESCRIPTION) if upload_boolean else local_destination
    shutil.copyfile(source, destination)

    validate_schema(soda["dataset_metadata"]["dataset_description"], SCHEMA_NAME_DATASET_DESCRIPTION)

    # write to excel file
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    ws1["D22"] = ""
    ws1["E22"] = ""
    ws1["D24"] = ""
    ws1["E24"] = ""
    ws1["D25"] = ""
    ws1["E25"] = ""

    # Populate the Metadata version (Required)
    ws1["D2"] = soda["dataset_metadata"]["dataset_description"]["metadata_version"]

    # Populate the Dataset Type (default to empty string if not present)
    ws1["D3"] = (
    soda.get("dataset_metadata", {})
        .get("dataset_description", {})
        .get("dataset_type", "")
    )

    standards_arr_len = populate_standards_info(ws1, soda)

    keyword_funding_array_len = populate_basic_info(ws1, soda)
    study_arr_len = populate_study_info(ws1, soda)
    contributor_arr_len = populate_contributor_info(ws1, soda)
    related_resource_arr_len = populate_related_resource_information(ws1, soda)
    populate_funding_info(ws1, soda)
    populate_participant_information(ws1, soda)
    dict_arr_len = data_dictionary_information(ws1, soda)

    max_len = max(
        keyword_funding_array_len,
        study_arr_len,
        contributor_arr_len,
        related_resource_arr_len,
        standards_arr_len,
        dict_arr_len,
    )

    # 3 is the first value column position
    extend_value_header(ws1, max_len, start_index=3)

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload_boolean:
        upload_metadata_file(
            "dataset_description.xlsx", soda, destination, True
        )

    return {"size": size}




def populate_study_info(workbook, soda):
    study_info = soda["dataset_metadata"]["dataset_description"]["study_information"]
    workbook["D20"] = study_info.get("study_purpose", "")
    workbook["D21"] = study_info.get("study_data_collection", "")
    workbook["D22"] = study_info.get("study_primary_conclusion", "")

    # Arrays
    organ_system = study_info.get("study_organ_system", [])
    approach = study_info.get("study_approach", [])
    technique = study_info.get("study_technique", [])

    for i, column in zip(range(len(organ_system)), excel_columns(start_index=3)):
        workbook[column + "23"] = organ_system[i]
    for i, column in zip(range(len(approach)), excel_columns(start_index=3)):
        workbook[column + "24"] = approach[i]
    for i, column in zip(range(len(technique)), excel_columns(start_index=3)):
        workbook[column + "25"] = technique[i]

    workbook["D26"] = study_info.get("study_collection_title", "")

    # Return the max length of the arrays, or 1 if all are empty
    return max(1, len(organ_system), len(approach), len(technique))

def populate_standards_info(workbook, soda):
    standards_info = soda["dataset_metadata"]["dataset_description"]["standards_information"]
    # this is an array with multiple entries
    for col, standard in zip(excel_columns(start_index=3), standards_info):
        workbook[col + "5"] = standard.get("data_standard", "")
        workbook[col + "6"] = standard.get("data_standard_version", "")

    return max(1, len(standards_info))

def populate_basic_info(workbook, soda):
    basic_info = soda["dataset_metadata"]["dataset_description"]["basic_information"]
    workbook["D8"] = basic_info.get("title", "")
    workbook["D9"] = basic_info.get("subtitle", "")
    workbook["D10"] = basic_info.get("description", "")

    # Write keywords array across columns in row 11 (D11, E11, F11, ...)
    keywords = basic_info.get("keywords", [])
    for col, keyword in zip(excel_columns(start_index=3), keywords):
        workbook[f"{col}11"] = keyword

    funding = basic_info.get("funding", [])
    for col, funding_source in zip(excel_columns(start_index=3), funding):
        workbook[f"{col}12"] = funding_source
    workbook["D13"] = basic_info.get("acknowledgments", "")
    workbook["D14"] = basic_info.get("license", "")

    # Return the length of the keywords array, or 1 if empty
    return max(1, len(keywords), len(funding))


def populate_funding_info(workbook, soda):
    funding_info = soda["dataset_metadata"]["dataset_description"]["funding_information"]
    workbook["D16"] = funding_info["funding_consortium"]
    workbook["D17"] = funding_info["funding_agency"]
    workbook["D18"] = funding_info["award_number"]



def populate_contributor_info(workbook, soda):
    contributor_info = soda["dataset_metadata"]["dataset_description"].get("contributor_information", [])
    for contributor, column in zip(contributor_info, excel_columns(start_index=3)):
        workbook[column + "28"] = contributor.get("contributor_name", "")
        workbook[column + "29"] = contributor.get("contributor_orcid_id", "")
        workbook[column + "30"] = contributor.get("contributor_affiliation", "")
        roles = contributor.get("contributor_roles", [])
        roles_str = ", ".join(roles) if roles else ""
        workbook[column + "31"] = roles_str
    # Return the length of the contributor array, or 1 if empty
    return max(1, len(contributor_info))


def populate_related_resource_information(workbook, soda):
    related_resource_information = soda["dataset_metadata"]["dataset_description"].get("related_resource_information", [])
    for info, column in zip(related_resource_information, excel_columns(start_index=3)):
        workbook[column + "33"] = info.get("identifier_description", "")
        workbook[column + "34"] = info.get("relation_type", "")
        workbook[column + "35"] = info.get("identifier", "")
        workbook[column + "36"] = info.get("identifier_type", "")
    # Return the length of the related resource array, or 1 if empty
    return max(1, len(related_resource_information))



def populate_participant_information(workbook, soda):
    participant_info = soda["dataset_metadata"]["dataset_description"]["participant_information"]
    workbook["D38"] = participant_info.get("number_of_subjects", 0)
    workbook["D39"] = participant_info.get("number_of_samples", 0)
    workbook["D40"] = participant_info.get("number_of_sites", 0)
    workbook["D41"] = participant_info.get("number_of_performances", 0)


def data_dictionary_information(workbook, soda):
    """
    This function is a placeholder for future implementation.
    It currently does not populate any data in the workbook.
    """
    # Placeholder for future implementation
    data_dictionary_info = soda["dataset_metadata"]["dataset_description"].get("data_dictionary_information", [])

    for column, entry in zip(excel_columns(start_index=3), data_dictionary_info):
        workbook[column + "43"] = entry.get("data_dictionary_path", "")
        workbook[column + "44"] = entry.get("data_dictionary_type", "")
        workbook[column + "45"] = entry.get("data_dictionary_description", "")

    return max(1, len(data_dictionary_info))

def grayout_subheaders(workbook, col):
    """
    Gray out the cells at workbook[row][col] for the specified cells in 
    positions 2, 3, 4, 7, 8, 9, 10, 13, 14, 15, 19, 20, 21, 22, 26, 27, 32, 37, 
    38, 39, 40, 41, 42
    """
    gray_out_rows_for_column = [2, 3, 4, 7, 8, 9, 10, 13, 14, 15, 19, 20, 21, 22, 26, 27, 32, 37, 
    38, 39, 40, 41, 42]

    for row in gray_out_rows_for_column:    
        cell = workbook[col + str(row)]
        if row in [4, 7, 15, 19, 27, 32, 37, 42]:
            fillColor("b2b2b2", cell)
        else:
            fillColor("cccccc", cell)


def fillColor(color, cell):
    colorFill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    cell.fill = colorFill


def apply_calibri_bold_12(cell):
    """Apply Calibri Bold 12pt font formatting to a cell"""
    calibri_bold_font = Font(name='Calibri', size=12, bold=True)
    cell.font = calibri_bold_font


def set_cell_alignment(cell, horizontal='left', vertical='top', wrap_text=False):
    """Set text alignment for a cell
    
    Args:
        cell: The cell to format
        horizontal: 'left', 'center', 'right', 'justify', 'distributed'
        vertical: 'top', 'center', 'bottom', 'justify', 'distributed'
        wrap_text: Boolean to enable text wrapping
    """
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text
    )


def apply_dashed_border(cell, workbook):
    """Apply border copied from cell A1 to the target cell"""
    # Copy the border from cell A1
    cell.border = copy(workbook["A1"].border)
    


def extend_value_header(workbook, max_len, start_index):
    """
    The headers starting at G1 are the 'Value' headers that correspond to the maximum number of entries for either the 
    keywords, contributor information, or data dictionary information arrays. This function extends those headers based on the max_len of the 
    three arrays.
    """

    column_list = excel_columns(start_index=start_index)

    # if max len is less than 4 then no need to extend headers
    if max_len < 4:
        return
    
    # replace the 4th value with header value 'Value 4' and so on for max len
    for i in range(4, max_len + 1):
        header_cell = workbook[column_list[i - 1] + "1"]
        header_cell.value = f"Value {i}"
        # make the new header blue
        fillColor("9cc2e5", header_cell)
        apply_calibri_bold_12(header_cell)
        set_cell_alignment(header_cell, horizontal='center', vertical='center')
        apply_dashed_border(header_cell, workbook)
        grayout_subheaders(workbook, column_list[i - 1])
