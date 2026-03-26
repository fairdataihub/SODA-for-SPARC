from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_SITES, SCHEMA_NAME_SITES
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill, Font
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema
from .helpers import upload_metadata_file, get_template_path

def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path(SDS_FILE_SITES)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_SITES) if upload_boolean else local_destination

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    sites = soda["dataset_metadata"]["sites"]

    validate_schema(sites, SCHEMA_NAME_SITES)


    # get the ascii column headers
    row = 2
    ascii_headers = excel_columns(start_index=0)
    for performance in sites: 
        ws1[ascii_headers[0] + str(row)] = performance.get("site_id", "")
        ws1[ascii_headers[0] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[1] + str(row)] = performance.get("specimen_id", "")
        ws1[ascii_headers[1] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[2] + str(row)] = performance.get("site_type", "")
        ws1[ascii_headers[2] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[3] + str(row)] = performance.get("laboratory_internal_id", "")
        ws1[ascii_headers[3] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[4] + str(row)] = performance.get("coordinate_system", "")
        ws1[ascii_headers[4] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[5] + str(row)] = performance.get("coordinate_system_position", "")
        ws1[ascii_headers[5] + str(row)].font = Font(bold=False, size=11, name="Arial")
        
        ws1[ascii_headers[6] + str(row)] = performance.get("more...", "")
        ws1[ascii_headers[6] + str(row)].font = Font(bold=False, size=11, name="Arial")
        row += 1

    wb.save(destination)

    size = getsize(destination)


    ## if generating directly on Pennsieve, call upload function
    if upload_boolean:
        upload_metadata_file(SDS_FILE_SITES, soda,  destination, True)

    return {"size": size}
    



