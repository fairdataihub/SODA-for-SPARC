### Import required python modules
from flask import abort
import platform
import os
import itertools
from os import makedirs, mkdir, walk
from openpyxl import load_workbook
from os.path import (
    isdir,
    join,
    splitext,
    basename,
    exists,
    expanduser,
    getsize,
)
import pandas as pd
import time
import shutil
import subprocess
import re
from timeit import default_timer as timer
from datetime import timedelta
import gevent
import pathlib
from datetime import datetime, timezone
import requests 
from permissions import pennsieve_get_current_user_permissions
from utils import get_dataset_id, create_request_headers
from namespaces import NamespaceEnum, get_namespace_logger
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from utils import load_metadata_to_dataframe
from constants import PENNSIEVE_URL

namespace_logger = get_namespace_logger(NamespaceEnum.ORGANIZE_DATASETS)
from authentication import get_access_token



### Global variables
create_soda_json_progress = 0
create_soda_json_total_items = 0
create_soda_json_completed = 0
curateprogress = " "
curatestatus = " "
curateprintstatus = " "
total_dataset_size = 1
curated_dataset_size = 0
start_time = 0

userpath = expanduser("~")
configpath = join(userpath, ".pennsieve", "config.ini")
submitdataprogress = " "
submitdatastatus = " "
submitprintstatus = " "
total_file_size = 1
uploaded_file_size = 0
start_time_bf_upload = 0
start_submit = 0
metadatapath = join(userpath, "SODA", "SODA_metadata")

myds = ""
initial_bfdataset_size = 0
upload_directly_to_ps = 0
initial_bfdataset_size_submit = 0

forbidden_characters = '<>:"/\|?*'
forbidden_characters_bf = '\/:*?"<>'


from namespaces import NamespaceEnum, get_namespace_logger
namespace_logger = get_namespace_logger(NamespaceEnum.MANAGE_DATASETS)

METADATA_FILES_SPARC = [
        "submission.xlsx",
        "submission.csv",
        "submission.json",
        "dataset_description.xlsx",
        "dataset_description.csv",
        "dataset_description.json",
        "subjects.xlsx",
        "subjects.csv",
        "subjects.json",
        "samples.xlsx",
        "samples.csv",
        "samples.json",
        "README.txt",
        "CHANGES.txt",
        "code_description.xlsx",
        "inputs_metadata.xlsx",
        "outputs_metadata.xlsx",
    ]

### Internal functions
def TZLOCAL():
    """
        Get current local timezone
    """
    return datetime.now(timezone.utc).astimezone().tzinfo

## these subsequent CheckLeafValue and traverseForLeafNodes functions check for the validity of file paths,
## and folder and file size
def checkLeafValue(leafName, leafNodeValue):
    error, c = "", 0
    total_dataset_size = 1
    curatestatus = ""

    if isinstance(leafNodeValue, list):
        filePath = leafNodeValue[0]

        if exists(filePath):
            filePathSize = getsize(filePath)
            if filePathSize == 0:
                c += 1
                error = error + filePath + " is 0 KB <br>"
            else:
                total_dataset_size += filePathSize
        else:
            c += 1
            error = error + filePath + " doesn not exist! <br>"

    elif isinstance(leafNodeValue, dict):
        c += 1
        error = error + leafName + " is empty <br>"

    if c > 0:
        error += "<br>Please remove invalid files/folders from your dataset and try again"
        curatestatus = "Done"
        abort(400, error)


    return [True, total_dataset_size - 1]


def traverseForLeafNodes(datasetStructure):
    total_dataset_size = 1

    for key in datasetStructure:
        if isinstance(datasetStructure[key], list):

            returnedOutput = checkLeafValue(key, datasetStructure[key])

            if returnedOutput[0]:
                total_dataset_size += returnedOutput[1]

        else:

            if len(datasetStructure[key]) == 0:
                returnedOutput = checkLeafValue(key, datasetStructure[key])

            else:
                # going one step down in the object tree
                traverseForLeafNodes(datasetStructure[key])

    return total_dataset_size


######## CREATE FILES FOR CURATE_DATASET FUNCTION
def createFiles(jsonpath, fileKey, distdir, listallfiles):
    # fileKey is the key in json structure that is a file (meaning that its value is an array)
    # note: this fileKey can be the new name of the file (if renamed by users in SODA)
    # (or can be the original basename)

    srcfile = jsonpath[fileKey][0]
    distfile = distdir
    listallfiles.append([srcfile, distfile])


def ignore_empty_high_level_folders(jsonObject):
    items_to_delete = [
        folder for folder in jsonObject.keys() if len(jsonObject[folder].keys()) == 0
    ]

    for item in items_to_delete:
        del jsonObject[item]

    return jsonObject


def generate_dataset_locally(destinationdataset, pathdataset, newdatasetname, jsonpath):
    """
    Associated with 'Generate' button in the 'Generate dataset' section of SODA interface
    Checks validity of files / paths / folders and then generates the files and folders
    as requested along with progress status

    Args:
        destinationdataset: type of destination dataset ('modify existing', 'create new', or 'upload to Pennsieve')
        pathdataset: destination path of new dataset if created locally or name of Pennsieve account (string)
        newdatasetname: name of the local dataset or name of the dataset on Pennsieve (string)
        manifeststatus: boolean to check if user request manifest files
        jsonpath: path of the files to be included in the dataset (dictionary)
    """
    global curatestatus  # set to 'Done' when completed or error to stop progress tracking from front-end
    global curateprogress  # GUI messages shown to user to provide update on progress
    global curateprintstatus  # If = "Curating" Progress messages are shown to user
    global total_dataset_size  # total size of the dataset to be generated
    global curated_dataset_size  # total size of the dataset generated (locally or on Pennsieve) at a given time
    global start_time
    global myds
    global upload_directly_to_ps
    global start_submit
    global initial_bfdataset_size

    curateprogress = " "
    curatestatus = ""
    curateprintstatus = " "
    error, c = "", 0
    curated_dataset_size = 0
    start_time = 0
    upload_directly_to_ps = 0
    start_submit = 0
    initial_bfdataset_size = 0

    jsonstructure_non_empty = ignore_empty_high_level_folders(jsonpath)

    if destinationdataset == "create new":
        if not isdir(pathdataset):
            curatestatus = "Done"
            abort(400, "Please select a valid folder for new dataset")
        if not newdatasetname:
            curatestatus = "Done"
            abort(400, "Please enter a valid name for new dataset folder")
        if check_forbidden_characters(newdatasetname):
            curatestatus = "Done"
            abort(400,  "A folder name cannot contain any of the following characters " + forbidden_characters)

    total_dataset_size = 1

    # check if path in jsonpath are valid and calculate total dataset size
    total_dataset_size = traverseForLeafNodes(jsonstructure_non_empty)
    total_dataset_size = total_dataset_size - 1

    # CREATE NEW
    if destinationdataset == "create new":
        try:
            listallfiles = []
            pathnewdatasetfolder = join(pathdataset, newdatasetname)
            pathnewdatasetfolder = return_new_path(pathnewdatasetfolder)
            open_file(pathnewdatasetfolder)

            curateprogress = "Started"
            curateprintstatus = "Curating"
            start_time = time.time()
            start_submit = 1

            pathdataset = pathnewdatasetfolder
            mkdir(pathdataset)
            create_dataset(pathdataset, jsonstructure_non_empty, listallfiles)

            curateprogress = "New dataset created"
            curateprogress = "Success: COMPLETED!"
            curatestatus = "Done"

        except Exception as e:
            curatestatus = "Done"
            raise e


def create_folder_level_manifest(jsonpath, jsondescription):
    """
    Function to create manifest files for each SPARC folder.
    Files are created in a temporary folder

    Args:
        datasetpath: path of the dataset (string)
        jsonpath: all paths in json format with key being SPARC folder names (dictionary)
        jsondescription: description associated with each path (dictionary)
    Action:
        Creates manifest files in xslx format for each SPARC folder
    """
    global total_dataset_size
    local_timezone = TZLOCAL()

    try:
        datasetpath = metadatapath
        shutil.rmtree(datasetpath) if isdir(datasetpath) else 0
        makedirs(datasetpath)
        folders = list(jsonpath.keys())
        # In each SPARC folder, generate a manifest file
        for folder in folders:
            if jsonpath[folder] != []:
                # Initialize dataframe where manifest info will be stored
                df = pd.DataFrame(
                    columns=[
                        "filename",
                        "timestamp",
                        "description",
                        "file type",
                        "Additional Metadata",
                    ]
                )
                # Get list of files/folders in the the folder
                # Remove manifest file from the list if already exists
                folderpath = join(datasetpath, folder)
                allfiles = jsonpath[folder]
                alldescription = jsondescription[folder + "_description"]

                countpath = -1
                for pathname in allfiles:
                    countpath += 1
                    if (
                        basename(pathname) == "manifest.csv"
                        or basename(pathname) == "manifest.xlsx"
                    ):
                        allfiles.pop(countpath)
                        alldescription.pop(countpath)

                # Populate manifest dataframe
                filename, timestamp, filetype, filedescription = [], [], [], []
                countpath = -1
                for paths in allfiles:
                    if isdir(paths):
                        key = basename(paths)
                        alldescription.pop(0)
                        for subdir, dirs, files in os.walk(paths):
                            for file in files:
                                gevent.sleep(0)
                                filepath = pathlib.Path(paths) / subdir / file
                                mtime = filepath.stat().st_mtime
                                lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                                    local_timezone
                                )
                                timestamp.append(
                                    lastmodtime.isoformat()
                                    .replace(".", ",")
                                    .replace("+00:00", "Z")
                                )
                                full_filename = filepath.name

                                if folder == "main":  # if file in main folder
                                    filename.append(
                                        full_filename
                                    ) if folder == "" else filename.append(
                                        join(folder, full_filename)
                                    )
                                else:
                                    subdirname = os.path.relpath(
                                        subdir, paths
                                    )  # gives relative path of the directory of the file w.r.t paths
                                    if subdirname == ".":
                                        filename.append(join(key, full_filename))
                                    else:
                                        filename.append(
                                            join(key, subdirname, full_filename)
                                        )

                                fileextension = splitext(full_filename)[1]
                                if (
                                    not fileextension
                                ):  # if empty (happens e.g. with Readme files)
                                    fileextension = "None"
                                filetype.append(fileextension)
                                filedescription.append("")
                    else:
                        gevent.sleep(0)
                        countpath += 1
                        filepath = pathlib.Path(paths)
                        file = filepath.name
                        filename.append(file)
                        mtime = filepath.stat().st_mtime
                        lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                            local_timezone
                        )
                        timestamp.append(
                            lastmodtime.isoformat()
                            .replace(".", ",")
                            .replace("+00:00", "Z")
                        )
                        filedescription.append(alldescription[countpath])
                        if isdir(paths):
                            filetype.append("folder")
                        else:
                            fileextension = splitext(file)[1]
                            if (
                                not fileextension
                            ):  # if empty (happens e.g. with Readme files)
                                fileextension = "None"
                            filetype.append(fileextension)

                df["filename"] = filename
                df["timestamp"] = timestamp
                df["file type"] = filetype
                df["description"] = filedescription

                makedirs(folderpath)
                # Save manifest as Excel sheet
                manifestfile = join(folderpath, "manifest.xlsx")
                df.to_excel(manifestfile, index=None, header=True)
                wb = load_workbook(manifestfile)
                ws = wb.active
                blueFill = PatternFill(
                    start_color="9DC3E6", fill_type="solid"
                )
                greenFill = PatternFill(
                    start_color="A8D08D", fill_type="solid"
                )
                yellowFill = PatternFill(
                    start_color="FFD965", fill_type="solid"
                )
                ws['A1'].fill = blueFill
                ws['B1'].fill = greenFill
                ws['C1'].fill = greenFill
                ws['D1'].fill = greenFill
                ws['E1'].fill = yellowFill
                wb.save(manifestfile)
                total_dataset_size += path_size(manifestfile)
                jsonpath[folder].append(manifestfile)

        return jsonpath

    except Exception as e:
        raise e


def check_forbidden_characters(my_string):
    """
    Check for forbidden characters in file/folder name

    Args:
        my_string: string with characters (string)
    Returns:
        False: no forbidden character
        True: presence of forbidden character(s)
    """
    regex = re.compile("[" + forbidden_characters + "]")
    if regex.search(my_string) == None and "\\" not in r"%r" % my_string:
        return False
    else:
        return True


def folder_size(path):
    """
    Provides the size of the folder indicated by path

    Args:
        path: path of the folder (string)
    Returns:
        total_size: total size of the folder in bytes (integer)
    """
    total_size = 0

    for path, _, files in walk(path):
        for f in files:
            fp = join(path, f)
            total_size += getsize(fp)
    return total_size


def open_file(file_path):
    """
    Opening folder on all platforms
    https://stackoverflow.com/questions/6631299/python-opening-a-folder-in-explorer-nautilus-mac-thingie

    Args:
        file_path: path of the folder (string)
    Action:
        Opens file explorer window to the given path
    """
    try:
        if platform.system() == "Windows":
            subprocess.Popen(r"explorer /select," + str(file_path))
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])
    except Exception as e:
        raise e



def path_size(path):
    """
    Returns size of the path, after checking if it's a folder or a file
    Args:
        path: path of the file/folder (string)
    Returns:
        total_size: total size of the file/folder in bytes (integer)
    """
    if isdir(path):
        return folder_size(path)
    else:
        return getsize(path)


def mycopyfile_with_metadata(src, dst, *, follow_symlinks=True):
    """
    Copy file src to dst with metadata (timestamp, permission, etc.) conserved

    Args:
        src: source file (string)
        dst: destination file (string)
    Returns:
        dst
    """
    if not follow_symlinks and os.path.islink(src):
        os.symlink(os.readlink(src), dst)
    else:
        with open(src, "rb") as fsrc:
            with open(dst, "wb") as fdst:
                mycopyfileobj(fsrc, fdst)
    shutil.copystat(src, dst)
    return dst


def mycopyfileobj(fsrc, fdst, length=16 * 1024 * 16):
    """
    Helper function to copy file

    Args:
        fsrc: source file opened in python (file-like object)
        fdst: destination file accessed in python (file-like object)
        length: copied buffer size in bytes (integer)
    """
    global curateprogress
    global total_dataset_size
    global curated_dataset_size
    while True:
        buf = fsrc.read(length)
        if not buf:
            break
        gevent.sleep(0)
        fdst.write(buf)
        curated_dataset_size += len(buf)


def return_new_path(topath):
    """
    This function checks if a folder already exists and in such cases,
    appends (2) or (3) etc. to the folder name

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """
    if exists(topath):
        i = 2
        while True:
            if not exists(topath + " (" + str(i) + ")"):
                return topath + " (" + str(i) + ")"
            i += 1
    else:
        return topath


def create_dataset(recursivePath, jsonStructure, listallfiles):
    """
    Associated with 'Create new dataset locally' option of SODA interface
    for creating requested folders and files to the destination path specified

    Args:
        recursivePath: destination path for creating new folders and files (recursively) (string)
        jsonStructure: all paths (dictionary)
    Action:
        Creates the folders and files specified
    """

    global curateprogress

    for key in jsonStructure:

        if isinstance(jsonStructure[key], dict):

            if not exists(join(recursivePath, key)):
                mkdir(join(recursivePath, key))

            outputpath = join(recursivePath, key)

            create_dataset(outputpath, jsonStructure[key], listallfiles)

        elif isinstance(jsonStructure[key], list):

            outputpath = join(recursivePath, key)

            createFiles(jsonStructure, key, outputpath, listallfiles)

            for fileinfo in listallfiles:
                srcfile = fileinfo[0]
                distfile = fileinfo[1]
                curateprogress = "Copying " + str(srcfile)

                mycopyfile_with_metadata(srcfile, distfile)

def create_soda_json_object_backend(
    soda_json_structure, root_folder_path, irregularFolders, replaced
):
    """
    This function is meant for importing local datasets into SODA.
    It creates a json object with the structure of the dataset.
    """
    global create_soda_json_progress  # amount of items counted during recursion
    global create_soda_json_total_items  # counts the total items in folder
    global create_soda_json_completed  # completed progress is either 0 or 1
    global METADATA_FILES_SPARC

    high_level_sparc_folders = [
        "code",
        "derivative",
        "docs",
        "primary",
        "protocol",
        "source",
    ]

    dataset_folder = soda_json_structure["dataset-structure"] = {"folders": {}}

    def recursive_structure_create(dataset_structure, folder_path, high_lvl_folder_name):
        global create_soda_json_progress
        # going within high level folders
        # add manifest details if manifest exists
        manifest_object = {
            "filename": "",
            "timestamp": "",
            "description": "",
            "additional-metadata": "",
        }

        lastSlash = folder_path.rfind("/") + 1
        folder_name = folder_path[lastSlash:]

        if folder_name in replaced.keys():
            folder_name = replaced[folder_name]

        # Check if folder is in irregular folders
        if folder_path in irregularFolders:
            index_check = irregularFolders.index(folder_path)
            modified_name = replaced[os.path.basename(folder_path)]
            folder_path = irregularFolders[index_check]
        entries = os.listdir(folder_path)
        for entry in entries:
            gevent.sleep(0)
            item_path = os.path.normpath(os.path.join(folder_path, entry))

            if os.path.isfile(item_path) is True:
                # check manifest to add metadata
                if entry[0:1] != ".":
                    create_soda_json_progress += 1
                if entry[0:1] != "." and entry[0:8] != "manifest":
                    # no hidden files or manifest files included
                    if high_lvl_folder_name in soda_json_structure["starting-point"]:
                        if (
                            soda_json_structure["starting-point"][high_lvl_folder_name]["path"]
                            != ""
                        ):
                            # checks if there is a path to a manifest
                            manifest_path = soda_json_structure["starting-point"][
                                high_lvl_folder_name
                            ]["path"]
                            ext_index = manifest_path.rfind(".")
                            extension = manifest_path[ext_index:]
                            if extension == ".xlsx":
                                for key in soda_json_structure["starting-point"][
                                    high_lvl_folder_name
                                ]["manifest"]:
                                    extra_columns = False
                                    if(len(key) > 5):
                                        extra_columns = True
                                        extra_columns_dict = dict(itertools.islice(key.items(), 5, len(key)))
                                    # description metadata
                                    if key["filename"] in item_path:
                                        if key["description"] != "":
                                            manifest_object["description"] = key[
                                                "description"
                                            ]
                                        else:
                                            manifest_object["description"] = ""
                                    # additional metadata
                                        if key["Additional Metadata"] != "":
                                            manifest_object["additional-metadata"] = key[
                                                "Additional Metadata"
                                            ]
                                        else:
                                            manifest_object["additional-metadata"] = ""

                                        if extra_columns:
                                            manifest_object["extra_columns"] = extra_columns_dict

                            elif extension == ".csv":
                                for key in soda_json_structure["starting-point"][
                                    folder_name
                                ]["manifest"]:
                                    extra_columns = False
                                    if len(key) > 5:
                                        extra_columns = True
                                        extra_columns_dict = dict(itertools.islice(key.items(), 5, len(key)))

                                    if (
                                        soda_json_structure["starting-point"][
                                            folder_name
                                        ]["manifest"][key]["filename"]
                                        == entry
                                    ):
                                        if (
                                            soda_json_structure["starting-point"][
                                                folder_name
                                            ][key]["description"]
                                            != None
                                        ):
                                            manifest_object[
                                                "description"
                                            ] = soda_json_structure["starting-point"][
                                                folder_name
                                            ][
                                                key
                                            ][
                                                "description"
                                            ]
                                        else:
                                            manifest_object["description"] = ""
                                    if (
                                        soda_json_structure["starting-point"][
                                            folder_name
                                        ]["manifest"][key]["Additional Metadata"]
                                        != None
                                    ):
                                        manifest_object[
                                            "additional-metadata"
                                        ] = soda_json_structure["starting-point"][
                                            folder_name
                                        ][
                                            "manifest"
                                        ][
                                            key
                                        ][
                                            "Additional Metadata"
                                        ]
                                    else:
                                        manifest_object["additional-metadata"] = ""

                                    if extra_columns:
                                        manifest_object["extra_columns"] = extra_columns_dict
                    # create json
                    if "extra_columns" in manifest_object:
                        dataset_structure["files"][entry] = {
                            "path": item_path,
                            "location": "local",
                            "action": ["existing"],
                            "description": manifest_object["description"],
                            "additional-metadata": manifest_object["additional-metadata"],
                            "extra_columns": manifest_object["extra_columns"]
                        }
                    else:
                        dataset_structure["files"][entry] = {
                            "path": item_path,
                            "location": "local",
                            "action": ["existing"],
                            "description": manifest_object["description"],
                            "additional-metadata": manifest_object["additional-metadata"],
                        }
            elif os.path.isdir(item_path) is True:
                create_soda_json_progress += 1
                if item_path in irregularFolders:
                    index_check = irregularFolders.index(item_path)
                    modified_name = replaced[os.path.basename(item_path)]

                    dataset_structure["folders"][modified_name] = {
                        "folders": {},
                        "files": {},
                        "path": item_path,
                        "location": "local",
                        "action": ["existing"],
                        "original-name": entry,
                    }
                    for folder in dataset_structure["folders"][modified_name][
                        "folders"
                    ]:
                        updated_path = dataset_structure["folders"][modified_name][
                            folder
                        ]["path"]
                        recursive_structure_create(
                            dataset_structure["folders"][modified_name][folder],
                            updated_path,
                            high_lvl_folder_name
                        )
                else:
                    dataset_structure["folders"][entry] = {
                        "folders": {},
                        "files": {},
                        "path": item_path,
                        "location": "local",
                        "action": ["existing"],
                    }

        for folder in dataset_structure["folders"]:
            updated_path = dataset_structure["folders"][folder]["path"]
            recursive_structure_create(
                dataset_structure["folders"][folder], updated_path, high_lvl_folder_name
            )

    # BEGIN

    # count the amount of items in folder
    create_soda_json_total_items = 0
    for root, dirs, filenames in os.walk(root_folder_path):
        # walk through all folders and it's subfolders
        for Dir in dirs:
            # does not take hidden folders or manifest folders
            if Dir[0:1] != "." and Dir[0:8] != "manifest":
                create_soda_json_total_items += 1
        for fileName in filenames:
            if root == root_folder_path and fileName in METADATA_FILES_SPARC:
                # goes through all files and does not count hidden files
                create_soda_json_total_items += 1
            else:
                if fileName[0:1] != ".":
                    create_soda_json_total_items += 1

    # reading high level folders
    create_soda_json_completed = 0
    create_soda_json_progress = 0
    entries = os.listdir(root_folder_path)


    for entry in entries:
        # begin going through high level folders
        gevent.sleep(0)
        item_path = root_folder_path + "/" + entry
        # high level folder paths
        if os.path.isfile(item_path) is True:
            if entry[0:1] != "." and entry in METADATA_FILES_SPARC:
                # is not a hidden folder
                create_soda_json_progress += 1
                soda_json_structure["metadata-files"][entry] = {
                    "path": item_path,
                    "location": "local",
                    "action": ["existing"],
                }
            # do file work here
        elif os.path.isdir(item_path) is True:
            create_soda_json_progress += 1
            # add item to soda
            if item_path in irregularFolders:
                index_check = irregularFolders.index(item_path)
                modified_name = replaced[index_check]
                folder_name = modified_name
                dataset_folder["folders"][folder_name] = {
                    "folders": {},
                    "files": {},
                    "path": item_path,
                    "location": "local",
                    "action": ["existing"],
                    "original-basename": item_path[(item_path.rfind("/") + 1) :],
                }
            else:
                if entry in high_level_sparc_folders:
                    dataset_folder["folders"][entry] = {
                        "folders": {},
                        "files": {},
                        "path": item_path,
                        "location": "local",
                        "action": ["existing"],
                    }
            soda_json_structure["starting-point"][entry] = {"path": ""}

    for folder in dataset_folder["folders"]:
        # go through high level folders again
        high_lvl_path = root_folder_path + "/" + folder
        temp_csv = high_lvl_path + "/manifest.csv"
        temp_xlsx = high_lvl_path + "/manifest.xlsx"
        if os.path.exists(temp_csv) == True:
            temp_file_path_csv = root_folder_path + "/" + folder + "/" + "manifest.csv"
            csv_data = pd.read_csv(temp_file_path_csv)
            csv_data.fillna("", inplace=True)
            json_format = csv_data.to_dict(orient="records")
            soda_json_structure["starting-point"][folder]["path"] = temp_file_path_csv
            soda_json_structure["starting-point"][folder]["manifest"] = json_format
        if os.path.exists(temp_xlsx) == True:
            temp_file_path_xlsx = (
                root_folder_path + "/" + folder + "/" + "manifest.xlsx"
            )
            excel_data = pd.read_excel(temp_file_path_xlsx, sheet_name="Sheet1")
            excel_data.fillna("", inplace=True)
            json_format = excel_data.to_dict(orient="records")
            soda_json_structure["starting-point"][folder]["path"] = temp_file_path_xlsx
            soda_json_structure["starting-point"][folder]["manifest"] = json_format
        recursive_structure_create(dataset_folder["folders"][folder], high_lvl_path, folder)

    create_soda_json_completed = 1
    return soda_json_structure


def monitor_local_json_progress():
    """
    Function for monitoring progress of json_object_creation
    Used for progress bar
    """
    global create_soda_json_completed
    global create_soda_json_total_items
    global create_soda_json_progress
    progress_percentage = (
        create_soda_json_progress / create_soda_json_total_items
    ) * 100

    return {
        "create_soda_json_progress": create_soda_json_progress,
        "create_soda_json_total_items": create_soda_json_total_items,
        "progress_percentage": progress_percentage,
        "create_soda_json_completed": create_soda_json_completed
    }





def monitor_pennsieve_json_progress():
    """
    Function for monitoring progress of soda_json object
    Used for progress bar
    """
    global create_soda_json_completed
    global create_soda_json_total_items
    global create_soda_json_progress

    if create_soda_json_progress != 0:
        progress_percentage = (
            create_soda_json_progress / create_soda_json_total_items
        ) * 100
    else:
        progress_percentage = 0
    return {
        "import_progress": create_soda_json_progress,
        "import_total_items": create_soda_json_total_items,
        "import_progress_percentage": progress_percentage,
        "import_completed_items": create_soda_json_completed,
    }


