import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Testing Checklist"

# Define styles
header_fill = PatternFill(start_color="549cdf", end_color="549cdf", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
section_fill = PatternFill(start_color="549cdf", end_color="549cdf", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=11)
subsection_fill = PatternFill(start_color="A3CEF1", end_color="A3CEF1", fill_type="solid")
subsection_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 80
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 40

# Create data validation for checkboxes
dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=False)
dv.prompt = "Select checkbox status"
dv.promptTitle = "Done Status"
ws.add_data_validation(dv)

# Add headers
headers = ["Test Description", "Done", "Notes"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row = 2

# Testing Setup Section
def add_section(title, row_num):
    ws.merge_cells(f'A{row_num}:C{row_num}')
    cell = ws.cell(row=row_num, column=1)
    cell.value = title
    cell.fill = section_fill
    cell.font = section_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 4):
        ws.cell(row=row_num, column=col).border = border
    return row_num + 1

def add_subsection(title, row_num):
    ws.merge_cells(f'A{row_num}:C{row_num}')
    cell = ws.cell(row=row_num, column=1)
    cell.value = title
    cell.fill = subsection_fill
    cell.font = subsection_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 4):
        ws.cell(row=row_num, column=col).border = border
    return row_num + 1

def add_test_item(description, row_num):
    ws.cell(row=row_num, column=1).value = description
    # Add checkbox with dropdown validation
    done_cell = ws.cell(row=row_num, column=2)
    done_cell.value = "☐"
    done_cell.font = Font(name='Segoe UI Symbol', size=14)
    dv.add(done_cell)  # Add dropdown validation to this cell
    
    ws.cell(row=row_num, column=3).value = ""
    
    for col in range(1, 4):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='top')
    
    return row_num + 1

# Testing Setup
row = add_section("TESTING SETUP", row)
row = add_test_item("Jacob: Windows OS (Latest Version)", row)
row = add_test_item("Aaron: Mac OS (Latest Version)", row)
row = add_test_item("Light workflow for Ubuntu (Latest Version)", row)
row = add_test_item("NOTE: Make sure .DNT is enabled so analytics numbers don't get messy", row)

# Workflow Testing + Pennsieve Integration
row += 1
row = add_section("WORKFLOW TESTING + PENNSIEVE INTEGRATION LOGIC TESTING", row)

row = add_subsection("Initial Setup", row)
row = add_test_item("Create a new user profile on the OS", row)
row = add_test_item("Install the test build of SODA", row)
row = add_test_item("Do not connect to Pennsieve or install the Pennsieve Agent", row)

row = add_subsection("Guided Mode Flow - Start Dataset", row)
row = add_test_item("Start a new dataset", row)
row = add_test_item("Create a dataset with Subs, Sam, Sites, Performances, and a Code folder", row)
row = add_test_item("Add files to every data category (Experimental, Code, Doc, Prot)", row)
row = add_test_item("Enter IDs manually", row)
row = add_test_item("Use at least 2 subjects. Have at least one sub-entity for each subject", row)
row = add_test_item("Fill out at least one entity of each included with all required and optional metadata", row)

row = add_subsection("File Association", row)
row = add_test_item("Associate at least one file with each Sub, Sam, and Site", row)
row = add_test_item("Create at least 2 perfs and associate at least 1 file to each", row)
row = add_test_item("Create modality. Associate at least 1 file", row)
row = add_test_item("Ensure that files filtered out (either through the search or not belonging to certain entities) are not modified when not displayed in the UI", row)

row = add_subsection("Manifest File Review", row)
row = add_test_item("Check that all files and folders are included", row)
row = add_test_item("Check that all entities are listed properly and that hierarchy is consistent (e.g., site-1 sam-2, site-3 sam-1)", row)
row = add_test_item("Check that there are no extra columns or misnamed columns", row)
row = add_test_item("Add custom metadata for at least 2 rows (reference at generation time)", row)
row = add_test_item("Create local copy + check it has edited fields", row)

row = add_subsection("Dataset Structure Review", row)
row = add_test_item("Make sure the dataset structure review folder hierarchy is correct", row)
row = add_test_item("Save & Exit", row)
row = add_test_item("Continue saved dataset + ensure it is correct", row)

row = add_subsection("Funding Information", row)
row = add_test_item("Choose NIH", row)
row = add_test_item("Choose SPARC", row)
row = add_test_item("Fill out form", row)
row = add_test_item("Add contributors", row)
row = add_test_item("Add a protocol", row)
row = add_test_item("Fill out dataset description and all dataset metadata", row)

row = add_subsection("Generate Dataset", row)
row = add_test_item("Generate a local copy", row)
row = add_test_item("Ensure it is correct by checking the manifest, subjects, samples, sites, and any other metadata file created", row)
row = add_test_item("Verify relationships created (e.g., 'sub-1 sam-2')", row)
row = add_test_item("Check headers are named properly, accepted values are correct, colors are correct, and order is correct", row)
row = add_test_item("Generate a Pennsieve dataset (add sign in information to PS and Agent now when prompted)", row)
row = add_test_item("Check information on Pennsieve is good", row)

row = add_subsection("Test Dataset Rename", row)
row = add_test_item("Create a dataset with invalid characters for Pennsieve datasets (e.g., ':')", row)
row = add_test_item("Generate the dataset locally - verify output name has OS invalid characters swapped but dataset description title is unchanged", row)
row = add_test_item("Generate the dataset on Pennsieve - verify user is prompted to change dataset name but title in dataset description file remains unchanged", row)

row = add_subsection("Test Guided Mode Contributor Workflow", row)
row = add_test_item("Remove Pennsieve Agent and .pennsieve folder", row)
row = add_test_item("With a user account with access to the sparc workspace, create a dataset and invite a guest account as 'Editor'", row)
row = add_test_item("Accept the invitation as a guest contributor", row)
row = add_test_item("Create an API Key on Pennsieve", row)
row = add_test_item("Use Manage Accounts to add the guest user to SODA", row)
row = add_test_item("Do minimal version of Guided Mode", row)
row = add_test_item("On the upload options page try to create a new dataset (this should fail as a guest contributor)", row)
row = add_test_item("Choose the existing dataset shared from the main account", row)
row = add_test_item("Make sure the dataset is empty (can only pull and update empty datasets in GM)", row)
row = add_test_item("Upload (should skip to the data file upload step)", row)
row = add_test_item("Check accuracy on Pennsieve of the upload", row)
row = add_test_item("Ensure dataset ownership has not changed", row)
row = add_test_item("Note: Verify SODA will not add any Pennsieve metadata, only Dataset Metadata", row)

row = add_subsection("Test Guided Mode Unit - Import Dataset with Invalid Files", row)
row = add_test_item("Import a dataset with invalid file names, folder names, and hidden files - all variations", row)

row = add_subsection("Dataset Content Page - Various Configurations", row)
row = add_test_item("Just subs, associate files, check manifest, generate, verify files", row)
row = add_test_item("Just subs and sams, check manifest, and generate, verify files", row)
row = add_test_item("Subs, sams, and sites", row)
row = add_test_item("Just code and no sub, sam, or sites", row)
row = add_test_item("Generate entity ids with Excel for subs + sams + possible entities then generate, verify files", row)
row = add_test_item("Generate a dataset that has entities without files associated locally and on Pennsieve (verify files post generation)", row)

row = add_subsection("Verify Dataset Metadata on Pennsieve", row)
row = add_test_item("Subtitle", row)
row = add_test_item("Description (dataset description should populate the description section)", row)
row = add_test_item("Tags", row)
row = add_test_item("Banner image", row)
row = add_test_item("Contributors", row)
row = add_test_item("DOI", row)
row = add_test_item("License", row)
row = add_test_item("ORCID account linked", row)
row = add_test_item("Publishing status (change the status at least once in the app and see it updated on Pennsieve Publishing tab)", row)
row = add_test_item("Check permissions are set properly", row)

row = add_subsection("Additional Tests", row)
row = add_test_item("Test automatic retry by turning off Pennsieve Agent while uploading once a few files are up. Verify files. Check accuracy on Pennsieve", row)
row = add_test_item("Create a saved dataset for SPARC and Save & Exit. Change to a different workspace in Manage Accounts and ensure dataset is not continuable without having to switch workspaces first", row)
row = add_test_item("Add multiple subjects then add samples and sites. Ensure the relationships are accurate in the UI", row)
row = add_test_item("Import up to 50K files. Go through basic workflow. Ensure nothing freezes or breaks", row)
row = add_test_item("Dataset Entity Metadata: Create enough subjects to require scrolling. Ensure scrolling works", row)

# Upload Dataset Workflow Testing
row += 1
row = add_section("UPLOAD DATASET WORKFLOW TESTING", row)

row = add_subsection("Basic Upload Workflow", row)
row = add_test_item("Remove Pennsieve Agent and .pennsieve folder", row)
row = add_test_item("Import a dataset", row)
row = add_test_item("Change the workspace back-and-forth and see that the dataset list gets updated to reflect the datasets on the workspace", row)
row = add_test_item("Pick a new dataset", row)
row = add_test_item("Generate a manifest file - check headers, colors, spelling - and add custom entries", row)
row = add_test_item("Generate the dataset", row)
row = add_test_item("Verify files", row)

row = add_subsection("Upload Dataset Update Existing Dataset", row)
row = add_test_item("Remove the Pennsieve Agent and .pennsieve folder", row)
row = add_test_item("Import a dataset", row)
row = add_test_item("Change the workspace back-and-forth and see that the dataset list gets updated", row)
row = add_test_item("Import an existing dataset", row)
row = add_test_item("Select 'Merge' and 'Skip'", row)
row = add_test_item("Ensure Manifest files cannot be created for this workflow", row)
row = add_test_item("Upload your files", row)
row = add_test_item("Verify files were uploaded", row)

row = add_subsection("Test Upload Dataset Contributor Workflow", row)
row = add_test_item("Remove Pennsieve Agent and .pennsieve folder (if not done in latest round)", row)
row = add_test_item("Create guest contributor account (if not done in latest round)", row)
row = add_test_item("Start Upload Dataset", row)
row = add_test_item("Try to create a new dataset - should be impossible as a guest", row)
row = add_test_item("Try to pull down the dataset you were given Editor access to", row)
row = add_test_item("Upload with 'Merge' and 'Skip'", row)
row = add_test_item("Check accuracy", row)

row = add_subsection("Upload Dataset Unit Testing", row)
row = add_test_item("Import a valid SDS dataset that is only accepted high level folders. SODA should allow the import if the folders are not empty", row)
row = add_test_item("Import an invalid SDS dataset that has at least one non-compliant high level folder. SODA should disallow the import", row)
row = add_test_item("Import a valid SDS dataset that has every high level folder and every metadata file (manifest included)", row)
row = add_test_item("Import a dataset with hidden files. Should see that SODA asks about the files", row)
row = add_test_item("Import a dataset with invalid file and folder names. Allow SODA to rename - check the rename happened in the manifest", row)
row = add_test_item("Import a dataset with invalid file and folder names. Allow SODA to not import those files/folder - check the proper files are imported and created in the manifest", row)
row = add_test_item("Automatic retry in new dataset upload (fail once some files pushed)", row)
row = add_test_item("Automatic retry in Contributor workflow (fail once some files pushed)", row)
row = add_test_item("Use Merge + Replace (ensure files are replaced)", row)
row = add_test_item("Use Merge + Skip (ensure only new files are added)", row)

row = add_subsection("Guest Workspace Testing", row)
row = add_test_item("Remove pennsieve folder and Pennsieve agent", row)
row = add_test_item("Sign in with API key and secret in Manage Accounts", row)
row = add_test_item("Aaron: Enable tracking at end and Upload a dataset in GM and UD and make sure the information is accurate in Kombucha", row)

# SDS Model Testing
row += 1
row = add_section("SDS MODEL TESTING (SPECIFIC TESTS TO ENSURE SODA MATCHES SDS3.0)", row)

row = add_test_item("Test that performances can only be files that are already linked to a subject or sample", row)
row = add_test_item("Test that invalid fields cannot be entered in SODA for metadata files", row)

# Freeze panes
ws.freeze_panes = "A2"

# Save the workbook
wb.save("SODA_Testing_Checklist.xlsx")
print("Excel file 'SODA_Testing_Checklist.xlsx' created successfully!")
print("\nInstructions:")
print("- Click any cell in the 'Done' column to see a dropdown")
print("- Select ☑ to mark as complete, or ☐ for not done")
print("- You can also type ☑ or ☐ directly into the cells")
print("- All test descriptions and notes are editable - click any cell to modify")

