# -*- coding: utf-8 -*-

### Import required python modules
import requests
import platform
import os
from os import listdir, makedirs, mkdir, walk, rename
from os.path import (
    isdir,
    isfile,
    join,
    splitext,
    basename,
    exists,
    expanduser,
    dirname,
    getsize,
    abspath,
)
import pandas as pd
import time
from timeit import default_timer as timer
from datetime import timedelta
import shutil
import subprocess
import gevent
import pathlib
from flask import abort
import requests
from datetime import datetime
from permissions import pennsieve_get_current_user_permissions
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils import connect_pennsieve_client, get_dataset_id, create_request_headers, TZLOCAL
from manifest import create_high_lvl_manifest_files_existing_ps_starting_point, create_high_level_manifest_files, get_auto_generated_manifest_files
from authentication import get_access_token

from pysodaUtils import (
    check_forbidden_characters_ps,
    get_agent_installation_location,
)

from organizeDatasets import import_pennsieve_dataset

from namespaces import NamespaceEnum, get_namespace_logger

namespace_logger = get_namespace_logger(NamespaceEnum.CURATE_DATASETS)


### Global variables
curateprogress = " "
curatestatus = " "
curateprintstatus = " "
total_dataset_size = 1
curated_dataset_size = 0
start_time = 0
uploaded_folder_counter = 0
current_size_of_uploaded_files = 0
generated_dataset_id = None
# the pennsieve python client used for uploading dataset files 
client = None 

userpath = expanduser("~")
configpath = join(userpath, ".pennsieve", "config.ini")
submitdataprogress = " "
submitdatastatus = " "
submitprintstatus = " "
total_file_size = 1
uploaded_file_size = 0
start_time_bf_upload = 0
start_submit = 0
metadatapath = join(userpath, "SODA", "SODA_metadata")
ps_recognized_file_extensions = [
    ".cram",
    ".jp2",
    ".jpx",
    ".lsm",
    ".ndpi",
    ".nifti",
    ".oib",
    ".oif",
    ".roi",
    ".rtf",
    ".swc",
    ".abf",
    ".acq",
    ".adicht",
    ".adidat",
    ".aedt",
    ".afni",
    ".ai",
    ".avi",
    ".bam",
    ".bash",
    ".bcl",
    ".bcl.gz",
    ".bin",
    ".brik",
    ".brukertiff.gz",
    ".continuous",
    ".cpp",
    ".csv",
    ".curv",
    ".cxls",
    ".czi",
    ".data",
    ".dcm",
    ".df",
    ".dicom",
    ".doc",
    ".docx",
    ".e",
    ".edf",
    ".eps",
    ".events",
    ".fasta",
    ".fastq",
    ".fcs",
    ".feather",
    ".fig",
    ".gif",
    ".h4",
    ".h5",
    ".hdf4",
    ".hdf5",
    ".hdr",
    ".he2",
    ".he5",
    ".head",
    ".hoc",
    ".htm",
    ".html",
    ".ibw",
    ".img",
    ".ims",
    ".ipynb",
    ".jpeg",
    ".jpg",
    ".js",
    ".json",
    ".lay",
    ".lh",
    ".lif",
    ".m",
    ".mat",
    ".md",
    ".mef",
    ".mefd.gz",
    ".mex",
    ".mgf",
    ".mgh",
    ".mgh.gz",
    ".mgz",
    ".mnc",
    ".moberg.gz",
    ".mod",
    ".mov",
    ".mp4",
    ".mph",
    ".mpj",
    ".mtw",
    ".ncs",
    ".nd2",
    ".nev",
    ".nex",
    ".nex5",
    ".nf3",
    ".nii",
    ".nii.gz",
    ".ns1",
    ".ns2",
    ".ns3",
    ".ns4",
    ".ns5",
    ".ns6",
    ".nwb",
    ".ogg",
    ".ogv",
    ".ome.btf",
    ".ome.tif",
    ".ome.tif2",
    ".ome.tif8",
    ".ome.tiff",
    ".ome.xml",
    ".openephys",
    ".pdf",
    ".pgf",
    ".png",
    ".ppt",
    ".pptx",
    ".ps",
    ".pul",
    ".py",
    ".r",
    ".raw",
    ".rdata",
    ".rh",
    ".rhd",
    ".sh",
    ".sldasm",
    ".slddrw",
    ".smr",
    ".spikes",
    ".svg",
    ".svs",
    ".tab",
    ".tar",
    ".tar.gz",
    ".tcsh",
    ".tdm",
    ".tdms",
    ".text",
    ".tif",
    ".tiff",
    ".tsv",
    ".txt",
    ".vcf",
    ".webm",
    ".xlsx",
    ".xml",
    ".yaml",
    ".yml",
    ".zip",
    ".zsh",
]

myds = ""
initial_bfdataset_size = 0
upload_directly_to_bf = 0
initial_bfdataset_size_submit = 0

total_files = 0 # the total number of files in a given dataset that need to be uploaded to Pennsieve
total_bytes_uploaded = 0 # current number of bytes uploaded to Pennsieve in the upload session
total_upload_size = 0 # total number of bytes to upload to Pennsieve in the upload session

forbidden_characters = '<>:"/\|?*'
forbidden_characters_bf = '\/:*?"<>'

# a global that tracks the amount of files that have been uploaded in an upload session;
# is reset once the session ends by success, or failure (is implicitly reset in case of Pennsieve Agent freeze by the user closing SODA)
main_curation_uploaded_files = 0

DEV_TEMPLATE_PATH = join(dirname(__file__), "..", "file_templates")

# once pysoda has been packaged with pyinstaller
# it becomes nested into the pysodadist/api directory
PROD_TEMPLATE_PATH = join(dirname(__file__), "..", "..", "file_templates")
TEMPLATE_PATH = DEV_TEMPLATE_PATH if exists(DEV_TEMPLATE_PATH) else PROD_TEMPLATE_PATH


PENNSIEVE_URL = "https://api.pennsieve.io"




def open_file(file_path):
    """
    Opening folder on all platforms
    https://stackoverflow.com/questions/6631299/python-opening-a-folder-in-explorer-nautilus-mac-thingie

    Args:
        file_path: path of the folder (string)
    Action:
        Opens file explorer window to the given path
    """
    try:
        if platform.system() == "Windows":
            subprocess.Popen(f"explorer /select,{str(file_path)}")
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])
    except Exception as e:
        raise e


def folder_size(path):
    """
    Provides the size of the folder indicated by path

    Args:
        path: path of the folder (string)
    Returns:
        total_size: total size of the folder in bytes (integer)
    """
    total_size = 0

    for path, dirs, files in walk(path):
        for f in files:
            fp = join(path, f)
            total_size += getsize(fp)
    return total_size


def path_size(path):
    """
    Returns size of the path, after checking if it's a folder or a file
    Args:
        path: path of the file/folder (string)
    Returns:
        total_size: total size of the file/folder in bytes (integer)
    """
    return folder_size(path) if isdir(path) else getsize(path)


def create_folder_level_manifest(jsonpath, jsondescription):
    """
    Function to create manifest files for each SPARC folder.
    Files are created in a temporary folder

    Args:
        datasetpath: path of the dataset (string)
        jsonpath: all paths in json format with key being SPARC folder names (dictionary)
        jsondescription: description associated with each path (dictionary)
    Action:
        Creates manifest files in xslx format for each SPARC folder
    """
    global total_dataset_size
    local_timezone = TZLOCAL()

    try:
        shutil.rmtree(metadatapath) if isdir(metadatapath) else 0
        makedirs(metadatapath)
        folders = list(jsonpath.keys())

        if "main" in folders:
            folders.remove("main")
        # In each SPARC folder, generate a manifest file
        for folder in folders:
            if jsonpath[folder] != []:
                # Initialize dataframe where manifest info will be stored
                df = pd.DataFrame(
                    columns=[
                        "filename",
                        "timestamp",
                        "description",
                        "file type",
                        "Additional Metadata",
                    ]
                )
                # Get list of files/folders in the the folder
                # Remove manifest file from the list if already exists
                folderpath = join(metadatapath, folder)
                allfiles = jsonpath[folder]
                alldescription = jsondescription[folder + "_description"]

                countpath = -1
                for pathname in allfiles:
                    countpath += 1
                    if basename(pathname) in ["manifest.csv", "manifest.xlsx"]:
                        allfiles.pop(countpath)
                        alldescription.pop(countpath)

                # Populate manifest dataframe
                filename, timestamp, filetype, filedescription = [], [], [], []
                countpath = -1
                for paths in allfiles:
                    if isdir(paths):
                        key = basename(paths)
                        alldescription.pop(0)
                        for subdir, dirs, files in os.walk(paths):
                            for file in files:
                                gevent.sleep(0)
                                filepath = pathlib.Path(paths) / subdir / file
                                mtime = filepath.stat().st_mtime
                                lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                                    local_timezone
                                )
                                timestamp.append(
                                    lastmodtime.isoformat()
                                    .replace(".", ",")
                                    .replace("+00:00", "Z")
                                )
                                full_filename = filepath.name

                                if folder == "main":  # if file in main folder
                                    filename.append(
                                        full_filename
                                    ) if folder == "" else filename.append(
                                        join(folder, full_filename)
                                    )
                                else:
                                    subdirname = os.path.relpath(
                                        subdir, paths
                                    )  # gives relative path of the directory of the file w.r.t paths
                                    if subdirname == ".":
                                        filename.append(join(key, full_filename))
                                    else:
                                        filename.append(
                                            join(key, subdirname, full_filename)
                                        )

                                fileextension = splitext(full_filename)[1]
                                if (
                                    not fileextension
                                ):  # if empty (happens e.g. with Readme files)
                                    fileextension = "None"
                                filetype.append(fileextension)
                                filedescription.append("")
                    else:
                        gevent.sleep(0)
                        countpath += 1
                        filepath = pathlib.Path(paths)
                        file = filepath.name
                        filename.append(file)
                        mtime = filepath.stat().st_mtime
                        lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                            local_timezone
                        )
                        timestamp.append(
                            lastmodtime.isoformat()
                            .replace(".", ",")
                            .replace("+00:00", "Z")
                        )
                        filedescription.append(alldescription[countpath])
                        if isdir(paths):
                            filetype.append("folder")
                        else:
                            fileextension = splitext(file)[1]
                            if (
                                not fileextension
                            ):  # if empty (happens e.g. with Readme files)
                                fileextension = "None"
                            filetype.append(fileextension)

                df["filename"] = filename
                df["timestamp"] = timestamp
                df["file type"] = filetype
                df["description"] = filedescription

                makedirs(folderpath)
                # Save manifest as Excel sheet
                manifestfile = join(folderpath, "manifest.xlsx")
                df.to_excel(manifestfile, index=None, header=True)
                wb = load_workbook(manifestfile)
                ws = wb.active

                blueFill = PatternFill(
                    start_color="9DC3E6", fill_type="solid"
                )
                greenFill = PatternFill(
                    start_color="A8D08D", fill_type="solid"
                )
                yellowFill = PatternFill(
                    start_color="FFD965", fill_type="solid"
                )
                ws['A1'].fill = blueFill
                ws['B1'].fill = greenFill
                ws['C1'].fill = greenFill
                ws['D1'].fill = greenFill
                ws['E1'].fill = yellowFill

                wb.save(manifestfile)
                total_dataset_size += path_size(manifestfile)
                jsonpath[folder].append(manifestfile)

        return jsonpath

    except Exception as e:
        raise e


def return_new_path(topath):
    """
    This function checks if a folder already exists and in such cases,
    appends (1) or (2) etc. to the folder name

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """

    if not exists(topath):
        return topath

    i = 1
    while True:
        if not exists(topath + " (" + str(i) + ")"):
            return topath + " (" + str(i) + ")"
        i += 1


def return_new_path_replace(topath):
    """
    This function checks if a folder already exists and in such cases,
    replace the existing folder (this is the opposite situation to the function return_new_path)

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """

    if not exists(topath):
        return topath
    i = 1
    while True:
        if not exists(topath + " (" + str(i) + ")"):
            return topath + " (" + str(i) + ")"
        i += 1


def time_format(elapsed_time):
    mins, secs = divmod(elapsed_time, 60)
    hours, mins = divmod(mins, 60)
    return "%dh:%02dmin:%02ds" % (hours, mins, secs)


def mycopyfileobj(fsrc, fdst, length=16 * 1024 * 16):
    """
    Helper function to copy file

    Args:
        fsrc: source file opened in python (file-like object)
        fdst: destination file accessed in python (file-like object)
        length: copied buffer size in bytes (integer)
    """
    global curateprogress
    global total_dataset_size
    global curated_dataset_size
    global main_generated_dataset_size

    while True:
        buf = fsrc.read(length)
        if not buf:
            break
        gevent.sleep(0)
        fdst.write(buf)
        curated_dataset_size += len(buf)
        main_generated_dataset_size += len(buf)


def mycopyfile_with_metadata(src, dst, *, follow_symlinks=True):
    """
    Copy file src to dst with metadata (timestamp, permission, etc.) conserved

    Args:
        src: source file (string)
        dst: destination file (string)
    Returns:
        dst
    """
    if not follow_symlinks and os.path.islink(src):
        os.symlink(os.readlink(src), dst)
    else:
        with open(src, "rb") as fsrc:
            with open(dst, "wb") as fdst:
                mycopyfileobj(fsrc, fdst)
    shutil.copystat(src, dst)
    return dst


def check_empty_files_folders(soda_json_structure):
    """
    Function to check for empty files and folders

    Args:
        soda_json_structure: soda dict with information about all specified files and folders
    Output:
        error: error message with list of non valid local data files, if any
    """
    try:
        def recursive_empty_files_check(my_folder, my_relative_path, error_files):
            for folder_key, folder in my_folder["folders"].items():
                relative_path = my_relative_path + "/" + folder_key
                error_files = recursive_empty_files_check(
                    folder, relative_path, error_files
                )

            for file_key in list(my_folder["files"].keys()):
                file = my_folder["files"][file_key]
                file_type = file["type"]
                if file_type == "local":
                    file_path = file["path"]
                    if isfile(file_path):
                        file_size = getsize(file_path)
                        if file_size == 0:
                            del my_folder["files"][file_key]
                            relative_path = my_relative_path + "/" + file_key
                            error_message = relative_path + " (path: " + file_path + ")"
                            error_files.append(error_message)

            return error_files

        def recursive_empty_local_folders_check(
                    my_folder,
                    my_folder_key,
                    my_folders_content,
                    my_relative_path,
                    error_folders,
                ):
            folders_content = my_folder["folders"]
            for folder_key in list(my_folder["folders"].keys()):
                folder = my_folder["folders"][folder_key]
                relative_path = my_relative_path + "/" + folder_key
                error_folders = recursive_empty_local_folders_check(
                    folder, folder_key, folders_content, relative_path, error_folders
                )

            if not my_folder["folders"] and not my_folder["files"]:
                ignore = False
                if "type" in my_folder and my_folder["type"] == "bf":
                    ignore = True
                if not ignore:
                    error_message = my_relative_path
                    error_folders.append(error_message)
                    del my_folders_content[my_folder_key]
            return error_folders

        error_files = []
        error_folders = []
        if "dataset-structure" in soda_json_structure.keys():
            dataset_structure = soda_json_structure["dataset-structure"]
            if "folders" in dataset_structure:
                for folder_key, folder in dataset_structure["folders"].items():
                    relative_path = folder_key
                    error_files = recursive_empty_files_check(
                        folder, relative_path, error_files
                    )

                folders_content = dataset_structure["folders"]
                for folder_key in list(dataset_structure["folders"].keys()):
                    folder = dataset_structure["folders"][folder_key]
                    relative_path = folder_key
                    error_folders = recursive_empty_local_folders_check(
                        folder,
                        folder_key,
                        folders_content,
                        relative_path,
                        error_folders,
                    )

        if "metadata-files" in soda_json_structure.keys():
            metadata_files = soda_json_structure["metadata-files"]
            for file_key in list(metadata_files.keys()):
                file = metadata_files[file_key]
                file_type = file["type"]
                if file_type == "local":
                    file_path = file["path"]
                    if isfile(file_path):
                        file_size = getsize(file_path)
                        if file_size == 0:
                            del metadata_files[file_key]
                            error_message = file_key + " (path: " + file_path + ")"
                            error_files.append(error_message)
            if not metadata_files:
                del soda_json_structure["metadata-files"]

        if len(error_files) > 0:
            error_message = [
                "The following local file(s) is/are empty (0 kb) and will be ignored."
            ]
            error_files = error_message + [] + error_files

        if len(error_folders) > 0:
            error_message = [
                "The SPARC dataset structure does not allow empty folders. The following empty folders will be removed from your dataset:"
            ]
            error_folders = error_message + [] + error_folders

        return {
            "empty_files": error_files, 
            "empty_folders": error_folders, 
            "soda_json_structure": soda_json_structure
        }

    except Exception as e:
        raise e


def check_local_dataset_files_validity(soda_json_structure):
    """
    Function to check that the local data files and folders specified in the dataset are valid

    Args:
        soda_json_structure: soda dict with information about all specified files and folders
    Output:
        error: error message with list of non valid local data files, if any
    """

    def recursive_local_file_check(my_folder, my_relative_path, error):
        for folder_key, folder in my_folder["folders"].items():
            relative_path = my_relative_path + "/" + folder_key
            error = recursive_local_file_check(folder, relative_path, error)

        for file_key in list(my_folder["files"].keys()):
            file = my_folder["files"][file_key]
            if file_key in ["manifest.xlsx", "manifest.csv"]:
                continue
            file_type = file["type"]
            if file_type == "local":
                file_path = file["path"]
                if file["type"] == "bf":
                    continue
                if not isfile(file_path):
                    relative_path = my_relative_path + "/" + file_key
                    error_message = relative_path + " (path: " + file_path + ")"
                    error.append(error_message)
                else:
                    file_size = getsize(file_path)
                    if file_size == 0:
                        del my_folder["files"][file_key]

        return error

    def recursive_empty_local_folder_remove(
        my_folder, my_folder_key, my_folders_content
    ):

        folders_content = my_folder["folders"]
        for folder_key in list(my_folder["folders"].keys()):
            folder = my_folder["folders"][folder_key]
            recursive_empty_local_folder_remove(folder, folder_key, folders_content)

        if not my_folder["folders"] and not my_folder["files"] and my_folder["type"] != "bf":
            del my_folders_content[my_folder_key]

    error = []
    if "dataset-structure" in soda_json_structure.keys():
        dataset_structure = soda_json_structure["dataset-structure"]
        # Remove 0kb files, files that can't be found, and any empty folders from the dataset data files
        if "folders" in dataset_structure:
            for folder_key, folder in dataset_structure["folders"].items():
                relative_path = folder_key
                error = recursive_local_file_check(folder, relative_path, error)

            folders_content = dataset_structure["folders"]
            for folder_key in list(dataset_structure["folders"].keys()):
                folder = dataset_structure["folders"][folder_key]
                recursive_empty_local_folder_remove(folder, folder_key, folders_content)

    if "metadata-files" in soda_json_structure.keys():
        metadata_files = soda_json_structure["metadata-files"]
        # Remove specified metadata files that do net exist at the specified paths or that are empty
        for file_key in list(metadata_files.keys()):
            file = metadata_files[file_key]
            file_type = file["type"]
            if file_type == "local":
                file_path = file["path"]
                if not isfile(file_path):
                    error_message = file_key + " (path: " + file_path + ")"
                    error.append(error_message)
                else:
                    file_size = getsize(file_path)
                    if file_size == 0:
                        del metadata_files[file_key]
        if not metadata_files:
            del soda_json_structure["metadata-files"]

    # Return list of all the files that were not found. 
    if len(error) > 0:
        error_message = [
            "Error: The following local files were not found. Specify them again or remove them."
        ]
        error = error_message + error

    return error


# path to local SODA folder for saving manifest files
manifest_sparc = ["manifest.xlsx", "manifest.csv"]
manifest_folder_path = join(userpath, "SODA", "manifest_files")



def check_JSON_size(jsonStructure):
    """
        This function is called to check size of files that will be created locally on a user's device.
    """
    global total_dataset_size
    total_dataset_size = 0

    try:
        def recursive_dataset_scan(folder):
            global total_dataset_size

            if "files" in folder.keys():
                for file_key, file in folder["files"].items():
                    if "deleted" not in file["action"]:
                        file_type = file["type"]
                        if file_type == "local":
                            file_path = file["path"]
                            if isfile(file_path):
                                total_dataset_size += getsize(file_path)

            if "folders" in folder.keys():
                for folder_key, folder in folder["folders"].items():
                    recursive_dataset_scan(folder)

        # scan dataset structure
        dataset_structure = jsonStructure["dataset-structure"]
        folderSection = dataset_structure["folders"]
        # gets keys like code, primary, source and their content...
        for keys, contents in folderSection.items():
            recursive_dataset_scan(contents)

        if "metadata-files" in jsonStructure.keys():
            metadata_files = jsonStructure["metadata-files"]
            for file_key, file in metadata_files.items():
                if file["type"] == "local":
                    metadata_path = file["path"]
                    if isfile(metadata_path) and "new" in file["action"]:
                        total_dataset_size += getsize(metadata_path)

        if "manifest-files" in jsonStructure.keys():
            manifest_files_structure = create_high_level_manifest_files(jsonStructure, manifest_folder_path)
            for key in manifest_files_structure.keys():
                manifestpath = manifest_files_structure[key]
                if isfile(manifestpath):
                    total_dataset_size += getsize(manifestpath)

        # returns in bytes
        return {"dataset_size": total_dataset_size}
    except Exception as e:
        raise e


def generate_dataset_locally(soda_json_structure):
    global namespace_logger

    namespace_logger.info("starting generate_dataset_locally")

    # Vars used for tracking progress on the frontend 
    global main_curate_progress_message
    global progress_percentage
    global main_total_generate_dataset_size
    global start_generate
    global main_curation_uploaded_files

    main_curation_uploaded_files = 0

    try:

        def recursive_dataset_scan(
            my_folder, my_folderpath, list_copy_files, list_move_files
        ):
            global main_total_generate_dataset_size

            if "folders" in my_folder.keys():
                for folder_key, folder in my_folder["folders"].items():
                    folderpath = join(my_folderpath, folder_key)
                    if not isdir(folderpath):
                        mkdir(folderpath)
                    list_copy_files, list_move_files = recursive_dataset_scan(
                        folder, folderpath, list_copy_files, list_move_files
                    )

            if "files" in my_folder.keys():
                for file_key, file in my_folder["files"].items():
                    if "deleted" not in file["action"]:
                        file_type = file["type"]
                        if file_type == "local":
                            file_path = file["path"]
                            if isfile(file_path):
                                destination_path = abspath(
                                    join(my_folderpath, file_key)
                                )
                                if not isfile(destination_path):
                                    if (
                                        "existing" in file["action"]
                                        and soda_json_structure["generate-dataset"][
                                            "if-existing"
                                        ]
                                        == "merge"
                                    ):
                                        list_move_files.append(
                                            [file_path, destination_path]
                                        )
                                    else:
                                        main_total_generate_dataset_size += getsize(
                                            file_path
                                        )
                                        list_copy_files.append(
                                            [file_path, destination_path]
                                        )
            return list_copy_files, list_move_files


        namespace_logger.info("generate_dataset_locally step 1")
        # 1. Create new folder for dataset or use existing merge with existing or create new dataset
        main_curate_progress_message = "Generating folder structure and list of files to be included in the dataset"
        dataset_absolute_path = soda_json_structure["generate-dataset"]["path"]
        if_existing = soda_json_structure["generate-dataset"]["if-existing"]
        dataset_name = soda_json_structure["generate-dataset"]["dataset-name"]
        datasetpath = join(dataset_absolute_path, dataset_name)
        datasetpath = return_new_path(datasetpath)
        mkdir(datasetpath)

        namespace_logger.info("generate_dataset_locally step 2")
        # 2. Scan the dataset structure and:
        # 2.1. Create all folders (with new name if renamed)
        # 2.2. Compile a list of files to be copied and a list of files to be moved (with new name recorded if renamed)
        list_copy_files = []
        list_move_files = []
        dataset_structure = soda_json_structure["dataset-structure"]
        for folder_key, folder in dataset_structure["folders"].items():
            folderpath = join(datasetpath, folder_key)
            mkdir(folderpath)
            list_copy_files, list_move_files = recursive_dataset_scan(
                folder, folderpath, list_copy_files, list_move_files
            )


        # 3. Add high-level metadata files in the list
        if "metadata-files" in soda_json_structure.keys():
            namespace_logger.info("generate_dataset_locally (optional) step 3 handling metadata-files")
            metadata_files = soda_json_structure["metadata-files"]
            for file_key, file in metadata_files.items():
                if file["type"] == "local":
                    metadata_path = file["path"]
                    if isfile(metadata_path):
                        destination_path = join(datasetpath, file_key)
                        if "existing" in file["action"]:
                            list_move_files.append([metadata_path, destination_path])
                        elif "new" in file["action"]:
                            main_total_generate_dataset_size += getsize(metadata_path)
                            list_copy_files.append([metadata_path, destination_path])

        # 4. Add manifest files in the list
        if "manifest-files" in soda_json_structure.keys():
            namespace_logger.info("generate_dataset_locally (optional) step 4 handling manifest-files")
            main_curate_progress_message = "Preparing manifest files"
            manifest_files_structure = create_high_level_manifest_files(
                soda_json_structure, manifest_folder_path
            )
            for key in manifest_files_structure.keys():
                manifestpath = manifest_files_structure[key]
                if isfile(manifestpath):
                    destination_path = join(datasetpath, key, "manifest.xlsx")
                    main_total_generate_dataset_size += getsize(manifestpath)
                    list_copy_files.append([manifestpath, destination_path])

        namespace_logger.info("generate_dataset_locally step 5 moving files to new location")
        # 5. Move files to new location
        main_curate_progress_message = "Moving files to new location"
        for fileinfo in list_move_files:
            srcfile = fileinfo[0]
            distfile = fileinfo[1]
            main_curate_progress_message = f"Moving file {str(srcfile)} to {str(distfile)}"
            shutil.move(srcfile, distfile)

        namespace_logger.info("generate_dataset_locally step 6 copying files to new location")
        # 6. Copy files to new location
        main_curate_progress_message = "Copying files to new location"
        start_generate = 1
        for fileinfo in list_copy_files:
            srcfile = fileinfo[0]
            distfile = fileinfo[1]
            main_curate_progress_message = f"Copying file {str(srcfile)} to {str(distfile)}"
            # track amount of copied files for loggin purposes
            mycopyfile_with_metadata(srcfile, distfile)
            main_curation_uploaded_files += 1


        namespace_logger.info("generate_dataset_locally step 7")
        # 7. Delete manifest folder and original folder if merge requested and rename new folder
        shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
        if if_existing == "merge":
            namespace_logger.info("generate_dataset_locally (optional) step 7.1 delete manifest folder if merge requested")
            main_curate_progress_message = "Finalizing dataset"
            original_dataset_path = join(dataset_absolute_path, dataset_name)
            shutil.rmtree(original_dataset_path)
            rename(datasetpath, original_dataset_path)
            open_file(join(dataset_absolute_path, original_dataset_path))
        else:
            open_file(join(dataset_absolute_path, datasetpath))
        return datasetpath, main_total_generate_dataset_size

    except Exception as e:
        raise e
    


def ps_create_new_dataset(datasetname, ps):
    """
    Args:
        datasetname: name of the dataset to be created (string)
        bf: Pennsieve account object
    Action:
        Creates dataset for the account specified
    """
    try:
        error, count = "", 0
        datasetname = datasetname.strip()

        if check_forbidden_characters_ps(datasetname):
            error = (
                f"{error}Error: A Pennsieve dataset name cannot contain any of the following characters: "
                + forbidden_characters_bf
                + "<br>"
            )
            count += 1

        if not datasetname:
            error = f"{error}Error: Please enter valid dataset name<br>"
            count += 1

        if datasetname.isspace():
            error = error + "Error: Please enter valid dataset name" + "<br>"
            count += 1

        if count > 0:
            abort(400, error)

        try:
            r = requests.get(f"{PENNSIEVE_URL}/datasets", headers=create_request_headers(ps))
            r.raise_for_status()
            dataset_dicts = r.json()
        except Exception as e:
            abort(500, "Error: Could not connect to Pennsieve. Please try again later.")


        dataset_list = [
            dataset_dict["content"]["name"] for dataset_dict in dataset_dicts
        ]
        if datasetname in dataset_list:
            abort(400, "Error: Dataset name already exists")
        else:
            r = requests.post(f"{PENNSIEVE_URL}/datasets", headers=create_request_headers(ps), json={"name": datasetname})
            r.raise_for_status()

        return r.json()

    except Exception as e:
        raise e

double_extensions = [
    ".ome.tiff",
    ".ome.tif",
    ".ome.tf2,",
    ".ome.tf8",
    ".ome.btf",
    ".ome.xml",
    ".brukertiff.gz",
    ".mefd.gz",
    ".moberg.gz",
    ".nii.gz",
    ".mgh.gz",
    ".tar.gz",
    ".bcl.gz",
]


def create_high_lvl_manifest_files_existing_ps(
    soda_json_structure, ps, my_tracking_folder
):
    """
    Function to create manifest files for each high-level SPARC folder.

    Args:
        soda_json_structure: soda dict with information about the dataset to be generated/modified
    Action:
        manifest_files_structure: dict including the local path of the manifest files
    """
    def get_name_extension(file_name):
        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break
        ext = ""
        name = ""
        if double_ext == False:
            name = os.path.splitext(file_name)[0]
            ext = os.path.splitext(file_name)[1]
        else:
            ext = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )
            name = os.path.splitext(os.path.splitext(file_name)[0])[0]
        return name, ext
    
    def recursive_import_ps_manifest_info(
        folder, my_relative_path, dict_folder_manifest, manifest_df
    ):
        """
        Import manifest information from the Pennsieve dataset for the given folder and its children.
        """
        
        if len(folder['children']) == 0:
            r = requests.get(f"{PENNSIEVE_URL}/packages/{folder['content']['id']}", headers=create_request_headers(ps), json={"include": "files"})
            r.raise_for_status()
            ps_folder = r.json()
            normalize_tracking_folder(ps_folder)
            folder['children'] = ps_folder['children']

        for _, folder_item in folder["children"]["folders"].items():
            folder_name = folder_item['content']['name']
            relative_path = generate_relative_path(
                my_relative_path, folder_name
            )
            dict_folder_manifest = recursive_import_ps_manifest_info(
                folder_item, relative_path, dict_folder_manifest, manifest_df
            )
        for _, file in folder["children"]["files"].items():
            if file['content']['name'] != "manifest":
                file_id = file['content']['id']
                r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/view", headers=create_request_headers(ps))
                r.raise_for_status()
                file_details = r.json()
                file_name = file_details[0]["content"]["name"]
                file_extension = splitext(file_name)[1]
                file_name_with_extension = (
                    splitext(file['content']['name'])[0] + file_extension
                )
                relative_path = generate_relative_path(
                    my_relative_path, file_name_with_extension
                )
                dict_folder_manifest["filename"].append(relative_path)
                # file type
                file_extension = get_name_extension(file_name)
                if file_extension == "":
                    file_extension = "None"
                dict_folder_manifest["file type"].append(file_extension)
                # timestamp, description, Additional Metadata
                if not manifest_df.empty:
                    if relative_path in manifest_df["filename"].values:
                        timestamp = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["timestamp"].iloc[0]
                        description = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["description"].iloc[0]
                        additional_metadata = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["Additional Metadata"].iloc[0]
                    else:
                        timestamp = ""
                        description = ""
                        additional_metadata = ""
                    dict_folder_manifest["timestamp"].append(timestamp)
                    dict_folder_manifest["description"].append(description)
                    dict_folder_manifest["Additional Metadata"].append(
                        additional_metadata
                    )
                else:
                    dict_folder_manifest["timestamp"].append("")
                    dict_folder_manifest["description"].append("")
                    dict_folder_manifest["Additional Metadata"].append("")
        return dict_folder_manifest
    
    # Merge existing folders
    def recursive_manifest_builder_existing_ps(
        my_folder,
        my_bf_folder,
        my_bf_folder_exists,
        my_relative_path,
        dict_folder_manifest,
    ):
        if "folders" in my_folder.keys():
            if my_bf_folder_exists:
                (
                    my_bf_existing_folders_name,
                ) = ps_get_existing_folders_details(my_bf_folder['children']['folders'])
            else:
                my_bf_existing_folders_name = []
            for folder_key, folder in my_folder["folders"].items():
                relative_path = generate_relative_path(my_relative_path, folder_key)
                if folder_key in my_bf_existing_folders_name:
                    bf_folder = my_bf_folder["children"]["folders"][folder_key]
                    bf_folder_exists = True
                else:
                    bf_folder = ""
                    bf_folder_exists = False
                dict_folder_manifest = recursive_manifest_builder_existing_ps(
                    folder,
                    bf_folder,
                    bf_folder_exists,
                    relative_path,
                    dict_folder_manifest,
                )
        if "files" in my_folder.keys():
            if my_bf_folder_exists:
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_bf_folder, bf)
            else:
                my_bf_existing_files = []
                my_bf_existing_files_name = []
                my_bf_existing_files_name_with_extension = []
            for file_key, file in my_folder["files"].items():
                if file["type"] == "local":
                    file_path = file["path"]
                    if isfile(file_path):
                        desired_name = splitext(file_key)[0]
                        file_extension = splitext(file_key)[1]
                        # manage existing file request
                        if existing_file_option == "skip" and file_key in my_bf_existing_files_name_with_extension:
                            continue
                        if existing_file_option == "replace" and file_key in my_bf_existing_files_name_with_extension:
                            # remove existing from manifest
                            filename = generate_relative_path(
                                my_relative_path, file_key
                            )
                            filename_list = dict_folder_manifest["filename"]
                            index_file = filename_list.index(filename)
                            del dict_folder_manifest["filename"][index_file]
                            del dict_folder_manifest["timestamp"][index_file]
                            del dict_folder_manifest["description"][index_file]
                            del dict_folder_manifest["file type"][index_file]
                            del dict_folder_manifest["Additional Metadata"][
                                index_file
                            ]
                            index_name = (
                                my_bf_existing_files_name_with_extension.index(
                                    file_key
                                )
                            )
                            del my_bf_existing_files[index_name]
                            del my_bf_existing_files_name[index_name]
                            del my_bf_existing_files_name_with_extension[
                                index_name
                            ]
                        if desired_name not in my_bf_existing_files_name:
                            final_name = file_key
                        else:
                            # expected final name
                            count_done = 0
                            final_name = desired_name
                            output = get_base_file_name(desired_name)
                            if output:
                                base_name = output[0]
                                count_exist = output[1]
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            base_name + "(" + str(count_exist) + ")"
                                        )
                                    else:
                                        count_done = 1
                            else:
                                count_exist = 0
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            desired_name
                                            + " ("
                                            + str(count_exist)
                                            + ")"
                                        )
                                    else:
                                        count_done = 1
                            final_name = final_name + file_extension
                            my_bf_existing_files_name.append(
                                splitext(final_name)[0]
                            )
                        # filename
                        filename = generate_relative_path(
                            my_relative_path, final_name
                        )
                        dict_folder_manifest["filename"].append(filename)
                        # timestamp
                        file_path = file["path"]
                        filepath = pathlib.Path(file_path)
                        mtime = filepath.stat().st_mtime
                        lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                            local_timezone
                        )
                        dict_folder_manifest["timestamp"].append(
                            lastmodtime.isoformat()
                            .replace(".", ",")
                            .replace("+00:00", "Z")
                        )
                        # description
                        if "description" in file.keys():
                            dict_folder_manifest["description"].append(
                                file["description"]
                            )
                        else:
                            dict_folder_manifest["description"].append("")
                        # file type
                        if file_extension == "":
                            file_extension = "None"
                        dict_folder_manifest["file type"].append(file_extension)
                        # addtional metadata
                        if "additional-metadata" in file.keys():
                            dict_folder_manifest["Additional Metadata"].append(
                                file["additional-metadata"]
                            )
                        else:
                            dict_folder_manifest["Additional Metadata"].append("")
        return dict_folder_manifest

    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]

    try:
        # create local folder to save manifest files temporarly (delete any existing one first)
        shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
        makedirs(manifest_folder_path)

        # import info about files already on ps
        dataset_structure = soda_json_structure["dataset-structure"]
        manifest_dict_save = {}
        for high_level_folder_key, high_level_folder in my_tracking_folder["children"]["folders"].items():
            if (
                high_level_folder_key in dataset_structure["folders"].keys()
            ):

                relative_path = ""
                dict_folder_manifest = {}
                # Initialize dict where manifest info will be stored
                dict_folder_manifest["filename"] = []
                dict_folder_manifest["timestamp"] = []
                dict_folder_manifest["description"] = []
                dict_folder_manifest["file type"] = []
                dict_folder_manifest["Additional Metadata"] = []

                # pull manifest file into if exists 
                manifest_df = pd.DataFrame()
                for file_key, file in high_level_folder['children']['files'].items():
                    file_id = file['content']['id']
                    r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/view", headers=create_request_headers(ps))
                    r.raise_for_status()
                    file_details = r.json()
                    file_name_with_extension = file_details[0]["content"]["name"]
                    if file_name_with_extension in manifest_sparc:
                        file_id_2 = file_details[0]["content"]["id"]
                        r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/files/{file_id_2}", headers=create_request_headers(ps))
                        r.raise_for_status()
                        file_url_info = r.json()
                        file_url = file_url_info["url"]
                        manifest_df = pd.read_excel(file_url, engine="openpyxl")
                        manifest_df = manifest_df.fillna("")
                        if (
                            "filename" not in manifest_df.columns
                            or "description" not in manifest_df.columns
                            or "Additional Metadata" not in manifest_df.columns
                        ):
                            manifest_df = pd.DataFrame()
                        break

                # store the data frame pulled from Pennsieve into a dictionary
                dict_folder_manifest =  recursive_import_ps_manifest_info(
                    high_level_folder, relative_path, dict_folder_manifest, manifest_df
                )

                manifest_dict_save[high_level_folder_key] = {
                    "manifest": dict_folder_manifest,
                    "bf_folder": high_level_folder,
                }

        # import info from local files to be uploaded
        local_timezone = TZLOCAL()
        manifest_files_structure = {}
        existing_folder_option = soda_json_structure["generate-dataset"]["if-existing"]
        existing_file_option = soda_json_structure["generate-dataset"][
            "if-existing-files"
        ]
        for folder_key, folder in dataset_structure["folders"].items():
            relative_path = ""

            if (
                folder_key in manifest_dict_save.keys()
                and existing_folder_option == "merge"
            ):
                bf_folder = manifest_dict_save[folder_key]["bf_folder"]
                bf_folder_exists = True
                dict_folder_manifest = manifest_dict_save[folder_key]["manifest"]

            elif (
                folder_key in manifest_dict_save.keys()
                and folder_key not in my_tracking_folder["children"]["folders"].keys()
                and existing_folder_option == "skip"
            ):
                continue

            else:
                bf_folder = ""
                bf_folder_exists = False
                dict_folder_manifest = {}
                dict_folder_manifest["filename"] = []
                dict_folder_manifest["timestamp"] = []
                dict_folder_manifest["description"] = []
                dict_folder_manifest["file type"] = []
                dict_folder_manifest["Additional Metadata"] = []

            dict_folder_manifest = recursive_manifest_builder_existing_ps(
                folder, bf_folder, bf_folder_exists, relative_path, dict_folder_manifest
            )

            # create high-level folder at the temporary location
            folderpath = join(manifest_folder_path, folder_key)
            makedirs(folderpath)

            # save manifest file
            manifestfilepath = join(folderpath, "manifest.xlsx")
            df = pd.DataFrame.from_dict(dict_folder_manifest)
            df.to_excel(manifestfilepath, index=None, header=True)
            wb = load_workbook(manifestfilepath)
            ws = wb.active

            blueFill = PatternFill(
                start_color="9DC3E6", fill_type="solid"
            )
            greenFill = PatternFill(
                start_color="A8D08D", fill_type="solid"
            )
            yellowFill = PatternFill(
                start_color="FFD965", fill_type="solid"
            )
            ws['A1'].fill = blueFill
            ws['B1'].fill = greenFill
            ws['C1'].fill = greenFill
            ws['D1'].fill = greenFill
            ws['E1'].fill = yellowFill
            wb.save(manifestfilepath)

            manifest_files_structure[folder_key] = manifestfilepath

        return manifest_files_structure

    except Exception as e:
        raise e





def generate_relative_path(x, y):
    return x + "/" + y if x else y


def ps_get_existing_folders_details(ps_folders):
    ps_existing_folders = [ps_folders[folder] for folder in ps_folders if ps_folders[folder]["content"]["packageType"] == "Collection"]
    ps_existing_folders_name = [folder['content']["name"] for folder in ps_existing_folders]

    return ps_existing_folders, ps_existing_folders_name


def ps_get_existing_files_details(ps_folder, ps):
    # TODO: Dorian -> ["extensions doesn't seem to be returned anymore by the endpoint"]
    def verify_file_name(file_name, extension):
        if extension == "":
            return file_name

        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break

        extension_from_name = ""

        if double_ext == False:
            extension_from_name = os.path.splitext(file_name)[1]
        else:
            extension_from_name = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )

        if extension_from_name == ("." + extension):
            return file_name
        else:
            return file_name + ("." + extension)

    files = ps_folder["children"]["files"]
    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]


    bf_existing_files_name = [splitext(files[file]['content']["name"])[0] for file in files]
    bf_existing_files_name_with_extension = []

    # determine if we are at the root of the dataset
    content = ps_folder["content"]
    if (str(content['id'])[2:9]) == "dataset":
        r = requests.get(f"{PENNSIEVE_URL}/datasets/{content['id']}", headers=create_request_headers(get_access_token())) 
        r.raise_for_status()
        root_folder = r.json()
        root_children = root_folder["children"]
        for item in root_children:
            file_name_with_extension = ""
            item_id = item["content"]["id"]
            item_name = item["content"]["name"]
            if item_id[2:9] == "package":
                if("extension" not in root_children):
                    file_name_with_extension = verify_file_name(item_name,"")
                else:
                    file_name_with_extension = verify_file_name(item_name, root_children["extension"])

            if file_name_with_extension == "":
                continue
            bf_existing_files_name_with_extension.append(file_name_with_extension)
    else:
        #is collection - aka a folder in the dataset
        for file_key, file in files.items():
            file_name_with_extension = ""
            file_name = file["content"]["name"]
            file_id = file["content"]["id"]
            if file_id[2:9] == "package":
                if "extension" not in file:
                    file_name_with_extension = verify_file_name(file_name,"")
                else:
                    file_name_with_extension = verify_file_name(file_name, file["extension"])
            if file_name_with_extension == "":
                continue
            bf_existing_files_name_with_extension.append(file_name_with_extension)


    return (
        bf_existing_files_name,
        bf_existing_files_name_with_extension,
    )


def check_if_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def get_base_file_name(file_name):
    output = []
    if file_name[-1] == ")":
        string_length = len(file_name)
        count_start = string_length
        character = file_name[count_start - 1]
        while character != "(" and count_start >= 0:
            count_start -= 1
            character = file_name[count_start - 1]
        if character == "(":
            base_name = file_name[:count_start - 1]
            num = file_name[count_start : string_length - 1]
            if check_if_int(num):
                output = [base_name, int(num)]
    return output


def ps_update_existing_dataset(soda_json_structure, ds, ps):
    global namespace_logger

    namespace_logger.info("Starting ps_update_existing_dataset")

    global main_curate_progress_message
    global main_total_generate_dataset_size
    global start_generate
    global main_initial_bfdataset_size

    # Delete any files on Pennsieve that have been marked as deleted
    def recursive_file_delete(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if "deleted" in folder["files"][item]["action"]:
                    file_path = folder["files"][item]["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [file_path]})
                    r.raise_for_status()
                    # remove the file from the soda json structure
                    del folder["files"][item]

        for item in list(folder["folders"]):
            recursive_file_delete(folder["folders"][item])
        return

    # Delete any files on Pennsieve that have been marked as deleted
    def metadata_file_delete(soda_json_structure):
        if "metadata-files" in soda_json_structure.keys():
            folder = soda_json_structure["metadata-files"]
            for item in list(folder):
                if "deleted" in folder[item]["action"]:
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [folder[item]["path"]]})
                    r.raise_for_status()
                    del folder[item]

        return

    def recursive_item_path_create(folder, path):
        """
        Recursively create the path for the item    # Add a new key containing the path to all the files and folders on the
        local data structure.
        Allows us to see if the folder path of a specfic file already
        exists on Pennsieve.
        """
        
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if item in ["manifest.xslx", "manifest.csv"]:
                    continue
                if "folderpath" not in folder["files"][item]:
                    namespace_logger.info(f"FOLDERPATH FOR FILE (path[:]) for recursive_item_path_create: {path[:]}")
                    folder["files"][item]["folderpath"] = path[:]

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                if "folderpath" not in folder["folders"][item]:
                    folder["folders"][item]["folderpath"] = path[:]
                    folder["folders"][item]["folderpath"].append(item)
                namespace_logger.info(f"FOLDERPATH FOR FOLDER (path[:]) for recursive_item_path_create: {folder['folders'][item]['folderpath'][:]}")
                recursive_item_path_create(
                    folder["folders"][item], folder["folders"][item]["folderpath"][:]
                )

        return

    # Check and create any non existing folders for the file move process (Used in the recursive_check_moved_files function)
    def recursive_check_and_create_ps_file_path(
        folderpath, index, current_folder_structure
    ):
        namespace_logger.info("Starting recursive_check_and_create_ps_file_path")
        namespace_logger.info(f"folderpath: {folderpath}")
        namespace_logger.info(f"current_folder_structure: {current_folder_structure}")
        folder = folderpath[index]

        if folder not in current_folder_structure["folders"]:
            if index == 0:
                r = requests.post(f"{PENNSIEVE_URL}/packages", json={"name": folder, "parent": f"{current_folder_structure['path']}", "packageType": "collection", "dataset": ds['content']['id']},  headers=create_request_headers(ps))
                r.raise_for_status()
                new_folder = r.json()
            else:
                r = requests.post(f"{PENNSIEVE_URL}/packages", json={"name": folder, "parent": f"{current_folder_structure['path']}", "packageType": "collection", "dataset": ds['content']['id']},  headers=create_request_headers(ps))
                r.raise_for_status()
                new_folder = r.json()
            
            current_folder_structure["folders"][folder] = {
                "type": "bf",
                "action": ["existing"],
                "path": new_folder['content']['id'],
                "folders": {},
                "files": {},
            }

        index += 1

        if index < len(folderpath):
            return recursive_check_and_create_ps_file_path(
                folderpath, index, current_folder_structure["folders"][folder]
            )
        else:
            return current_folder_structure["folders"][folder]["path"]

    # Check for any files that have been moved and verify paths before moving
    def recursive_check_moved_files(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if (
                    "moved" in folder["files"][item]["action"]
                    and folder["files"][item]["type"] == "bf"
                ):
                    # create the folders if they do not exist
                    new_folder_id = ""
                    new_folder_id = recursive_check_and_create_ps_file_path(
                        folder["files"][item]["folderpath"].copy(), 0, dataset_structure
                    )
                    # move the file into the target folder on Pennsieve
                    r = requests.post(f"{PENNSIEVE_URL}/data/move",  json={"things": [folder["files"][item]["path"]], "destination": new_folder_id}, headers=create_request_headers(ps),)
                    r.raise_for_status()

        for item in list(folder["folders"]):
            recursive_check_moved_files(folder["folders"][item])

        return

    # Rename any files that exist on Pennsieve
    def recursive_file_rename(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if (
                    "renamed" in folder["files"][item]["action"]
                    and folder["files"][item]["type"] == "bf"
                ):
                    # rename the file on Pennsieve
                    r = requests.put(f"{PENNSIEVE_URL}/packages/{folder['files'][item]['path']}?updateStorage=true", json={"name": item}, headers=create_request_headers(ps))
                    r.raise_for_status()

        for item in list(folder["folders"]):
            recursive_file_rename(folder["folders"][item])

        return

    def recursive_folder_delete(folder):
        """
        Delete any stray folders that exist on Pennsieve
        Only top level files are deleted since the api deletes any
        files and folders that exist inside.
        """
        for item in list(folder["folders"]):
            if folder["folders"][item]["type"] == "bf":
                if "moved" in folder["folders"][item]["action"]:
                    file_path = folder["folders"][item]["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [file_path]})
                    r.raise_for_status()
                if "deleted" in folder["folders"][item]["action"]:
                    file_path = folder["folders"][item]["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [file_path]})
                    r.raise_for_status()
                    del folder["folders"][item]
                else:
                    recursive_folder_delete(folder["folders"][item])
            else:
                recursive_folder_delete(folder["folders"][item])

        return

    # Rename any folders that still exist.
    def recursive_folder_rename(folder, mode):
        for item in list(folder["folders"]):
            if (
                folder["folders"][item]["type"] == "bf"
                and "action" in folder["folders"][item].keys()
                and mode in folder["folders"][item]["action"]
            ):
                folder_id = folder["folders"][item]["path"]
                r = requests.put(f"{PENNSIEVE_URL}/packages/{folder_id}?updateStorage=true", headers=create_request_headers(ps), json={"name": item})
                r.raise_for_status()
            recursive_folder_rename(folder["folders"][item], mode)

        return

    ps_dataset = ""
    start = timer()
    # 1. Remove all existing files on Pennsieve, that the user deleted.
    namespace_logger.info("ps_update_existing_dataset step 1 remove existing files on Pennsieve the user deleted")
    main_curate_progress_message = "Checking Pennsieve for deleted files"
    dataset_structure = soda_json_structure["dataset-structure"]
    recursive_file_delete(dataset_structure)
    main_curate_progress_message = (
        "Files on Pennsieve marked for deletion have been deleted"
    )

    # 2. Rename any deleted folders on Pennsieve to allow for replacements.
    namespace_logger.info("ps_update_existing_dataset step 2 rename deleted folders on Pennsieve to allow for replacements")
    main_curate_progress_message = "Checking Pennsieve for deleted folders"
    dataset_structure = soda_json_structure["dataset-structure"]
    recursive_folder_rename(dataset_structure, "deleted")
    main_curate_progress_message = "Folders on Pennsieve have been marked for deletion"

    # 2.5 Rename folders that need to be in the final destination.
    namespace_logger.info("ps_update_existing_dataset step 2.5 rename folders that need to be in the final destination")
    main_curate_progress_message = "Renaming any folders requested by the user"
    recursive_folder_rename(dataset_structure, "renamed")
    main_curate_progress_message = "Renamed all folders requested by the user"

    # 3. Get the status of all files currently on Pennsieve and create
    # the folderpath for all items in both dataset structures.
    namespace_logger.info("ps_update_existing_dataset step 3 get the status of all files currently on Pennsieve and create the folderpath for all items in both dataset structures")
    main_curate_progress_message = "Fetching files and folders from Pennsieve"
    current_bf_dataset_files_folders = import_pennsieve_dataset(
        soda_json_structure.copy()
    )["soda_object"]
    ps_dataset = current_bf_dataset_files_folders["dataset-structure"]
    main_curate_progress_message = "Creating file paths for all files on Pennsieve"
    recursive_item_path_create(dataset_structure, [])
    recursive_item_path_create(ps_dataset, [])
    main_curate_progress_message = "File paths created"

    # 4. Move any files that are marked as moved on Pennsieve.
    # Create any additional folders if required
    namespace_logger.info("ps_update_existing_dataset step 4 move any files that are marked as moved on Pennsieve")
    main_curate_progress_message = "Moving any files requested by the user"
    recursive_check_moved_files(dataset_structure)
    main_curate_progress_message = "Moved all files requested by the user"

    # 5. Rename any Pennsieve files that are marked as renamed.
    namespace_logger.info("ps_update_existing_dataset step 5 rename any Pennsieve files that are marked as renamed")
    main_curate_progress_message = "Renaming any files requested by the user"
    recursive_file_rename(dataset_structure)
    main_curate_progress_message = "Renamed all files requested by the user"

    # 6. Delete any Pennsieve folders that are marked as deleted.
    namespace_logger.info("ps_update_existing_dataset step 6 delete any Pennsieve folders that are marked as deleted")
    main_curate_progress_message = (
        "Deleting any additional folders present on Pennsieve"
    )
    recursive_folder_delete(dataset_structure)
    main_curate_progress_message = "Deletion of additional folders complete"

    # 7. Delete any metadata files that are marked as deleted.
    namespace_logger.info("ps_update_existing_dataset step 8 delete any metadata files that are marked as deleted")
    main_curate_progress_message = "Removing any metadata files marked for deletion"
    metadata_file_delete(soda_json_structure)
    main_curate_progress_message = "Removed metadata files marked for deletion"

    # 8. Run the original code to upload any new files added to the dataset.
    namespace_logger.info("ps_update_existing_dataset step 9 run the ps_create_new_dataset code to upload any new files added to the dataset")
    if "manifest-files" in soda_json_structure.keys():
        if "auto-generated" in soda_json_structure["manifest-files"].keys():
            soda_json_structure["manifest-files"] = {"destination": "bf", "auto-generated": True}
        else:
            soda_json_structure["manifest-files"] = {"destination": "bf"}


    namespace_logger.info(f"before modifying generate dataset: {soda_json_structure['generate-dataset']}")
    soda_json_structure["generate-dataset"] = {
        "destination": "bf",
        "if-existing": "merge",
        "if-existing-files": "replace",
        "generate-option": "existing-bf"
    }

    end = timer()
    namespace_logger.info(f"Time for ps_update_existing_dataset function: {timedelta(seconds=end - start)}")
    ps_upload_to_dataset(soda_json_structure, ps, ds)

    return


def normalize_tracking_folder(tracking_folder):
    """
    Normalize the tracking folder object to be a dictonary with the shape: {files: {}, folders: {}}. 
    This shape matches our dataset structure object. Recall, the tracking folder receives information about what folders and 
    files are stored on Pennsieve. We update this as we update Pennsieve's state. 
    """
    if tracking_folder == "":
        return {"folders": {}, "files": {} }
    else:
        temp_children = {"folders": {}, "files": {}}

        # add the files and folders to the temp_children structure 
        for child in tracking_folder["children"]:
            if child["content"]["packageType"] == "Collection":
                # add the folders ( designated collection on Pennsieve ) to the temp_children structure under folders
                temp_children["folders"][child["content"]["name"]] = child
            else:
                # add the files (anything not designated a collection) to the temp_children structure under files
                temp_children["files"][child["content"]["name"]] = child

        # replace the non-normalized children structure with the normalized children structure
        tracking_folder["children"] = temp_children


def build_create_folder_request(folder_name, folder_parent_id, dataset_id):
    """
    Create a folder on Pennsieve. 
    """
    body = {}

    # if creating a folder at the root of the dataset the api does not require a parent key
    if folder_parent_id.find("N:dataset") == -1:
        body["parent"] = folder_parent_id
    
    body["name"] = folder_name
    body["dataset"] = dataset_id
    body["packageType"] = "collection"

    return body


bytes_uploaded_per_file = {}
total_bytes_uploaded = {"value": 0}
current_files_in_subscriber_session = 0

bytes_uploaded_per_file = {}
total_bytes_uploaded = {"value": 0}
current_files_in_subscriber_session = 0

def ps_upload_to_dataset(soda_json_structure, ps, ds):
    global namespace_logger

    # Progress tracking variables that are used for the frontend progress bar.
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global start_generate
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global uploaded_folder_counter
    global current_size_of_uploaded_files
    global total_files
    global total_bytes_uploaded # current number of bytes uploaded to Pennsieve in the current session
    global client
    global files_uploaded
    global total_dataset_files
    global current_files_in_subscriber_session


    total_files = 0
    total_dataset_files = 0
    total_metadata_files = 0
    total_manifest_files = 0
    main_curation_uploaded_files = 0
    total_bytes_uploaded = {"value": 0}
    files_uploaded = 0

    uploaded_folder_counter = 0
    current_size_of_uploaded_files = 0
    start = timer()
    try:
        # Set the Pennsieve Python Client's dataset to the Pennsieve dataset that will be uploaded to.
        selected_id = ds["content"]["id"]
        ps.use_dataset(selected_id)
        
        def gather_manifest_files(soda_json_structure):
            """
            Gather the manifest files from the soda json structure.
            Output: A list of the file path of the manifest files.
            """
            global total_files
            global main_total_generate_dataset_size
            nonlocal total_manifest_files

            list_manifest_files = []
            if "auto-generated" in soda_json_structure["manifest-files"] and soda_json_structure["manifest-files"]["auto-generated"] == True:
                    manifest_files_structure = get_auto_generated_manifest_files(soda_json_structure)
                    for key in manifest_files_structure.keys():
                        manifestpath = manifest_files_structure[key]
                        list_manifest_files.append([manifestpath, key])
                        total_files += 1
                        total_manifest_files += 1
                        main_total_generate_dataset_size += getsize(manifestpath)
            
            namespace_logger.info(f"list_manifest_files: {list_manifest_files}")
            return list_manifest_files
        
        def gather_metadata_files(soda_json_structure, ):
            """
            Gather the metadata files from the soda json structure.
            Output: A list of the file path of the metadata files.
            """
            global main_total_generate_dataset_size
            global total_files
            nonlocal total_metadata_files

            metadata_files = soda_json_structure["metadata-files"]
            metadata_files_list = []
            for file_key, file in metadata_files.items():
                if file["type"] == "local":
                    metadata_path = file["path"]
                    if isfile(metadata_path):
                        metadata_files_list.append(metadata_path)
                        main_total_generate_dataset_size += getsize(metadata_path)
                        total_files += 1
                        total_metadata_files += 1

            return metadata_files_list

        def recursive_dataset_scan_for_new_upload(dataset_structure, list_upload_files, my_relative_path):
            """
            This function recursively gathers the files and folders in the dataset that will be uploaded to Pennsieve.
            It assumes the dataset is new based on the generate_option value and will spend less time comparing what is on Pennsieve.
            It will gather all the relative paths for the files and folders to pass along to the Pennsieve agent.
            Input:
            dataset_structure,
            my_relative_path

            Output:
            two lists in one tuple, the first list will have all the local file paths that will be uploaded to Pennsieve
            The second list will have the relative files paths according to the dataset structure.
            If the folder does not existing yet on Pennsieve the agent will create it.
            """
            print("recursive_dataset_scan_for_upload")
            global main_total_generate_dataset_size
            list_of_local_file_paths = []
            # First loop will take place in the root of the dataset
            if "folders" in dataset_structure.keys():
                for folder_key, folder in dataset_structure["folders"].items():
                    relative_path = generate_relative_path(my_relative_path, folder_key)
                    list_upload_files = recursive_dataset_scan_for_new_upload(folder, list_upload_files, relative_path)
            if "files" in dataset_structure.keys():
                list_local_files = []
                list_projected_names = []
                list_desired_names = []
                list_final_names = []
                additional_upload_lists = []

                list_initial_names = []
                for file_key, file in dataset_structure["files"].items():
                    # relative_path = generate_relative_path(my_relative_path, file_key)
                    file_path = file["path"]
                    if isfile(file_path) and file["type"] == "local":
                        projected_name = splitext(basename(file_path))[0]
                        projected_extension = splitext(basename(file_path))[1]
                        projected_name_w_extension = basename(file_path)
                        desired_name = splitext(file_key)[0]
                        desired_extension = splitext(file_key)[1]
                        desired_name_with_extension = file_key
                        

                        if projected_name != desired_name:
                            list_initial_names.append(projected_name)
                            list_local_files.append(file_path)
                            list_projected_names.append(projected_name_w_extension)
                            list_desired_names.append(desired_name_with_extension)
                            list_final_names.append(desired_name)
                        else:
                            list_local_files.append(file_path)
                            list_projected_names.append(projected_name_w_extension)
                            list_desired_names.append(desired_name_with_extension)
                            list_final_names.append(desired_name)
                            list_initial_names.append(projected_name)

                        main_total_generate_dataset_size += getsize(file_path)

                if list_local_files:
                    list_upload_files.append([
                        list_local_files,
                        list_projected_names,
                        list_desired_names,
                        list_final_names,
                        my_relative_path
                    ])

            return list_upload_files
        
        # See how to create folders with the Pennsieve agent
        def recursive_create_folder_for_ps(
            my_folder, my_tracking_folder, existing_folder_option
        ):
            """
            Creates a folder on Pennsieve for each folder in the dataset structure if they aren't already present in the dataset.
            Input:
                my_folder: The dataset structure to be created on Pennsieve. Pass in the soda json object to start. 
                my_tracking_folder: Tracks what folders have been created on Pennsieve thus far. Starts as an empty dictionary.
                existing_folder_option: Dictates whether to merge, duplicate, replace, or skip existing folders.
            """

            # Check if the current folder has any subfolders that already exist on Pennsieve. Important step to appropriately handle replacing and merging folders.
            if len(my_tracking_folder["children"]["folders"]) == 0 and my_tracking_folder["content"]["id"].find("N:dataset") == -1:
                r = requests.get(f"{PENNSIEVE_URL}/packages/{my_tracking_folder['content']['id']}", headers=create_request_headers(ps), json={"include": "files"})
                r.raise_for_status()
                ps_folder = r.json()
                normalize_tracking_folder(ps_folder)
                my_tracking_folder["children"] = ps_folder["children"]

            # create/replace/skip folder
            if "folders" in my_folder.keys():
                for folder_key, folder in my_folder["folders"].items():
                    if existing_folder_option == "skip":
                        if folder_key not in my_tracking_folder["children"]["folders"]:
                            r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(ps), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                            r.raise_for_status()
                            ps_folder = r.json()
                            normalize_tracking_folder(ps_folder)
                        else:
                            ps_folder = my_tracking_folder["children"]["folders"][folder_key]
                            normalize_tracking_folder(ps_folder)

                    elif existing_folder_option == "create-duplicate":
                        r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(ps), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                        r.raise_for_status()
                        ps_folder = r.json()
                        normalize_tracking_folder(ps_folder)

                    elif existing_folder_option == "replace":
                        # if the folder exists on Pennsieve remove it
                        if folder_key in my_tracking_folder["children"]["folders"]:
                            ps_folder = my_tracking_folder["children"]["folders"][folder_key]

                            r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [ps_folder["content"]["id"]]})
                            r.raise_for_status()

                            # remove from ps_folder 
                            del my_tracking_folder["children"]["folders"][folder_key]

                        r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(ps), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                        r.raise_for_status()
                        ps_folder = r.json()
                        normalize_tracking_folder(ps_folder)

                    elif existing_folder_option == "merge":
                        if folder_key in my_tracking_folder["children"]["folders"]:
                            ps_folder = my_tracking_folder["children"]["folders"][folder_key]
                            normalize_tracking_folder(ps_folder)
                        else:
                            # We are merging but this is a new folder - not one that already exists in the current dataset - so we create it.
                            r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(ps), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                            r.raise_for_status()
                            ps_folder = r.json()
                            normalize_tracking_folder(ps_folder)


                    my_tracking_folder["children"]["folders"][folder_key] = ps_folder
                    tracking_folder = my_tracking_folder["children"]["folders"][folder_key] # get the folder we just added to the tracking folder
                    recursive_create_folder_for_ps(
                        folder, tracking_folder, existing_folder_option
                    )
        
        def recursive_dataset_scan_for_ps(
            my_folder,
            my_tracking_folder,
            existing_file_option,
            list_upload_files,
            my_relative_path,
        ):
            """
                Delete files that are marked to be replaced in the dataset. Create a list of files to upload to Pennsieve.
            """

            global main_total_generate_dataset_size


            # folder children are packages such as collections and files stored on the Pennsieve dataset
            ps_folder_children = my_tracking_folder["children"] #ds (dataset)



            if "folders" in my_folder.keys():
                for folder_key, folder in my_folder["folders"].items():
                    relative_path = generate_relative_path(my_relative_path, folder_key)

                    if existing_folder_option == "skip" and folder_key in my_tracking_folder["children"]["folders"]:
                        continue

                    tracking_folder = ps_folder_children["folders"][folder_key]
                    list_upload_files = recursive_dataset_scan_for_ps(
                        folder,
                        tracking_folder,
                        existing_file_option,
                        list_upload_files,
                        relative_path,
                    )

            if "files" in my_folder.keys() and my_tracking_folder["content"]["id"].find("N:dataset") == -1: 

                # delete files to be deleted
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_tracking_folder, ps)

                for file_key, file in my_folder["files"].items():
                    # if local then we are either adding a new file to an existing/new dataset or replacing a file in an existing dataset
                    if file["type"] == "local":
                        file_path = file["path"]
                        if isfile(file_path) and existing_file_option == "replace" and file_key in ps_folder_children["files"]:
                            my_file = ps_folder_children["files"][file_key]
                            # delete the package ( aka file ) from the dataset 
                            r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(ps), json={"things": [f"{my_file['content']['id']}"]})
                            r.raise_for_status()
                            del ps_folder_children["files"][file_key]


                # create list of files to be uploaded with projected and desired names saved
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_tracking_folder, ps)

                list_local_files = []
                list_projected_names = []
                list_desired_names = []
                list_final_names = []
                additional_upload_lists = []

                list_initial_names = []

                # add the files that are set to be uploaded to Pennsieve to a list 
                # handle renaming files and creating duplicates
                for file_key, file in my_folder["files"].items():
                    if file["type"] == "local":
                        file_path = file["path"]
                        if isfile(file_path):
                            initial_name = splitext(basename(file_path))[0]
                            initial_extension = splitext(basename(file_path))[1]
                            initial_name_with_extension = basename(file_path)
                            desired_name = splitext(file_key)[0]
                            desired_name_extension = splitext(file_key)[1]
                            desired_name_with_extension = file_key
                            if existing_file_option == "skip" and desired_name_with_extension in my_bf_existing_files_name_with_extension:
                                continue

                            # check if initial filename exists on Pennsieve dataset and get the projected name of the file after upload
                            # used when a local file has a name that matches an existing name on Pennsieve
                            count_done = 0
                            count_exist = 0
                            projected_name = initial_name_with_extension
                            while count_done == 0:
                                if (
                                    projected_name
                                    in my_bf_existing_files_name_with_extension
                                ):
                                    count_exist += 1
                                    projected_name = (
                                        initial_name
                                        + " ("
                                        + str(count_exist)
                                        + ")"
                                        + initial_extension
                                    )
                                else:
                                    count_done = 1

                            # expected final name
                            count_done = 0
                            final_name = desired_name_with_extension
                            if output := get_base_file_name(desired_name):
                                base_name = output[0]
                                count_exist = output[1]
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            base_name
                                            + "("
                                            + str(count_exist)
                                            + ")"
                                            + desired_name_extension
                                        )
                                    else:
                                        count_done = 1
                            else:
                                count_exist = 0
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            desired_name
                                            + " ("
                                            + str(count_exist)
                                            + ")"
                                            + desired_name_extension
                                        )
                                    else:
                                        count_done = 1

                            # save in list accordingly
                            if (
                                initial_name in list_initial_names
                                or initial_name in list_final_names
                                or projected_name in list_final_names
                                or final_name in list_projected_names
                            ):
                                additional_upload_lists.append(
                                    [
                                        [file_path],
                                        ps_folder_children,
                                        [projected_name],
                                        [desired_name],
                                        [final_name],
                                        my_tracking_folder,
                                        my_relative_path,
                                    ]
                                )
                            else:
                                list_local_files.append(file_path)
                                list_projected_names.append(projected_name)
                                list_desired_names.append(desired_name_with_extension)
                                list_final_names.append(final_name)
                                list_initial_names.append(initial_name)

                            my_bf_existing_files_name.append(final_name)
                            if initial_extension in ps_recognized_file_extensions:
                                my_bf_existing_files_name_with_extension.append(
                                    final_name
                                )
                            else:
                                my_bf_existing_files_name_with_extension.append(
                                    final_name + initial_extension
                                )

                            # add to projected dataset size to be generated
                            main_total_generate_dataset_size += getsize(file_path)

                if list_local_files:
                    list_upload_files.append(
                        [
                            list_local_files,
                            ps_folder_children,
                            list_projected_names,
                            list_desired_names,
                            list_final_names,
                            my_tracking_folder,
                            my_relative_path,
                        ]
                    )

                for item in additional_upload_lists:
                    list_upload_files.append(item)

            return list_upload_files

        def monitor_subscriber_progress(events_dict):
            """
            Monitors the progress of a subscriber and unsubscribes once the upload finishes. 
            """
            global files_uploaded
            global total_bytes_uploaded
            global current_files_in_subscriber_session
            global bytes_uploaded_per_file
            global main_curation_uploaded_files

            if events_dict["type"] == 1:  # upload status: file_id, total, current, worker_id
                file_id = events_dict["upload_status"].file_id
                total_bytes_to_upload = events_dict["upload_status"].total
                current_bytes_uploaded = events_dict["upload_status"].current

                # get the previous bytes uploaded for the given file id - use 0 if no bytes have been uploaded for this file id yet
                previous_bytes_uploaded = bytes_uploaded_per_file.get(file_id, 0)

                # update the file id's current total bytes uploaded value 
                bytes_uploaded_per_file[file_id] = current_bytes_uploaded

                # sometimes a user uploads the same file to multiple locations in the same session. Edge case. Handle it by resetting the value to 0 if it is equivalent to the 
                # total bytes for that file 
                if previous_bytes_uploaded == total_bytes_to_upload:
                    previous_bytes_uploaded = 0 


                # calculate the additional amount of bytes that have just been uploaded for the given file id
                total_bytes_uploaded["value"] += current_bytes_uploaded - previous_bytes_uploaded

                # check if the given file has finished uploading
                if current_bytes_uploaded == total_bytes_to_upload and  file_id != "":
                    files_uploaded += 1
                    main_curation_uploaded_files += 1


                # check if the upload has finished
                if files_uploaded == current_files_in_subscriber_session:
                    namespace_logger.info("Upload complete")
                    # unsubscribe from the agent's upload messages since the upload has finished
                    ps.unsubscribe(10)

        # Set variables needed throughout generation flow
        list_upload_files = []
        list_upload_metadata_files = []
        list_upload_manifest_files = []
        list_of_files_to_rename = {}
        brand_new_dataset = False
        dataset_structure = soda_json_structure["dataset-structure"]
        generate_option = soda_json_structure["generate-dataset"]["generate-option"]
        starting_point = soda_json_structure["starting-point"]["type"]
        relative_path = ds["content"]["name"]
        existing_folder_option = soda_json_structure["generate-dataset"]["if-existing"]
        existing_file_option = soda_json_structure["generate-dataset"][
            "if-existing-files"
        ]
             
        main_curate_progress_message = "Preparing a list of files to upload"
        # 1. Scan the dataset structure and create a list of files/folders to be uploaded with the desired renaming
        if generate_option == "new" and starting_point == "new":
            # we can assume no files/folders exist in the dataset since the generate option is new and starting point is also new
            # therefore, we can assume the dataset structure is the same as the tracking structure
            brand_new_dataset = True
            list_upload_files = recursive_dataset_scan_for_new_upload(dataset_structure, list_upload_files, relative_path)

            if "metadata-files" in soda_json_structure.keys():
                list_upload_metadata_files = gather_metadata_files(soda_json_structure)
            
            if "manifest-files" in soda_json_structure.keys():
                list_upload_manifest_files = gather_manifest_files(soda_json_structure)
                
        else:
            # we will need a tracking structure to compare against
            tracking_json_structure = ds
            normalize_tracking_folder(tracking_json_structure)
            recursive_create_folder_for_ps(dataset_structure, tracking_json_structure, existing_folder_option)
            list_upload_files = recursive_dataset_scan_for_ps(
                dataset_structure,
                tracking_json_structure,
                existing_file_option,
                list_upload_files,
                relative_path,
            )
        
            # 3. Add high-level metadata files to a list
            if "metadata-files" in soda_json_structure.keys():
                namespace_logger.info("ps_create_new_dataset (optional) step 3 create high level metadata list")
                (
                    my_bf_existing_files_name,
                    _,
                ) = ps_get_existing_files_details(ds, ps)
                metadata_files = soda_json_structure["metadata-files"]
                for file_key, file in metadata_files.items():
                    if file["type"] == "local":
                        metadata_path = file["path"]
                        if isfile(metadata_path):
                            initial_name = splitext(basename(metadata_path))[0]
                            if (
                                existing_file_option == "replace"
                                and initial_name in my_bf_existing_files_name
                            ):
                                my_file = ds['children']['files'][file_key]
                                # delete the file from Pennsieve
                                r = requests.post(f"{PENNSIEVE_URL}/data/delete", json={"things": [my_file['content']['id']]}, headers=create_request_headers(get_access_token()))
                                r.raise_for_status()
                            if (
                                existing_file_option == "skip"
                                and initial_name in my_bf_existing_files_name
                            ):
                                continue
                            
                            list_upload_metadata_files.append(metadata_path)
                            main_total_generate_dataset_size += getsize(metadata_path)
                            total_files += 1
                            total_metadata_files += 1

            # 4. Prepare and add manifest files to a list
            if "manifest-files" in soda_json_structure.keys():
                namespace_logger.info("ps_create_new_dataset (optional) step 4 create manifest list")
                # create local folder to save manifest files temporarly (delete any existing one first)
                if "auto-generated" in soda_json_structure["manifest-files"]:
                    if soda_json_structure["manifest-files"]["auto-generated"] == True:
                        manifest_files_structure = (
                            get_auto_generated_manifest_files(soda_json_structure)
                        )
                # TODO: Dorian -> Verify if this function is still needed
                else:
                    # prepare manifest files
                    if soda_json_structure["starting-point"]["type"] == "bf":
                        # get auto generated manifest file
                        manifest_files_structure = (
                            create_high_lvl_manifest_files_existing_ps_starting_point(
                                soda_json_structure, 
                                manifest_folder_path
                            )
                        )
                    else:
                        if (
                            soda_json_structure["generate-dataset"]["destination"] == "bf"
                            and "dataset-name" not in soda_json_structure["generate-dataset"]
                        ):
                            # generating dataset on an existing bf dataset - account for existing files and manifest files
                            # get auto generated manifest file
                            manifest_files_structure = (
                                create_high_lvl_manifest_files_existing_ps(
                                    soda_json_structure, ps, tracking_json_structure
                                )
                            )
                        else:
                            manifest_files_structure = create_high_level_manifest_files(
                                soda_json_structure, manifest_folder_path
                            )

                # add manifest files to list after deleting existing ones
                for key in manifest_files_structure.keys():
                    manifestpath = manifest_files_structure[key]
                    folder = tracking_json_structure["children"]["folders"][key]

                    # delete existing manifest files
                    for child_key in folder["children"]["files"]:
                        file_name_no_ext = os.path.splitext(folder['children']['files'][child_key]['content']['name'])[0]
                        if file_name_no_ext.lower() == "manifest":
                            # delete the manifest file from the given folder 
                            r = requests.post(f"{PENNSIEVE_URL}/data/delete", json={"things": [folder['children']['files'][child_key]['content']['id']]}, headers=create_request_headers(get_access_token()))
                            r.raise_for_status()

                    # upload new manifest files
                    # the number of files to upload and the total also determines when the upload subscribers should stop listening to the dataset upload progress ( when files uploaded == total files stop listening )
                    list_upload_manifest_files.append([manifestpath, key])
                    total_files += 1
                    total_manifest_files += 1
                    main_total_generate_dataset_size += getsize(manifestpath)
        
        
        # 2. Count how many files will be uploaded to inform frontend
        for folderInformation in list_upload_files:
            file_paths_count = len(folderInformation[0])
            total_files += file_paths_count
            total_dataset_files += file_paths_count

        # 3. Upload files and add to tracking list
        start_generate = 1
        main_curate_progress_message = ("Queuing dataset files for upload with the Pennsieve Agent..." + "<br>" + "This may take some time.")

        # create a manifest for files - IMP: We use a single file to start with since creating a manifest requires a file path.  We need to remove this at the end. 
        if len(list_upload_files) > 0:
            list_of_files_to_renamed = {}
            first_file_local_path = list_upload_files[0][0][0]
            
            if brand_new_dataset:
                first_relative_path = list_upload_files[0][4]
                first_final_name = list_upload_files[0][2][0]
            else:
                first_relative_path = list_upload_files[0][6]
                first_final_name = list_upload_files[0][4][0]

            folder_name = first_relative_path[first_relative_path.index("/")+1:]

            if first_final_name != basename(first_file_local_path):
                # if file name is not the same as local path, then it has been renamed in SODA
                if folder_name not in list_of_files_to_rename:
                    list_of_files_to_rename[folder_name] = {}
                if basename(first_file_local_path) not in list_of_files_to_rename[folder_name]:
                    list_of_files_to_rename[folder_name][basename(first_file_local_path)] = {
                        "final_file_name": first_final_name,
                        "id": "",
                    },

            manifest_data = ps.manifest.create(first_file_local_path, folder_name)
            manifest_id = manifest_data.manifest_id

            # remove the item just added to the manifest 
            list_upload_files[0][0].pop(0)

            loc = get_agent_installation_location()

            # there are files to add to the manifest if there are more than one file in the first folder or more than one folder
            if len(list_upload_files[0][0]) > 1 or len(list_upload_files) > 1:
                index_skip = True
                index_skip = True
                for folderInformation in list_upload_files:
                    list_file_paths = folderInformation[0]
                    if brand_new_dataset:
                        relative_path = folderInformation[4]
                        final_file_name_list = folderInformation[2]
                    else:
                        relative_path = folderInformation[6]
                        final_file_name_list = folderInformation[4]
                    # get the substring from the string relative_path that starts at the index of the / and contains the rest of the string
                    try:
                        folder_name = relative_path[relative_path.index("/")+1:]
                    except ValueError as e:
                        folder_name = relative_path

                    # Add files to manfiest"
                    final_files_index = 1 if index_skip else 0
                    index_skip = False
                    namespace_logger.info(f"files folder names: {folder_name}")
                    for file_path in list_file_paths:
                        file_file_name = final_file_name_list[final_files_index]
                        if file_file_name != basename(file_path):
                            # save the relative path, final name and local path of the file to be renamed
                            if folder_name not in list_of_files_to_rename:
                                list_of_files_to_rename[folder_name] = {}
                            if basename(file_path) not in list_of_files_to_rename[folder_name]:
                                list_of_files_to_rename[folder_name][basename(file_path)] = {
                                    "final_file_name": file_file_name,
                                    "id": "",
                                }
                        ps.manifest.add(file_path, folder_name, manifest_id)
                        final_files_index += 1

            # reset global variables used in the subscriber monitoring function
            bytes_uploaded_per_file = {}
            file_uploaded = 0
            total_bytes_uploaded = {"value": 0}
            current_files_in_subscriber_session = total_dataset_files


            # upload the manifest files
            try: 
                ps.manifest.upload(manifest_id)

                main_curate_progress_message = ("Uploading data files...")

                # subscribe to the manifest upload so we wait until it has finished uploading before moving on
                ps.subscribe(10, False, monitor_subscriber_progress)
            except Exception as e:
                namespace_logger.error("Error uploading dataset files")
                namespace_logger.error(e)
                raise Exception("The Pennsieve Agent has encountered an issue while uploading. Please retry the upload. If this issue persists please follow this <a href='https://docs.sodaforsparc.io/docs/how-to/how-to-reinstall-the-pennsieve-agent'> guide</a> on performing a full reinstallation of the Pennsieve Agent to fix the problem.")

        # 4. Upload metadata files
        if list_upload_metadata_files:
            namespace_logger.info("ps_create_new_dataset (optional) step 6 upload metadata files")
            main_curate_progress_message = (f"Uploading metadata files in high-level dataset folder {str(ds['content']['name'])}")

            # create the manifest 
            manifest_data = ps.manifest.create(list_upload_metadata_files[0])
            manifest_id = manifest_data.manifest_id

            loc = get_agent_installation_location()

            # add the files to the manifest
            for manifest_path in list_upload_metadata_files[1:]:
                # subprocess call to the pennsieve agent to add the files to the manifest
                ps.manifest.add(manifest_path, target_base_path="", manifest_id=manifest_id)

            # reset global variables used in the subscriber monitoring function
            bytes_uploaded_per_file = {}
            current_files_in_subscriber_session = total_metadata_files
            files_uploaded = 0

            # upload the manifest 
            try:
                ps.manifest.upload(manifest_id)

                # subscribe to the manifest upload so we wait until it has finished uploading before moving on
                ps.subscribe(10, False, monitor_subscriber_progress)
            except Exception as e:
                namespace_logger.error("Error uploading metadata files")
                namespace_logger.error(e)
                raise Exception("The Pennsieve Agent has encountered an issue while uploading. Please retry the upload. If this issue persists please follow this <a href='https://docs.sodaforsparc.io/docs/how-to/how-to-reinstall-the-pennsieve-agent'> guide</a> on performing a full reinstallation of the Pennsieve Agent to fix the problem.")

        # 5. Upload manifest files
        if list_upload_manifest_files:
            namespace_logger.info("ps_create_new_dataset (optional) step 7 upload manifest files")

            ps_folder = list_upload_manifest_files[0][1]
            manifest_data = ps.manifest.create(list_upload_manifest_files[0][0], ps_folder)
            manifest_id = manifest_data.manifest_id

            loc = get_agent_installation_location()

            if len(list_upload_manifest_files) > 1:
                for item in list_upload_manifest_files[1:]:
                    manifest_file = item[0]
                    ps_folder = item[1]
                    main_curate_progress_message = ( f"Uploading manifest file in {ps_folder} folder" )

                    # add the files to the manifest
                    ps.manifest.add(manifest_file, ps_folder, manifest_id)

            # reset global variable used in the subscriber monitoring function
            bytes_uploaded_per_file = {}
            current_files_in_subscriber_session = total_manifest_files
            files_uploaded = 0

            try: 
                # upload the manifest 
                ps.manifest.upload(manifest_id)

                ps.subscribe(10, False, monitor_subscriber_progress)
            except Exception as e:
                namespace_logger.error("Error uploading manifest files")
                namespace_logger.error(e)
                raise Exception("The Pennsieve Agent has encountered an issue while uploading. Please retry the upload. If this issue persists please follow this <a href='https://docs.sodaforsparc.io/docs/how-to/how-to-reinstall-the-pennsieve-agent'> guide</a> on performing a full reinstallation of the Pennsieve Agent to fix the problem.")

        # wait for all of the Agent's processes to finish to avoid errors when deleting files on Windows
        time.sleep(1)

        # 6. Rename files
        namespace_logger.info("===========================BEGINNING OF RENAME FILES===========================")
        if list_of_files_to_rename:
            namespace_logger.info("ps_create_new_dataset (optional) step 8 rename files")
            main_curate_progress_message = ("Renaming files...")
            dataset_id = ds["content"]["id"]
            collection_ids = {}
            # gets the high level folders in the dataset
            r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(ps))
            r.raise_for_status()
            dataset_content = r.json()["children"]

            namespace_logger.info(f"FIRST INITIAL FETCH: {dataset_content}")

            if dataset_content == []:
                while dataset_content == []:
                    r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(ps))
                    r.raise_for_status()
                    dataset_content = r.json()["children"]

            collections_found = False
            while not collections_found:
                for item in dataset_content:
                    # high lvl folders' ids are stored to be used to find the file IDS
                    if item["content"]["packageType"] == "Collection":
                        collections_found = True
                        collection_ids[item["content"]["name"]] = {
                            "id": item["content"]["nodeId"],
                        }

                if not collections_found:
                    # No collections were found, metadata files were processed but not the high level folders
                    r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(ps))
                    r.raise_for_status()
                    dataset_content = r.json()["children"]

            for key in list_of_files_to_rename.keys():
                # split the key up if there are multiple folders in the relative path
                relative_path = key.split("/")
                high_lvl_folder_name = relative_path[0]
                last_folder_name = relative_path[-1]
                subfolder_level = 0
                subfolder_amount = len(relative_path) - 1

                namespace_logger.info(f"subfolder_amount FOR {key}: {subfolder_amount}")
                namespace_logger.info(f"subfolder_level FOR {key}: {subfolder_level}")
                namespace_logger.info(f"relative_path list FOR {key}: {relative_path}")
                namespace_logger.info(f"last folder needed FOR {key}: {last_folder_name}")

                if high_lvl_folder_name in collection_ids:
                    # subfolder_amount will be the amount of subfolders we need to call until we can get the file ID to rename
                    high_lvl_folder_id = collection_ids[high_lvl_folder_name]["id"]
                    r = requests.get(f"{PENNSIEVE_URL}/packages/{high_lvl_folder_id}", headers=create_request_headers(ps))
                    r.raise_for_status()
                    dataset_content = r.json()["children"]

                    if dataset_content == []:
                        # request until there is no children content, (folder is empty so files have not been processed yet)
                        while dataset_content == []:
                            r = requests.get(f"{PENNSIEVE_URL}/packages/{high_lvl_folder_id}", headers=create_request_headers(ps))
                            r.raise_for_status()
                            dataset_content = r.json()["children"]
                    
                    if subfolder_amount == 0:
                        # the file is in the high level folder
                        if "id" not in list_of_files_to_rename[key]:
                            # store the id of the folder to be used again in case the file id is not found (happens when not all files have been processed yet)
                            list_of_files_to_rename[key]["id"] = high_lvl_folder_id

                        
                        for item in dataset_content:
                            if item["content"]["packageType"] != "Collection":
                                file_name = item["content"]["name"]
                                file_id = item["content"]["nodeId"]

                                if file_name in list_of_files_to_rename[key]:
                                    # store the package id for now
                                    list_of_files_to_rename[key][file_name]["id"] = file_id
                    else:
                        # file is within a subfolder and we recursively iterate until we get to the last subfolder needed
                        subfolder_id = collection_ids[high_lvl_folder_name]["id"]
                        while subfolder_level != subfolder_amount:
                            if dataset_content == []:
                                # subfolder has no content so request again
                                while dataset_content == []:
                                    r = requests.get(f"{PENNSIEVE_URL}/packages/{subfolder_id}", headers=create_request_headers(ps))
                                    r.raise_for_status()
                                    dataset_content = r.json()["children"]

                            for item in dataset_content:
                                if item["content"]["packageType"] == "Collection":
                                    folder_name = item["content"]["name"]
                                    folder_id = item["content"]["nodeId"]

                                    if folder_name in relative_path:
                                        # we have found the folder we need to iterate through
                                        subfolder_level += 1
                                        namespace_logger.info(f"folder_name we are gonna get content from (in relative path): {folder_name}")
                                        namespace_logger.info(f"relative_path: {relative_path}, we need folder level: {subfolder_level}")
                                        namespace_logger.info(f"folder id for {folder_name}: {folder_id}")
                                        r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}", headers=create_request_headers(ps))
                                        r.raise_for_status()

                                        if subfolder_level != subfolder_amount:
                                            dataset_content = r.json()["children"]
                                            if dataset_content == []:
                                                while dataset_content == []:
                                                    # subfolder has no content so request again
                                                    r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}", headers=create_request_headers(ps))
                                                    r.raise_for_status()
                                                    dataset_content = r.json()["children"]
                                            namespace_logger.info(f"subfolder_amount: {subfolder_amount}")
                                            namespace_logger.info(f"subfolder_level: {subfolder_level}")
                                            namespace_logger.info(f"subfolder id for {folder_name}: {subfolder_id}")
                                            subfolder_id = folder_id
                                            break
                                        else:
                                            # we are at the last folder in the relative path, we can get the file id
                                            if "id" not in list_of_files_to_rename[key]:
                                                # store the id of the last folder to directly call later in case not all files get an id
                                                list_of_files_to_rename[key]["id"] = folder_id
                                            for item in r.json()["children"]:
                                                if item["content"]["packageType"] != "Collection":
                                                    file_name = item["content"]["name"]
                                                    file_id = item["content"]["nodeId"]

                                                    if file_name in list_of_files_to_rename[key]:
                                                        # store the package id for renaming
                                                        list_of_files_to_rename[key][file_name]["id"] = file_id
                                    else:
                                        continue

            # 8.5 Rename files - All or most ids have been fetched now rename the files or gather the ids again if not all files have been processed at this time
            namespace_logger.info(f"list of files to rename after: {list_of_files_to_rename}")
            for relative_path in list_of_files_to_rename.keys():
                for file in list_of_files_to_rename[relative_path].keys():
                    collection_id = list_of_files_to_rename[relative_path]["id"]
                    if file == "id":
                        continue
                    new_name = list_of_files_to_rename[relative_path][file]["final_file_name"]
                    file_id = list_of_files_to_rename[relative_path][file]["id"]
                    
                    if file_id != "":
                        # id was found so make api call to rename with final file name
                        r = requests.put(f"{PENNSIEVE_URL}/packages/{file_id}?updateStorage=true", json={"name": new_name}, headers=create_request_headers(ps))
                        r.raise_for_status()
                        if r.status_code == 500:
                            continue
                    else:
                        # id was not found so keep trying to get the id until it is found
                        all_ids_found = False
                        while not all_ids_found:
                            collection_id = list_of_files_to_rename[relative_path]["id"]
                            if file == "id":
                                continue

                            r = requests.put(f"{PENNSIEVE_URL}/packages/{collection_id}?updateStorage=true", headers=create_request_headers(ps))
                            r.raise_for_status()
                            dataset_content = r.json()["children"]
                            for item in dataset_content:
                                if item["content"]["packageType"] != "Collection":
                                    file_name = item["content"]["name"]
                                    file_id = item["content"]["nodeId"]

                                    if file_name == file:
                                        # id was found so make api call to rename with file file name
                                        r = requests.put(f"{PENNSIEVE_URL}/packages/{file_id}", json={"name": new_name}, headers=create_request_headers(ps))
                                        r.raise_for_status()
                                        all_ids_found = True
                                        break

        namespace_logger.info("===========================END OF RENAME FILES===========================")
        shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
        end = timer()
        namespace_logger.info(f"Time for ps_upload_to_dataset function: {timedelta(seconds=end - start)}")
    except Exception as e:
        raise e

main_curate_status = ""
main_curate_print_status = ""
main_curate_progress_message = ""
main_total_generate_dataset_size = 1
main_generated_dataset_size = 0
start_generate = 0
generate_start_time = 0
main_generate_destination = ""
main_initial_bfdataset_size = 0
myds = ""



def handle_duplicate_package_name_error(e, soda_json_structure):
    if e.response.text == '{"type":"BadRequest","message":"package name must be unique","code":400}':
        if "if-existing-files" in soda_json_structure["generate-dataset"]:
            if soda_json_structure["generate-dataset"]["if-existing-files"] == "create-duplicate":
                return

    raise e

def ps_check_dataset_files_validity(soda_json_structure, ps):
    """
    Function to check that the bf data files and folders specified in the dataset are valid

    Args:
        dataset_structure: soda dict with information about all specified files and folders
        ps: pennsieve http object
    Output:
        error: error message with list of non valid local data files, if any
    """
    def check_folder_validity(folder_id, folder_dict, folder_path, error):
        """
        Function to verify that the subfolders and files specified in the dataset are valid

        Args:
            folder_id: id of the folder in the dataset
            folder_dict: dict with information about the folder
            folder_path: path of the folder in the dataset
            error: error message with list of non valid files/folders, if any
        Output:
            error: error message with list of non valid files/folders, if any
        """
        global PENNSIEVE_URL
        # get the folder content through Pennsieve api
        r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}", headers=create_request_headers(ps))
        r.raise_for_status()
        folder_content = r.json()["children"]

        # check that the subfolders and files specified in the dataset are valid
        if "files" in folder_dict.keys():
            for file_key, file in folder_dict["files"].items():
                file_type = file["type"]
                relative_path = (f"{folder_path}/{file_key}")
                # If file is from Pennsieve we verify if file exists on Pennsieve
                if file_type == "bf":
                    file_actions = file["action"]
                    file_id = file["path"]
                    if "moved" in file_actions:
                        try:
                            r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/view", headers=create_request_headers(ps))
                            r.raise_for_status()
                        except Exception as e:
                            error.append(f"{relative_path} id: {file_id}")
                        continue
                    if next((item for item in folder_content if item["content"]["id"] == file_id), None) is None:
                        error.append(f"{relative_path} id: {file_id}")
        
        if "folders" in folder_dict.keys():
            for folder_key, folder in folder_dict["folders"].items():
                folder_type = folder["type"]
                relative_path = (f"{folder_path}/{folder_key}")
                if folder_type == "bf":
                    folder_id = folder["path"]
                    folder_action = folder["action"]
                    if "moved" in folder_action:
                        try:
                            r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}", headers=create_request_headers(ps))
                            r.raise_for_status()
                        except Exception as e:
                            error.append(f"{relative_path} id: {folder_id}")
                        continue
                    if next((item for item in folder_content if item["content"]["id"] == folder_id), None) is None:
                        error.append(f"{relative_path} id: {folder_id}")
                    else:
                        check_folder_validity(folder_id, folder, relative_path, error)

        return error

    global PENNSIEVE_URL
    error = []
    # check that the files and folders specified in the dataset are valid
    dataset_name = soda_json_structure["bf-dataset-selected"]["dataset-name"]
    dataset_id = get_dataset_id(ps, dataset_name)
    r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(ps))
    r.raise_for_status()
    root_folder = r.json()["children"]

    if "dataset-structure" in soda_json_structure.keys():
        dataset_structure = soda_json_structure["dataset-structure"]
        if "folders" in dataset_structure:
            for folder_key, folder in dataset_structure["folders"].items():
                folder_type = folder["type"]
                relative_path = folder_key
                if folder_type == "bf":
                    collection_id = folder["path"]
                    collection_actions = folder["action"]
                    if "moved" in collection_actions:
                        try:
                            r = requests.get(f"{PENNSIEVE_URL}/packages/{collection_id}/view", headers=create_request_headers(ps))
                            r.raise_for_status()
                        except Exception as e:
                            error.append(f"{relative_path} id: {collection_id}")
                        continue
                    if next((item for item in root_folder if item["content"]["id"] == collection_id), None) is None:
                        error.append(f"{relative_path} id: {collection_id}")
                    else:
                        # recursively check all files + subfolders of collection_id
                        error = check_folder_validity(collection_id, folder, relative_path, error)

    if "metadata-files" in soda_json_structure.keys():
        # check that the metadata files specified in the dataset are valid
        for file_key, file in soda_json_structure["metadata-files"].items():
            if file["type"] == "bf":
                file_id = file["path"]
                if next((item for item in root_folder if item["content"]["id"] == file_id), None) is None:
                    error.append(f"{file_key} id: {file_id}")

    # if there are items in the error list, check if they have been "moved"

    if len(error) > 0:
        error_message = [
            "Error: The following Pennsieve files/folders are invalid. Specify them again or remove them."
        ]
        error = error_message + error

    return error


def check_server_access_to_files(file_list):
    # Return two lists, one that the server can open, and one that it can not.
    # This is to avoid the server trying to open files that it does not have access to.cf
    accessible_files = []
    inaccessible_files = []
    for file in file_list:
        if os.path.isfile(file) or os.path.isdir(file):
            accessible_files.append(file)
        else:
            inaccessible_files.append(file)

    return {"accessible_files": accessible_files, "inaccessible_files": inaccessible_files}


def clean_json_structure(soda_json_structure):
    global namespace_logger
    # Delete any files on Pennsieve that have been marked as deleted
    def recursive_file_delete(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if item in ["manifest.xlsx", "manifest.csv"]:
                    continue
                if "deleted" in folder["files"][item]["action"]:
                    # remove the file from the soda json structure
                    del folder["files"][item]

        for item in list(folder["folders"]):
            recursive_file_delete(folder["folders"][item])
        return


    # Rename any files that exist on Pennsieve
    def recursive_file_rename(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if (
                    "renamed" in folder["files"][item]["action"]
                    and folder["files"][item]["type"] == "bf"
                ):
                    continue

        for item in list(folder["folders"]):
            recursive_file_rename(folder["folders"][item])

        return

    def recursive_folder_delete(folder):
        """
        Delete any stray folders that exist on Pennsieve
        Only top level files are deleted since the api deletes any
        files and folders that exist inside.
        """

        for folder_item in list(folder["folders"]):
            if folder["folders"][folder_item]["type"] == "bf":
                if "deleted" in folder["folders"][folder_item]["action"]:
                    namespace_logger.info(f"Deleting folder {folder_item}")
                    del folder["folders"][folder_item]
                else:
                    recursive_folder_delete(folder["folders"][folder_item])
            else:
                recursive_folder_delete(folder["folders"][folder_item])
        return

    main_keys = soda_json_structure.keys()
    dataset_structure = soda_json_structure["dataset-structure"]

    if ("dataset-structure" not in main_keys and "metadata-files" not in main_keys):
        abort(400, "Error: Your dataset is empty. Please add valid files and non-empty folders to your dataset.")

    if "generate-dataset" in main_keys:
        # Check that local files/folders exist
        try:
            if error := check_local_dataset_files_validity(soda_json_structure):
                abort(400, error)
            # check that dataset is not empty after removing all the empty files and folders
            if not soda_json_structure["dataset-structure"]["folders"] and "metadata-files" not in soda_json_structure:
                abort(400, "Error: Your dataset is empty. Please add valid files and non-empty folders to your dataset.")
        except Exception as e:
            raise e

    if "starting-point" in main_keys:
        if soda_json_structure["starting-point"]["type"] == "bf" or soda_json_structure["starting-point"]["type"] == "local":
            # remove deleted files and folders from the json
            recursive_file_delete(dataset_structure)
            recursive_folder_delete(dataset_structure)
            soda_json_structure["dataset-structure"] = dataset_structure

    # here will be clean up the soda json object before creating the manifest file cards
    return {"soda_json_structure": soda_json_structure}


def main_curate_function(soda_json_structure):
    global namespace_logger

    namespace_logger.info("Starting main_curate_function")
    namespace_logger.info(f"main_curate_function metadata generate-options={soda_json_structure['generate-dataset']}")

    global main_curate_status
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global main_generated_dataset_size
    global start_generate
    global generate_start_time
    global main_generate_destination
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global uploaded_folder_counter

    global myds
    global generated_dataset_id

    start_generate = 0
    start = timer()
    generate_start_time = time.time()

    # variables for tracking the progress of the curate process on the frontend 
    main_curate_status = ""
    main_curate_progress_message = "Starting..."
    main_total_generate_dataset_size = 0
    main_generated_dataset_size = 0
    main_curation_uploaded_files = 0
    uploaded_folder_counter = 0
    generated_dataset_id = None

    main_curate_status = "Curating"
    main_curate_progress_message = "Starting dataset curation"
    main_generate_destination = ""
    main_initial_bfdataset_size = 0

    myds = ""
    main_keys = soda_json_structure.keys()
    error = ""

    # 1] Check for potential errors
    namespace_logger.info("main_curate_function step 1")

    # 1.1. If the dataset is being generated locally then check that the local destination is valid
    if "generate-dataset" in main_keys and soda_json_structure["generate-dataset"]["destination"] == "local":
        main_curate_progress_message = "Checking that the local destination selected for generating your dataset is valid"
        generate_dataset = soda_json_structure["generate-dataset"]
        local_dataset_path = generate_dataset["path"]
        if not isdir(local_dataset_path):
            error_message = (
                "Error: The Path "
                + local_dataset_path
                + " is not found. Please select a valid destination folder for the new dataset"
            )
            main_curate_status = "Done"
            error = error_message
            abort(400, error)

    namespace_logger.info("main_curate_function step 1.2")

    # 1.2. If generating dataset to Pennsieve or any other Pennsieve actions are requested check that the destination is valid
    if "bf-account-selected" in soda_json_structure:
        # check that the Pennsieve account is valid
        try:
            main_curate_progress_message = (
                "Checking that the selected Pennsieve account is valid"
            )
            accountname = soda_json_structure["bf-account-selected"]["account-name"]
            ps = connect_pennsieve_client(accountname)
        except Exception as e:
            main_curate_status = "Done"
            abort(400, "Please select a valid Pennsieve account.")

    # if uploading on an existing bf dataset
    if "bf-dataset-selected" in soda_json_structure:
        # check that the Pennsieve dataset is valid
        try:
            main_curate_progress_message = (
                "Checking that the selected Pennsieve dataset is valid"
            )
            bfdataset = soda_json_structure["bf-dataset-selected"]["dataset-name"]
            token = get_access_token()
            selected_dataset_id = get_dataset_id(token, bfdataset)

        except Exception as e:
            main_curate_status = "Done"
            abort(400, "Error: Please select a valid Pennsieve dataset")

        # check that the user has permissions for uploading and modifying the dataset
        main_curate_progress_message = "Checking that you have required permissions for modifying the selected dataset"
        role = pennsieve_get_current_user_permissions(selected_dataset_id, token)["role"]
        if role not in ["owner", "manager", "editor"]:
            main_curate_status = "Done"
            abort(403, "Error: You don't have permissions for uploading to this Pennsieve dataset")

    namespace_logger.info("main_curate_function step 1.3")

    # 1.3. Check that specified dataset files and folders are valid (existing path) if generate dataset is requested
    # Note: Empty folders and 0 kb files will be removed without warning (a warning will be provided on the front end before starting the curate process)
    if "generate-dataset" in main_keys:

        # Check at least one file or folder are added to the dataset
        try:
            main_curate_progress_message = "Checking that the dataset is not empty"
            if (
                "dataset-structure" not in soda_json_structure
                and "metadata-files" not in soda_json_structure
            ):
                main_curate_status = "Done" 
                abort(400, "Error: Your dataset is empty. Please add valid files and non-empty folders to your dataset.")
        except Exception as e:
            main_curate_status = "Done"
            raise e

        namespace_logger.info("main_curate_function step 1.3.1")

        # Check that local files/folders exist
        try:
            if error := check_local_dataset_files_validity(soda_json_structure):
                main_curate_status = "Done"
                abort(400, error)

            # check that dataset is not empty after removing all the empty files and folders
            if not soda_json_structure["dataset-structure"]["folders"] and "metadata-files" not in soda_json_structure:
                main_curate_status = "Done"
                abort(400, "Error: Your dataset is empty. Please add valid files and non-empty folders to your dataset.")
        except Exception as e:
            main_curate_status = "Done"
            raise e

        namespace_logger.info("main_curate_function step 1.3.2")
        # Check that bf files/folders exist (Only used for when generating from an existing Pennsieve dataset)
        generate_option = soda_json_structure["generate-dataset"]["generate-option"]
        if generate_option == "existing-bf":
            try:
                main_curate_progress_message = (
                    "Checking that the Pennsieve files and folders are valid"
                )
                if soda_json_structure["generate-dataset"]["destination"] == "bf":
                    if error := ps_check_dataset_files_validity(soda_json_structure, ps):
                        namespace_logger.info("Failed to validate dataset files")
                        namespace_logger.info(error)
                        main_curate_status = "Done"
                        abort(400, error)
            except Exception as e:
                main_curate_status = "Done"
                raise e


    namespace_logger.info("main_curate_function step 3")

    # 2] Generate
    if "generate-dataset" in main_keys:
        main_curate_progress_message = "Generating dataset"
        try:
            # Generate dataset locally
            if soda_json_structure["generate-dataset"]["destination"] == "local":
                main_generate_destination = soda_json_structure["generate-dataset"][
                    "destination"
                ]
                _, main_total_generate_dataset_size = generate_dataset_locally(
                    soda_json_structure
                )

            # Generate dataset to Pennsieve
            if soda_json_structure["generate-dataset"]["destination"] == "bf":
                main_generate_destination = soda_json_structure["generate-dataset"][
                    "destination"
                ]
                if generate_option == "existing-bf":
                    # make an api request to pennsieve to get the dataset details
                    namespace_logger.info(f"this is the selected_dataset_id for existing-bf {selected_dataset_id}")
                    r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}", headers=create_request_headers(get_access_token()))
                    r.raise_for_status()
                    myds = r.json()
                    ps_update_existing_dataset(soda_json_structure, myds, ps)

                elif generate_option == "new":
                    # if dataset name is in the generate-dataset section, we are generating a new dataset
                    if "dataset-name" in soda_json_structure["generate-dataset"]:
                        dataset_name = soda_json_structure["generate-dataset"][
                            "dataset-name"
                        ]
                        ds = ps_create_new_dataset(dataset_name, ps)
                        selected_dataset_id = ds["content"]["id"]


                    # whether we are generating a new dataset or merging, we want the dataset information for later steps
                    r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}", headers=create_request_headers(get_access_token()))
                    r.raise_for_status()
                    myds = r.json()

                    ps_upload_to_dataset(soda_json_structure, ps, myds)
        except Exception as e:
            main_curate_status = "Done"
            raise e

    namespace_logger.info("main_curate_function finished")

    main_curate_status = "Done"
    main_curate_progress_message = "Success: COMPLETED!"
    end = timer()
    namespace_logger.info(f"Time for main_curate_function function: {timedelta(seconds=end - start)}")
    return {
        "main_curate_progress_message": main_curate_progress_message,
        "main_total_generate_dataset_size": main_total_generate_dataset_size,
        "main_curation_uploaded_files": main_curation_uploaded_files,
    }


def main_curate_function_progress():
    """
    Function frequently called by front end to help keep track of the dataset generation progress
    """

    global main_curate_status  # empty if curate on going, "Done" when main curate function stopped (error or completed)
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global main_generated_dataset_size
    global start_generate
    global generate_start_time
    global main_generate_destination
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global total_bytes_uploaded # current number of bytes uploaded to Pennsieve in the upload session
    global myds 


    elapsed_time = time.time() - generate_start_time
    elapsed_time_formatted = time_format(elapsed_time)

    return {
        "main_curate_status": main_curate_status,
        "start_generate": start_generate,
        "main_curate_progress_message": main_curate_progress_message,
        "main_total_generate_dataset_size": main_total_generate_dataset_size,
        "main_generated_dataset_size": total_bytes_uploaded["value"],
        "elapsed_time_formatted": elapsed_time_formatted,
        "total_files_uploaded": main_curation_uploaded_files,
        "generated_dataset_id": myds["content"]["id"] if myds != "" else None, # when a new dataset gets generated log its id to our analytics
    }


def preview_dataset(soda_json_structure):
    """
    Associated with 'Preview' button in the SODA interface
    Creates a folder for preview and adds mock files based on the files specified in the UI by the user (same name as origin but 0 kb in size)
    Opens the dialog box to showcase the files / folders added

    Args:
        soda_json_structure: soda dict with information about all specified files and folders
    Action:
        Opens the dialog box at preview_path
    Returns:
        preview_path: path of the folder where the preview files are located
    """

    preview_path = join(userpath, "SODA", "Preview_dataset")

    # remove empty files and folders from dataset
    try:
        check_empty_files_folders(soda_json_structure)
    except Exception as e:
        raise e

    # create Preview_dataset folder
    try:
        if isdir(preview_path):
            shutil.rmtree(preview_path, ignore_errors=True)
        makedirs(preview_path)
    except Exception as e:
        raise e

    try:

        if "dataset-structure" in soda_json_structure.keys():
            # create folder structure
            def recursive_create_mock_folder_structure(my_folder, my_folderpath):
                if "folders" in my_folder.keys():
                    for folder_key, folder in my_folder["folders"].items():
                        folderpath = join(my_folderpath, folder_key)
                        if not isdir(folderpath):
                            mkdir(folderpath)
                        recursive_create_mock_folder_structure(folder, folderpath)

                if "files" in my_folder.keys():
                    for file_key, file in my_folder["files"].items():
                        if not "deleted" in file["action"]:
                            open(join(my_folderpath, file_key), "a").close()

            dataset_structure = soda_json_structure["dataset-structure"]
            folderpath = preview_path
            recursive_create_mock_folder_structure(dataset_structure, folderpath)

            if "manifest-files" in soda_json_structure.keys():
                if "folders" in dataset_structure.keys():
                    for folder_key, folder in dataset_structure["folders"].items():
                        manifest_path = join(preview_path, folder_key, "manifest.xlsx")
                        if not isfile(manifest_path):
                            open(manifest_path, "a").close()

        if "metadata-files" in soda_json_structure.keys():
            for metadata_key in soda_json_structure["metadata-files"].keys():
                open(join(preview_path, metadata_key), "a").close()

        if len(listdir(preview_path)) > 0:
            folder_in_preview = listdir(preview_path)[0]
            open_file(join(preview_path, folder_in_preview))
        else:
            open_file(preview_path)

        return preview_path

    except Exception as e:
        raise e


def generate_manifest_file_locally(generate_purpose, soda_json_structure):
    """
    Function to generate manifest files locally
    """

    def recursive_item_path_create(folder, path):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if "folderpath" not in folder["files"][item]:
                    folder["files"][item]["folderpath"] = path[:]

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                if "folderpath" not in folder["folders"][item]:
                    folder["folders"][item]["folderpath"] = path[:]
                    folder["folders"][item]["folderpath"].append(item)
                recursive_item_path_create(
                    folder["folders"][item], folder["folders"][item]["folderpath"][:]
                )

        return

    def copytree(src, dst, symlinks=False, ignore=None):
        for item in os.listdir(src):
            s = os.path.join(src, item)
            d = os.path.join(dst, item)
            if os.path.isdir(s):
                if os.path.exists(d):
                    shutil.rmtree(d)
                shutil.copytree(s, d, symlinks, ignore)
            else:
                shutil.copy2(s, d)

    dataset_structure = soda_json_structure["dataset-structure"]
    starting_point = soda_json_structure["starting-point"]["type"]
    manifest_destination = soda_json_structure["manifest-files"]["local-destination"]

    recursive_item_path_create(dataset_structure, [])
    create_high_lvl_manifest_files_existing_ps_starting_point(soda_json_structure, manifest_folder_path)

    if generate_purpose == "edit-manifest":
        manifest_destination = os.path.join(manifest_destination, "SODA Manifest Files")

    else:
        manifest_destination = return_new_path(
            os.path.join(manifest_destination, "SODA Manifest Files")
        )

    copytree(manifest_folder_path, manifest_destination)

    if generate_purpose == "edit-manifest":
        return {"success_message_or_manifest_destination": manifest_destination}

    open_file(manifest_destination)
    return {"success_message_or_manifest_destination": "success"}


def generate_manifest_file_data(dataset_structure_obj):
    # modify this function here to handle paths from pennsieve
    # create path using bfpath key from json object

    local_timezone = TZLOCAL()

    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]

    def get_name_extension(file_name):
        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break

        ext = ""

        if double_ext == False:
            ext = os.path.splitext(file_name)[1]
        else:
            ext = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )
        return ext

    def guided_recursive_folder_traversal(folder, hlf_data_array, ds_struct_path):
        if "files" in folder.keys():
            standard_manifest_columns = ["filename", "timestamp", "description", "file type", "Additional Metadata"]
            if(len(hlf_data_array) < 1):
                hlf_data_array.append(standard_manifest_columns)
            for item in list(folder["files"]):
                # do not generate a manifest file entry for the manifest file itself
                if item in ["manifest.xlsx", "manifest.csv"]:
                    continue
                file_manifest_template_data = []
                local_path_to_file = folder["files"][item]["path"].replace("\\", "/")
                item_description = folder["files"][item]["description"]
                item_additional_info = folder["files"][item]["additional-metadata"]

                # The name of the file eg "file.txt"
                file_name = os.path.basename(local_path_to_file)
                namespace_logger.info(f"file_name based on file path: {file_name}")
                namespace_logger.info(f"file name based on key: {item}")
                if file_name != item:
                    file_name = item
                if len(ds_struct_path) > 0:
                    filename_entry = "/".join(ds_struct_path) + "/" + file_name
                else:
                    filename_entry = file_name

                # The extension of the file eg ".txt"
                file_type_entry = get_name_extension(file_name)

                # The timestamp of the file on the user's local machine
                file_path = pathlib.Path(local_path_to_file)
                mtime = file_path.stat().st_mtime
                last_mod_time = datetime.fromtimestamp(mtime, tz=local_timezone).fromtimestamp(mtime).astimezone(
                    local_timezone
                )
                timestamp_entry = last_mod_time.isoformat().replace(".", ",").replace("+00:00", "Z")

                if(filename_entry[0:1] == "/"):
                    file_manifest_template_data.append(filename_entry[:1])
                else:
                    file_manifest_template_data.append(filename_entry)
                
                file_manifest_template_data.append(timestamp_entry)
                file_manifest_template_data.append(item_description)
                file_manifest_template_data.append(file_type_entry)
                file_manifest_template_data.append(item_additional_info)

                # extra column key is an object of all extra columns of a manifest
                # key will be the column header and value will be the value of the column+row 
                # (from the excel) (now in the form of a dict)
                if "extra_columns" in folder["files"][item]:
                    for key in folder["files"][item]["extra_columns"]:
                        file_manifest_template_data.append(folder["files"][item]["extra_columns"][key])
                        if key not in hlf_data_array[0]:
                            # add column name to manifest column names array
                            hlf_data_array[0].append(key)

                hlf_data_array.append(file_manifest_template_data)

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                relative_structure_path.append(item)
                guided_recursive_folder_traversal(
                    folder["folders"][item], hlf_data_array, relative_structure_path
                )
                relative_structure_path.pop()
        return

    def pennsieve_recursive_folder_traversal(folder, hlf_data_array, ds_struct_path):
        if "files" in folder.keys():
            standard_manifest_columns = ["filename", "timestamp", "description", "file type", "Additional Metadata"]
            if(len(hlf_data_array) < 1):
                hlf_data_array.append(standard_manifest_columns)
            for item in list(folder["files"]):
                file_manifest_template_data = []
                if item in ["manifest.xlsx", "manifest.csv"]:
                    continue
                item_description = folder["files"][item]["description"]
                item_additional_info = folder["files"][item]["additional-metadata"]
                file_name = ""
                if folder["files"][item]["type"] == "bf": 
                    file_name = os.path.basename(item)
                    timestamp_entry = folder["files"][item]["timestamp"]
                else:
                    local_path_to_file = folder["files"][item]["path"].replace("\\", "/")
                    file_name = os.path.basename(local_path_to_file)
                    file_path = pathlib.Path(local_path_to_file)
                    mtime = file_path.stat().st_mtime
                    last_mod_time = datetime.fromtimestamp(mtime, tz=local_timezone).fromtimestamp(mtime).astimezone(local_timezone)
                    timestamp_entry = last_mod_time.isoformat().replace(".", ",").replace("+00:00", "Z")

                
                filename_entry = "/".join(ds_struct_path) + "/" + file_name
                file_type_entry = get_name_extension(file_name)

                if(filename_entry[0:1] == "/"):
                    file_manifest_template_data.append(filename_entry[1:])
                else:
                    file_manifest_template_data.append(filename_entry)
                file_manifest_template_data.append(timestamp_entry)
                file_manifest_template_data.append(item_description)
                file_manifest_template_data.append(file_type_entry)
                file_manifest_template_data.append(item_additional_info)
                # extra column key is an object of all extra columns of a manifest
                # key will be the column header and value will be the value of the column+row 
                # (from the excel) (now in the form of a dict)
                if "extra_columns" in folder["files"][item]:
                    for key in folder["files"][item]["extra_columns"]:
                        file_manifest_template_data.append(folder["files"][item]["extra_columns"][key])
                        if key not in hlf_data_array[0]:
                            # add column name to manifest column names array
                            hlf_data_array[0].append(key)

                hlf_data_array.append(file_manifest_template_data)

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                relative_structure_path.append(item)
                pennsieve_recursive_folder_traversal(
                    folder["folders"][item], hlf_data_array, relative_structure_path
                )
                relative_structure_path.pop()
        return

    # Initialize the array that the manifest data will be added to.
    hlf_manifest_data = {}
    # any additional columns created by the user will be appended with the high level folder when found


    # Loop through each high level folder and create a manifest data array for each.
    for high_level_folder in list(dataset_structure_obj["folders"]):
        hlf_data_array = []

        # create an array to keep track of the path to the obj being recursed over
        relative_structure_path = []
        # hlf_data_array.append(standard_manifest_columns)

        if "bfpath" in dataset_structure_obj["folders"][high_level_folder]:
            # means the json is from a pennsieve dataset
            pennsieve_recursive_folder_traversal(dataset_structure_obj["folders"][high_level_folder], hlf_data_array, relative_structure_path)
        else:
            guided_recursive_folder_traversal(dataset_structure_obj["folders"][high_level_folder], hlf_data_array, relative_structure_path)
        hlf_manifest_data[high_level_folder] = hlf_data_array

    return hlf_manifest_data


def handle_duplicate_package_name_error(e, soda_json_structure):
    if (
        e.response.text
        == '{"type":"BadRequest","message":"package name must be unique","code":400}'
    ):
        if "if-existing-files" in soda_json_structure["generate-dataset"]:
            if (
                soda_json_structure["generate-dataset"]["if-existing-files"]
                == "create-duplicate"
            ):
                return

    raise e
