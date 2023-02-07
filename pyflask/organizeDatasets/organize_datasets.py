### Import required python modules
from turtle import pensize
from flask import abort
import platform
import os
import itertools
from os import makedirs, mkdir, walk
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from os.path import (
    isdir,
    join,
    splitext,
    basename,
    exists,
    expanduser,
    getsize,
)
import pandas as pd
import time
import shutil
import subprocess
import re
import gevent
import pathlib
from datetime import datetime, timezone
import requests 
from permissions import bf_get_current_user_permission_agent_two, has_edit_permissions
from utils import connect_pennsieve_client, get_dataset_id, create_request_headers, authenticate_user_with_client
from namespaces import NamespaceEnum, get_namespace_logger
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from utils import load_manifest_to_dataframe

import json
namespace_logger = get_namespace_logger(NamespaceEnum.ORGANIZE_DATASETS)
from authentication import get_access_token




### Global variables
create_soda_json_progress = 0
create_soda_json_total_items = 0
create_soda_json_completed = 0
curateprogress = " "
curatestatus = " "
curateprintstatus = " "
total_dataset_size = 1
curated_dataset_size = 0
start_time = 0

userpath = expanduser("~")
configpath = join(userpath, ".pennsieve", "config.ini")
submitdataprogress = " "
submitdatastatus = " "
submitprintstatus = " "
total_file_size = 1
uploaded_file_size = 0
start_time_bf_upload = 0
start_submit = 0
metadatapath = join(userpath, "SODA", "SODA_metadata")

bf = ""
myds = ""
initial_bfdataset_size = 0
upload_directly_to_bf = 0
initial_bfdataset_size_submit = 0

forbidden_characters = '<>:"/\|?*'
forbidden_characters_bf = '\/:*?"<>'

PENNSIEVE_URL = "https://api.pennsieve.io"

from namespaces import NamespaceEnum, get_namespace_logger
namespace_logger = get_namespace_logger(NamespaceEnum.MANAGE_DATASETS)

### Internal functions
def TZLOCAL():
    return datetime.now(timezone.utc).astimezone().tzinfo

## these subsequent CheckLeafValue and traverseForLeafNodes functions check for the validity of file paths,
## and folder and file size
def checkLeafValue(leafName, leafNodeValue):
    error, c = "", 0
    total_dataset_size = 1
    curatestatus = ""

    if isinstance(leafNodeValue, list):
        filePath = leafNodeValue[0]

        if exists(filePath):
            filePathSize = getsize(filePath)
            if filePathSize == 0:
                c += 1
                error = error + filePath + " is 0 KB <br>"
            else:
                total_dataset_size += filePathSize
        else:
            c += 1
            error = error + filePath + " doesn not exist! <br>"

    elif isinstance(leafNodeValue, dict):
        c += 1
        error = error + leafName + " is empty <br>"

    if c > 0:
        error += "<br>Please remove invalid files/folders from your dataset and try again"
        curatestatus = "Done"
        abort(400, error)


    return [True, total_dataset_size - 1]


def traverseForLeafNodes(jsonStructure):
    total_dataset_size = 1

    for key in jsonStructure:
        if isinstance(jsonStructure[key], list):

            returnedOutput = checkLeafValue(key, jsonStructure[key])

            # returnedOutput = [True, total_dataset_size-1]
            if returnedOutput[0]:
                total_dataset_size += returnedOutput[1]

        elif len(jsonStructure[key]) == 0:
            returnedOutput = checkLeafValue(key, jsonStructure[key])

        else:
            # going one step down in the object tree
            traverseForLeafNodes(jsonStructure[key])

    return total_dataset_size


######## CREATE FILES FOR CURATE_DATASET FUNCTION
def createFiles(jsonpath, fileKey, distdir, listallfiles):
    # fileKey is the key in json structure that is a file (meaning that its value is an array)
    # note: this fileKey can be the new name of the file (if renamed by users in SODA)
    # (or can be the original basename)

    srcfile = jsonpath[fileKey][0]
    distfile = distdir
    listallfiles.append([srcfile, distfile])


def ignore_empty_high_level_folders(jsonObject):
    items_to_delete = [
        folder for folder in jsonObject.keys() if len(jsonObject[folder].keys()) == 0
    ]

    for item in items_to_delete:
        del jsonObject[item]

    return jsonObject


def generate_dataset_locally(destinationdataset, pathdataset, newdatasetname, jsonpath):
    """
    Associated with 'Generate' button in the 'Generate dataset' section of SODA interface
    Checks validity of files / paths / folders and then generates the files and folders
    as requested along with progress status

    Args:
        destinationdataset: type of destination dataset ('modify existing', 'create new', or 'upload to Pennsieve')
        pathdataset: destination path of new dataset if created locally or name of Pennsieve account (string)
        newdatasetname: name of the local dataset or name of the dataset on Pennsieve (string)
        manifeststatus: boolean to check if user request manifest files
        jsonpath: path of the files to be included in the dataset (dictionary)
    """
    global curatestatus  # set to 'Done' when completed or error to stop progress tracking from front-end
    global curateprogress  # GUI messages shown to user to provide update on progress
    global curateprintstatus  # If = "Curating" Progress messages are shown to user
    global total_dataset_size  # total size of the dataset to be generated
    global curated_dataset_size  # total size of the dataset generated (locally or on Pennsieve) at a given time
    global start_time
    global bf
    global myds
    global upload_directly_to_bf
    global start_submit
    global initial_bfdataset_size

    curateprogress = " "
    curatestatus = ""
    curateprintstatus = " "
    error, c = "", 0
    curated_dataset_size = 0
    start_time = 0
    upload_directly_to_bf = 0
    start_submit = 0
    initial_bfdataset_size = 0

    jsonstructure_non_empty = ignore_empty_high_level_folders(jsonpath)

    if destinationdataset == "create new":
        if not isdir(pathdataset):
            curatestatus = "Done"
            abort(400, "Please select a valid folder for new dataset")
        if not newdatasetname:
            curatestatus = "Done"
            abort(400, "Please enter a valid name for new dataset folder")
        if check_forbidden_characters(newdatasetname):
            curatestatus = "Done"
            abort(400,  "A folder name cannot contain any of the following characters " + forbidden_characters)

    total_dataset_size = 1

    # check if path in jsonpath are valid and calculate total dataset size
    total_dataset_size = traverseForLeafNodes(jsonstructure_non_empty)
    total_dataset_size = total_dataset_size - 1

    # CREATE NEW
    if destinationdataset == "create new":
        try:
            listallfiles = []
            pathnewdatasetfolder = join(pathdataset, newdatasetname)
            pathnewdatasetfolder = return_new_path(pathnewdatasetfolder)
            open_file(pathnewdatasetfolder)

            curateprogress = "Started"
            curateprintstatus = "Curating"
            start_time = time.time()
            start_submit = 1

            pathdataset = pathnewdatasetfolder
            mkdir(pathdataset)
            create_dataset(pathdataset, jsonstructure_non_empty, listallfiles)

            curateprogress = "New dataset created"
            curateprogress = "Success: COMPLETED!"
            curatestatus = "Done"

        except Exception as e:
            curatestatus = "Done"
            raise e


def create_folder_level_manifest(jsonpath, jsondescription):
    """
    Function to create manifest files for each SPARC folder.
    Files are created in a temporary folder

    Args:
        datasetpath: path of the dataset (string)
        jsonpath: all paths in json format with key being SPARC folder names (dictionary)
        jsondescription: description associated with each path (dictionary)
    Action:
        Creates manifest files in xslx format for each SPARC folder
    """
    global total_dataset_size
    local_timezone = TZLOCAL()

    try:
        datasetpath = metadatapath
        shutil.rmtree(datasetpath) if isdir(datasetpath) else 0
        makedirs(datasetpath)
        folders = list(jsonpath.keys())
        if "main" in folders:
            folders.remove("main")
        # In each SPARC folder, generate a manifest file
        for folder in folders:
            if jsonpath[folder] != []:
                # Initialize dataframe where manifest info will be stored
                df = pd.DataFrame(
                    columns=[
                        "filename",
                        "timestamp",
                        "description",
                        "file type",
                        "Additional Metadata",
                    ]
                )
                # Get list of files/folders in the the folder
                # Remove manifest file from the list if already exists
                folderpath = join(datasetpath, folder)
                allfiles = jsonpath[folder]
                alldescription = jsondescription[folder + "_description"]
                manifestexists = join(folderpath, "manifest.xlsx")

                countpath = -1
                for pathname in allfiles:
                    countpath += 1
                    if basename(pathname) in ["manifest.csv", "manifest.xlsx"]:
                        allfiles.pop(countpath)
                        alldescription.pop(countpath)

                # Populate manifest dataframe
                filename, timestamp, filetype, filedescription = [], [], [], []
                countpath = -1
                for paths in allfiles:
                    if isdir(paths):
                        key = basename(paths)
                        alldescription.pop(0)
                        for subdir, dirs, files in os.walk(paths):
                            for file in files:
                                gevent.sleep(0)
                                filepath = pathlib.Path(paths) / subdir / file
                                mtime = filepath.stat().st_mtime
                                lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                                    local_timezone
                                )
                                timestamp.append(
                                    lastmodtime.isoformat()
                                    .replace(".", ",")
                                    .replace("+00:00", "Z")
                                )
                                full_filename = filepath.name

                                if folder == "main":  # if file in main folder
                                    filename.append(
                                        full_filename
                                    ) if folder == "" else filename.append(
                                        join(folder, full_filename)
                                    )
                                else:
                                    subdirname = os.path.relpath(
                                        subdir, paths
                                    )  # gives relative path of the directory of the file w.r.t paths
                                    if subdirname == ".":
                                        filename.append(join(key, full_filename))
                                    else:
                                        filename.append(
                                            join(key, subdirname, full_filename)
                                        )

                                fileextension = splitext(full_filename)[1]
                                if (
                                    not fileextension
                                ):  # if empty (happens e.g. with Readme files)
                                    fileextension = "None"
                                filetype.append(fileextension)
                                filedescription.append("")
                    else:
                        gevent.sleep(0)
                        countpath += 1
                        filepath = pathlib.Path(paths)
                        file = filepath.name
                        filename.append(file)
                        mtime = filepath.stat().st_mtime
                        lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                            local_timezone
                        )
                        timestamp.append(
                            lastmodtime.isoformat()
                            .replace(".", ",")
                            .replace("+00:00", "Z")
                        )
                        filedescription.append(alldescription[countpath])
                        if isdir(paths):
                            filetype.append("folder")
                        else:
                            fileextension = splitext(file)[1] or "None"
                            filetype.append(fileextension)

                df["filename"] = filename
                df["timestamp"] = timestamp
                df["file type"] = filetype
                df["description"] = filedescription

                makedirs(folderpath)
                # Save manifest as Excel sheet
                manifestfile = join(folderpath, "manifest.xlsx")
                df.to_excel(manifestfile, index=None, header=True)
                wb = load_workbook(manifestfile)
                ws = wb.active
                blueFill = PatternFill(
                    start_color="9DC3E6", fill_type="solid"
                )
                greenFill = PatternFill(
                    start_color="A8D08D", fill_type="solid"
                )
                yellowFill = PatternFill(
                    start_color="FFD965", fill_type="solid"
                )
                ws['A1'].fill = blueFill
                ws['B1'].fill = greenFill
                ws['C1'].fill = greenFill
                ws['D1'].fill = greenFill
                ws['E1'].fill = yellowFill
                wb.save(manifestfile)
                total_dataset_size += path_size(manifestfile)
                jsonpath[folder].append(manifestfile)

        return jsonpath

    except Exception as e:
        raise e


def check_forbidden_characters(my_string):
    """
    Check for forbidden characters in file/folder name

    Args:
        my_string: string with characters (string)
    Returns:
        False: no forbidden character
        True: presence of forbidden character(s)
    """
    regex = re.compile("[" + forbidden_characters + "]")
    return regex.search(my_string) is not None or "\\" in r"%r" % my_string


def folder_size(path):
    """
    Provides the size of the folder indicated by path

    Args:
        path: path of the folder (string)
    Returns:
        total_size: total size of the folder in bytes (integer)
    """
    total_size = 0
    start_path = "."  # To get size of current directory
    for path, dirs, files in walk(path):
        for f in files:
            fp = join(path, f)
            total_size += getsize(fp)
    return total_size


def open_file(file_path):
    """
    Opening folder on all platforms
    https://stackoverflow.com/questions/6631299/python-opening-a-folder-in-explorer-nautilus-mac-thingie

    Args:
        file_path: path of the folder (string)
    Action:
        Opens file explorer window to the given path
    """
    try:
        if platform.system() == "Windows":
            subprocess.Popen(f"explorer /select,{str(file_path)}")
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])
    except Exception as e:
        raise e



def path_size(path):
    """
    Returns size of the path, after checking if it's a folder or a file
    Args:
        path: path of the file/folder (string)
    Returns:
        total_size: total size of the file/folder in bytes (integer)
    """
    return folder_size(path) if isdir(path) else getsize(path)


def mycopyfile_with_metadata(src, dst, *, follow_symlinks=True):
    """
    Copy file src to dst with metadata (timestamp, permission, etc.) conserved

    Args:
        src: source file (string)
        dst: destination file (string)
    Returns:
        dst
    """
    if not follow_symlinks and os.path.islink(src):
        os.symlink(os.readlink(src), dst)
    else:
        with open(src, "rb") as fsrc:
            with open(dst, "wb") as fdst:
                mycopyfileobj(fsrc, fdst)
    shutil.copystat(src, dst)
    return dst


def mycopyfileobj(fsrc, fdst, length=16 * 1024 * 16):
    """
    Helper function to copy file

    Args:
        fsrc: source file opened in python (file-like object)
        fdst: destination file accessed in python (file-like object)
        length: copied buffer size in bytes (integer)
    """
    global curateprogress
    global total_dataset_size
    global curated_dataset_size
    while True:
        buf = fsrc.read(length)
        if not buf:
            break
        gevent.sleep(0)
        fdst.write(buf)
        curated_dataset_size += len(buf)


def return_new_path(topath):
    """
    This function checks if a folder already exists and in such cases,
    appends (2) or (3) etc. to the folder name

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """
    if not exists(topath):
        return topath
    i = 2
    while True:
        if not exists(topath + " (" + str(i) + ")"):
            return topath + " (" + str(i) + ")"
        i += 1


def create_dataset(recursivePath, jsonStructure, listallfiles):
    """
    Associated with 'Create new dataset locally' option of SODA interface
    for creating requested folders and files to the destination path specified

    Args:
        recursivePath: destination path for creating new folders and files (recursively) (string)
        jsonStructure: all paths (dictionary)
    Action:
        Creates the folders and files specified
    """

    global curateprogress

    for key in jsonStructure:

        if isinstance(jsonStructure[key], dict):

            if not exists(join(recursivePath, key)):
                mkdir(join(recursivePath, key))

            outputpath = join(recursivePath, key)

            create_dataset(outputpath, jsonStructure[key], listallfiles)

        elif isinstance(jsonStructure[key], list):

            outputpath = join(recursivePath, key)

            createFiles(jsonStructure, key, outputpath, listallfiles)

            for fileinfo in listallfiles:
                srcfile = fileinfo[0]
                distfile = fileinfo[1]
                curateprogress = f"Copying {str(srcfile)}"

                mycopyfile_with_metadata(srcfile, distfile)


def create_soda_json_object_backend(
    soda_json_structure, root_folder_path, irregularFolders, replaced
):
    """
    Function for importing files from local machine into json structure
    """
    global create_soda_json_progress  # amount of items counted during recursion
    global create_soda_json_total_items  # counts the total items in folder
    global create_soda_json_completed  # completed progress is either 0 or 1

    high_level_sparc_folders = [
        "code",
        "derivative",
        "docs",
        "primary",
        "protocol",
        "source",
    ]

    dataset_folder = soda_json_structure["dataset-structure"] = {"folders": {}}

    def recursive_structure_create(dataset_structure, folder_path):
        global create_soda_json_progress
        # going within high level folders
        # add manifest details if manifest exists
        manifest_object = {
            "filename": "",
            "timestamp": "",
            "description": "",
            "additional-metadata": "",
        }

        lastSlash = folder_path.rfind("/") + 1
        folder_name = folder_path[lastSlash:]
        if folder_name in replaced:
            folder_name = (
                folder_path[:lastSlash]
                + dataset_structure[folder_name]["original-name"]
            )
        # finds the last / in the path and that is the folder name
        if folder_path in irregularFolders:
            index_check = irregularFolders.index(folder_path)
            modified_name = replaced[index_check]
            folder_path = irregularFolders[index_check]
        entries = os.listdir(folder_path)
        for entry in entries:
            gevent.sleep(0)
            check_path = folder_path + "/" + entry
            if os.path.isfile(check_path) is True:
                # check manifest to add metadata
                if entry[:1] != ".":
                    create_soda_json_progress += 1
                if entry[:1] != "." and entry[:8] != "manifest":
                    # no hidden files or manifest files included
                    if folder_name in soda_json_structure[
                        "starting-point"
                    ] and (
                        soda_json_structure["starting-point"][folder_name][
                            "path"
                        ]
                        != ""
                    ):
                        # checks if there is a path to a manifest
                        manifest_path = soda_json_structure["starting-point"][
                            folder_name
                        ]["path"]
                        ext_index = manifest_path.rfind(".")
                        extension = manifest_path[ext_index:]
                        if extension == ".xlsx":
                            for key in soda_json_structure["starting-point"][
                                folder_name
                            ]["manifest"]:
                                extra_columns = False
                                if(len(key) > 5):
                                    extra_columns = True
                                    extra_columns_dict = dict(itertools.islice(key.items(), 5, len(key)))
                                # description metadata
                                if key["filename"] == entry:
                                    if key["description"] != "":
                                        manifest_object["description"] = key[
                                            "description"
                                        ]
                                    else:
                                        manifest_object["description"] = ""
                                    # additional metadata
                                    if key["Additional Metadata"] != "":
                                        manifest_object["additional-metadata"] = key[
                                            "Additional Metadata"
                                        ]
                                    else:
                                        manifest_object["additional-metadata"] = ""

                                    if extra_columns:
                                        manifest_object["extra_columns"] = extra_columns_dict

                        elif extension == ".csv":
                            for key in soda_json_structure["starting-point"][
                                folder_name
                            ]["manifest"]:
                                extra_columns = False
                                if len(key) > 5:
                                    extra_columns = True
                                    extra_columns_dict = dict(itertools.islice(key.items(), 5, len(key)))

                                if (
                                    soda_json_structure["starting-point"][
                                        folder_name
                                    ]["manifest"][key]["filename"]
                                    == entry
                                ):
                                    if (
                                        soda_json_structure["starting-point"][
                                            folder_name
                                        ][key]["description"]
                                        != None
                                    ):
                                        manifest_object[
                                            "description"
                                        ] = soda_json_structure["starting-point"][
                                            folder_name
                                        ][
                                            key
                                        ][
                                            "description"
                                        ]
                                    else:
                                        manifest_object["description"] = ""
                                if (
                                    soda_json_structure["starting-point"][
                                        folder_name
                                    ]["manifest"][key]["Additional Metadata"]
                                    != None
                                ):
                                    manifest_object[
                                        "additional-metadata"
                                    ] = soda_json_structure["starting-point"][
                                        folder_name
                                    ][
                                        "manifest"
                                    ][
                                        key
                                    ][
                                        "Additional Metadata"
                                    ]
                                else:
                                    manifest_object["additional-metadata"] = ""

                                if extra_columns:
                                    manifest_object["extra_columns"] = extra_columns_dict
                    # create json
                    if "extra_columns" in manifest_object:
                        dataset_structure["files"][entry] = {
                            "path": check_path,
                            "type": "local",
                            "action": ["existing"],
                            "description": manifest_object["description"],
                            "additional-metadata": manifest_object["additional-metadata"],
                            "extra_columns": manifest_object["extra_columns"]
                        }
                    else:
                        dataset_structure["files"][entry] = {
                            "path": check_path,
                            "type": "local",
                            "action": ["existing"],
                            "description": manifest_object["description"],
                            "additional-metadata": manifest_object["additional-metadata"],
                        }
            elif os.path.isdir(check_path) is True:
                create_soda_json_progress += 1
                if check_path in irregularFolders:
                    index_check = irregularFolders.index(check_path)
                    modified_name = replaced[index_check]

                    dataset_structure["folders"][modified_name] = {
                        "folders": {},
                        "files": {},
                        "path": check_path,
                        "type": "local",
                        "action": ["existing"],
                        "original-name": entry,
                    }
                    for folder in dataset_structure["folders"][modified_name][
                        "folders"
                    ]:
                        updated_path = dataset_structure["folders"][modified_name][
                            folder
                        ]["path"]
                        recursive_structure_create(
                            dataset_structure["folders"][modified_name][folder],
                            updated_path,
                        )
                else:
                    dataset_structure["folders"][entry] = {
                        "folders": {},
                        "files": {},
                        "path": check_path,
                        "type": "local",
                        "action": ["existing"],
                    }

        for folder in dataset_structure["folders"]:
            updated_path = dataset_structure["folders"][folder]["path"]
            recursive_structure_create(
                dataset_structure["folders"][folder], updated_path
            )

    # BEGIN

    # count the amount of items in folder
    create_soda_json_total_items = 0
    for root, dirs, filenames in os.walk(root_folder_path):
        # walk through all folders and it's subfolders
        for Dir in dirs:
            # does not take hidden folders or manifest folders
            if Dir[:1] != "." and Dir[:8] != "manifest":
                create_soda_json_total_items += 1
        for fileName in filenames:
            # goes through all files and does not count hidden files
            if fileName[:1] != ".":
                create_soda_json_total_items += 1

    # reading high level folders
    create_soda_json_completed = 0
    create_soda_json_progress = 0
    entries = os.listdir(root_folder_path)

    for entry in entries:
        # begin going through high level folders
        gevent.sleep(0)
        item_path = root_folder_path + "/" + entry
        # high level folder paths
        if os.path.isfile(item_path) is True:
            if entry[:1] != ".":
                # is not a hidden folder
                create_soda_json_progress += 1
                soda_json_structure["metadata-files"][entry] = {
                    "path": item_path,
                    "type": "local",
                    "action": ["existing"],
                }
                    # do file work here
        elif os.path.isdir(item_path) is True:
            create_soda_json_progress += 1
            # add item to soda
            if item_path in irregularFolders:
                index_check = irregularFolders.index(item_path)
                modified_name = replaced[index_check]
                folder_name = modified_name
                dataset_folder["folders"][folder_name] = {
                    "folders": {},
                    "files": {},
                    "path": item_path,
                    "type": "local",
                    "action": ["existing"],
                    "original-basename": item_path[(item_path.rfind("/") + 1) :],
                }
            else:
                if entry in high_level_sparc_folders:
                    dataset_folder["folders"][entry] = {
                        "folders": {},
                        "files": {},
                        "path": item_path,
                        "type": "local",
                        "action": ["existing"],
                    }
            soda_json_structure["starting-point"][entry] = {"path": ""}

    for folder in dataset_folder["folders"]:
        # go through high level folders again
        high_lvl_path = root_folder_path + "/" + folder
        temp_csv = high_lvl_path + "/manifest.csv"
        temp_xlsx = high_lvl_path + "/manifest.xlsx"
        if os.path.exists(temp_csv) == True:
            temp_file_path_csv = root_folder_path + "/" + folder + "/" + "manifest.csv"
            csv_data = pd.read_csv(temp_file_path_csv)
            csv_data.fillna("", inplace=True)
            json_format = csv_data.to_dict(orient="records")
            soda_json_structure["starting-point"][folder]["path"] = temp_file_path_csv
            soda_json_structure["starting-point"][folder]["manifest"] = json_format
        if os.path.exists(temp_xlsx) == True:
            temp_file_path_xlsx = (
                root_folder_path + "/" + folder + "/" + "manifest.xlsx"
            )
            excel_data = pd.read_excel(temp_file_path_xlsx, sheet_name="Sheet1")
            excel_data.fillna("", inplace=True)
            json_format = excel_data.to_dict(orient="records")
            soda_json_structure["starting-point"][folder]["path"] = temp_file_path_xlsx
            soda_json_structure["starting-point"][folder]["manifest"] = json_format
        recursive_structure_create(dataset_folder["folders"][folder], high_lvl_path)

    create_soda_json_completed = 1
    return soda_json_structure


def monitor_local_json_progress():
    """
    Function for monitoring progress of json_object_creation
    Used for progress bar
    """
    global create_soda_json_completed
    global create_soda_json_total_items
    global create_soda_json_progress
    progress_percentage = (
        create_soda_json_progress / create_soda_json_total_items
    ) * 100

    return {
        "create_soda_json_progress": create_soda_json_progress,
        "create_soda_json_total_items": create_soda_json_total_items,
        "progress_percentage": progress_percentage,
        "create_soda_json_completed": create_soda_json_completed
    }




def import_pennsieve_dataset(soda_json_structure, requested_sparc_only=True):
    global namespace_logger
    high_level_sparc_folders = [
        "code",
        "derivative",
        "docs",
        "primary",
        "protocol",
        "source",
    ]
    manifest_sparc = ["manifest.xlsx", "manifest.csv"]
    high_level_metadata_sparc = [
        "submission.xlsx",
        "submission.csv",
        "submission.json",
        "dataset_description.xlsx",
        "dataset_description.csv",
        "dataset_description.json",
        "subjects.xlsx",
        "subjects.csv",
        "subjects.json",
        "samples.xlsx",
        "samples.csv",
        "samples.json",
        "README.txt",
        "CHANGES.txt",
        "code_description.xlsx",
        "inputs_metadata.xlsx",
        "outputs_metadata.xlsx",
    ]

    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]

    global create_soda_json_completed
    global create_soda_json_total_items
    global create_soda_json_progress
    create_soda_json_progress = 0
    create_soda_json_total_items = 0
    create_soda_json_completed = 0

    def verify_file_name(file_name, extension):
        global namespace_logger
        if extension == "":
            return file_name

        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break

        extension_from_name = ""

        if double_ext == False:
            extension_from_name = os.path.splitext(file_name)[1]
        else:
            extension_from_name = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )

        if extension_from_name == ("." + extension):
            return file_name
        else:
            return file_name + ("." + extension)


    def createFolderStructure(subfolder_json, pennsieve_client_or_token, manifest):
        """
            Function for creating the Pennsieve folder structure for a given dataset as an object stored locally.
            Arguments:
                subfolder_json: The json object containing the folder structure of the dataset
                pennsieve_client: The Pennsieve client object
                manifest: The manifest object for the dataset
        """
        # root level folder will pass subfolders into this function and will recursively check if there are subfolders while creating the json structure
        global namespace_logger
        global create_soda_json_progress

        collection_id = subfolder_json["path"]

        headers = create_request_headers(pennsieve_client_or_token)

        r = requests.get(f"{PENNSIEVE_URL}/packages/{collection_id}", headers=headers)
        r.raise_for_status()
        subfolder = r.json()

        children_content = subfolder["children"]
        for items in children_content:
            item_name = items["content"]["name"]
            create_soda_json_progress += 1
            item_id = items["content"]["id"]
            if item_id[2:9] == "package":
                # if it is a file name check if there are additional manifest information to attach to files
                if item_name[:8] != "manifest":  # manifest files are not being included json structure

                    #verify file name first
                    if("extension" not in children_content):
                        item_name = verify_file_name(item_name, "")
                    else:
                        item_name = verify_file_name(item_name, children_content["extension"])

                    ## verify timestamps
                    timestamp = items["content"]["createdAt"]
                    formatted_timestamp = timestamp.replace('.', ',')
                    subfolder_json["files"][item_name] = {
                        "action": ["existing"],
                        "path": item_id,
                        "bfpath": [],
                        "timestamp": formatted_timestamp,
                        "type": "bf",
                        "additional-metadata": "",
                        "description": "",
                    }
                    for paths in subfolder_json["bfpath"]:
                        subfolder_json["files"][item_name]["bfpath"].append(paths)


                    # creates path for item_name (stored in temp_name)
                    if len(subfolder_json["files"][item_name]["bfpath"]) > 1:
                        temp_name = ""
                        for i in range(
                            len(subfolder_json["files"][item_name]["bfpath"])
                        ):
                            if i == 0:
                                continue
                            temp_name += (
                                subfolder_json["files"][item_name]["bfpath"][i] + "/"
                            )
                        temp_name += item_name
                    else:
                        temp_name = item_name

                    if len(manifest.keys()) > 0:
                        extra_columns = False
                        if len(manifest.keys()) > 5:
                            # extra columns are in the manifest
                            # if length of keys is greater than 5 than extra custom columns were made
                            extra_columns = True
                            extra_columns_dict = dict(itertools.islice(manifest.items(), 5, len(manifest)))
                        if "filename" in manifest:
                            if temp_name in manifest["filename"].values():
                                location_index = list(manifest["filename"].values()).index(
                                    temp_name
                                )
                                if manifest["description"][location_index] != "":
                                    subfolder_json["files"][item_name][
                                        "description"
                                    ] = manifest["description"][location_index]
                                if manifest["Additional Metadata"] != "":
                                    subfolder_json["files"][item_name][
                                        "additional-metadata"
                                    ] = manifest["Additional Metadata"][location_index]
                                if manifest["file type"][location_index] != "":
                                        subfolder_json["files"][item_name]["file type"] = manifest["file type"][location_index]
                                if extra_columns:
                                    subfolder_json["files"][item_name]["extra_columns"] = {}
                                    starting_manifest_index = 5
                                    for index in extra_columns_dict.keys():
                                        subfolder_json["files"][item_name]["extra_columns"][index] = (manifest[index][location_index])
                                        starting_manifest_index += 1
                        elif "File Name" in manifest:
                            if temp_name in manifest["File Name"].values():
                                location_index = list(manifest["File Name"].values()).index(
                                    temp_name
                                )
                                if manifest["description"][location_index] != "":
                                    subfolder_json["files"][item_name][
                                        "description"
                                    ] = manifest["description"][location_index]
                                if manifest["Additional Metadata"] != "":
                                    subfolder_json["files"][item_name][
                                        "additional-metadata"
                                    ] = manifest["Additional Metadata"][location_index]
                                if manifest["file type"][location_index] != "":
                                        subfolder_json["files"][item_name]["file type"] = manifest["file type"][location_index]
                            # extra columns are in the manifest


            else:  # another subfolder found
                subfolder_json["folders"][item_name] = {
                    "action": ["existing"],
                    "path": item_id,
                    "bfpath": [],
                    "files": {},
                    "folders": {},
                    "type": "bf",
                }
                for paths in subfolder_json["bfpath"]:
                    subfolder_json["folders"][item_name]["bfpath"].append(paths)
                subfolder_json["folders"][item_name]["bfpath"].append(item_name)

                # go through recursive again through subfolder

        if len(subfolder_json["folders"].keys()) != 0:  # there are subfolders
            for folder in subfolder_json["folders"].keys():
                subfolder = subfolder_json["folders"][folder]
                createFolderStructure(subfolder, pennsieve_client_or_token, manifest)

    # START

    error = []

    # check that the Pennsieve account is valid
    try:
        bf_account_name = soda_json_structure["bf-account-selected"]["account-name"]
    except Exception as e:
        raise e 

    token = get_access_token()

    # check that the Pennsieve dataset is valid
    try:
        bf_dataset_name = soda_json_structure["bf-dataset-selected"]["dataset-name"]
    except Exception as e:
        raise e

    selected_dataset_id = get_dataset_id(token, bf_dataset_name)


    # check that the user has permission to edit this dataset
    try:
        role = bf_get_current_user_permission_agent_two(selected_dataset_id, token)["role"]
        if role not in ["owner", "manager", "editor"]:
            curatestatus = "Done"
            raise Exception("You don't have permissions for uploading to this Pennsieve dataset")
    except Exception as e:
        abort(401, "You do not have permissions to edit upload this Pennsieve dataset.")


    # surface layer of dataset is pulled. then go through through the children to get information on subfolders
    manifest_dict = {}
    manifest_error_message = []
    soda_json_structure["dataset-structure"] = {
        "files": {},
        "folders": {},
    }


    # headers for making requests to Pennsieve's api
    headers = create_request_headers(token)


    # root of dataset is pulled here
    # root_children is the files and folders within root
    r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}", headers=headers)
    r.raise_for_status()
    root_folder = r.json()

    # root's packages 
    r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}/packageTypeCounts", headers=headers)
    r.raise_for_status()
    packages_list = r.json()


    # root's children files
    for count in packages_list.values():
        create_soda_json_total_items += int(count)
    root_children = root_folder["children"]

    for items in root_children:
        item_id = items["content"]["id"]
        item_name = items["content"]["name"]
        if items["content"]["packageType"] == "Collection" and item_name in high_level_sparc_folders:
            create_soda_json_progress += 1
            # is a SPARC folder and will be checked recursively
            soda_json_structure["dataset-structure"]["folders"][item_name] = {
                "type": "bf",
                "path": item_id,
                "action": ["existing"],
                "files": {},
                "folders": {},
                "bfpath": [item_name],
            }
        else:
            if item_name in high_level_metadata_sparc:
                create_soda_json_progress += 1
                # is a metadata file
                if "metadata-files" in soda_json_structure.keys():
                    soda_json_structure["metadata-files"][item_name] = {
                        "type": "bf",
                        "action": ["existing"],
                        "path": item_id,
                    }


    # manifest information is needed so it is looked for before the recursive calls are made
    if len(soda_json_structure["dataset-structure"]["folders"].keys()) != 0:
        for folder in soda_json_structure["dataset-structure"]["folders"].keys():
            collection_id = soda_json_structure["dataset-structure"]["folders"][folder][
                "path"
            ]

            r = requests.get(f"{PENNSIEVE_URL}/packages/{collection_id}", headers=headers)
            r.raise_for_status()
            subfolder = r.json()

            children_content = subfolder["children"]
            manifest_dict[folder] = {}
            if len(children_content) > 0:
                for items in children_content:
                    # check subfolders surface to see if manifest files exist to then use within recursive_subfolder_check
                    package_name = items["content"]["name"]
                    package_id = items["content"]["id"]
                    if package_name in manifest_sparc:
                        # item is manifest
                        r = requests.get(f"{PENNSIEVE_URL}/packages/{package_id}/view", headers=headers)
                        r.raise_for_status()
                        file_details = r.json()

                        file_id = file_details[0]["content"]["id"]
                        r = requests.get(f"{PENNSIEVE_URL}/packages/{package_id}/files/{file_id}", headers=headers)
                        r.raise_for_status()
                        manifest_url = r.json()["url"]

                        df = ""
                        try:                            
                            if package_name.lower() == "manifest.xlsx":
                                df = load_manifest_to_dataframe(package_id, "excel", token)
                                df = df.fillna("")
                            else:
                                df = load_manifest_to_dataframe(package_id, "csv", token)
                                df = df.fillna("")
                            # 
                            manifest_dict[folder].update(df.to_dict())
                        except Exception as e:
                            namespace_logger.info(f"Error reading manifest file: {e}")
                            manifest_error_message.append(
                                items["content"]["name"]
                            )

                subfolder_section = soda_json_structure["dataset-structure"]["folders"][
                    folder
                ]

                if folder in manifest_dict:
                    createFolderStructure(
                        subfolder_section, token, manifest_dict[folder]
                    )  # passing item's json and the collection ID

    success_message = (
        "Data files under a valid high-level SPARC folders have been imported"
    )
    create_soda_json_completed = 1


    return {
        "soda_object": soda_json_structure,
        "success_message": success_message,
        "manifest_error_message": manifest_error_message,
        "import_progress": create_soda_json_progress,
        "import_total_items": create_soda_json_total_items,
    }

def monitor_pennsieve_json_progress():
    """
    Function for monitoring progress of soda_json object
    Used for progress bar
    """
    global create_soda_json_completed
    global create_soda_json_total_items
    global create_soda_json_progress

    if create_soda_json_progress != 0:
        progress_percentage = (
            create_soda_json_progress / create_soda_json_total_items
        ) * 100
    else:
        progress_percentage = 0
    return {
        "import_progress": create_soda_json_progress,
        "import_total_items": create_soda_json_total_items,
        "import_progress_percentage": progress_percentage,
        "import_completed_items": create_soda_json_completed,
    }
